"""Template-manipulation utilities for template-bootstrap operations.

Most operations here are openpyxl-based (fast, no Excel needed):
  - strip orphan named ranges (S&P CapIQ junk leftover from add-ins)
  - remove specific defined names by exact name
  - remove hidden / vestigial sheets
  - add named ranges programmatically (for the engine cell map)

ONE operation requires Excel: `recalc_and_save()`. openpyxl strips cached
formula values when it saves. For workbooks with circular references resolved
via iterative calculation (common in LBO models where debt quantum depends on
fees that depend on debt quantum), the empty cached values force Excel to do
a cold-start recalc on first open, which can fail to converge or flag the
circular ref before iteration kicks in. After any openpyxl edit, call
`recalc_and_save()` to have Excel itself compute and write back the cached
values — the resulting file opens identically to one saved by Excel.

These are one-time operations producing a versioned-up template that operator
reviews before it becomes canonical — NOT used per-run.
"""
from __future__ import annotations

import hashlib
import re
from collections import Counter
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.workbook.defined_name import DefinedName


# Patterns for the S&P Capital IQ Pro add-in junk that leaks into Name Manager.
CAPIQ_NAME_PATTERNS = (
    re.compile(r"^IQ_"),
    re.compile(r"^_xleta\."),
    re.compile(r"^CIQWB"),
)


def hash_file(path: Path) -> str:
    h = hashlib.sha256()
    with path.open("rb") as f:
        for chunk in iter(lambda: f.read(65536), b""):
            h.update(chunk)
    return f"sha256:{h.hexdigest()}"


def _is_capiq_junk(name: str) -> bool:
    return any(p.match(name) for p in CAPIQ_NAME_PATTERNS)


def strip_capiq_names(src: Path, dst: Path) -> dict[str, int]:
    """Copy src → dst with CapIQ-pattern named ranges removed.

    Returns counts: {"removed": N, "kept": M}.
    """
    wb = load_workbook(src, keep_vba=False)
    dn = wb.defined_names
    removed = 0
    kept = 0
    # We must collect names first; mutating during iteration is unsafe.
    for name in list(dn):
        if _is_capiq_junk(name):
            del dn[name]
            removed += 1
        else:
            kept += 1
    wb.save(dst)
    return {"removed": removed, "kept": kept}


def remove_defined_names(path: Path, names: list[str]) -> dict[str, list[str]]:
    """Remove specific workbook-scope defined names. Operates in place.

    Returns {"removed": [...], "not_found": [...]}.
    """
    wb = load_workbook(path, keep_vba=False)
    dn = wb.defined_names
    removed: list[str] = []
    not_found: list[str] = []
    for name in names:
        if name in dn:
            del dn[name]
            removed.append(name)
        else:
            not_found.append(name)
    wb.save(path)
    return {"removed": removed, "not_found": not_found}


def remove_sheet(path: Path, sheet_name: str) -> bool:
    """Delete a sheet. Returns True if removed, False if not present.

    Refuses to delete the only visible sheet — Excel files must keep at least
    one visible sheet.
    """
    wb = load_workbook(path, keep_vba=False)
    if sheet_name not in wb.sheetnames:
        return False
    visible = [s for s in wb.sheetnames if wb[s].sheet_state == "visible"]
    if len(visible) == 1 and visible[0] == sheet_name:
        raise ValueError(f"Refusing to delete {sheet_name!r} — would leave no visible sheets")
    del wb[sheet_name]
    wb.save(path)
    return True


def add_named_range(path: Path, name: str, sheet: str, addr: str, *, overwrite: bool = False) -> None:
    """Add a workbook-scope named range pointing at `sheet!addr`.

    Args:
        name: workbook-scope defined name (no spaces, no leading digit).
        sheet: sheet name (no quotes; we'll quote if needed).
        addr: A1-form address (e.g. "I25" or "B5:F15"), $-prefix optional.
        overwrite: if True, replace an existing name; if False, error.
    """
    if not _VALID_NAME.match(name):
        raise ValueError(f"Invalid defined-name {name!r}: must start with letter/underscore, no spaces")
    wb = load_workbook(path, keep_vba=False)
    if sheet not in wb.sheetnames:
        raise ValueError(f"Sheet {sheet!r} not in workbook")
    if name in wb.defined_names:
        if not overwrite:
            raise ValueError(f"Defined name {name!r} already exists (pass overwrite=True to replace)")
        del wb.defined_names[name]

    qsheet = f"'{sheet}'" if (" " in sheet or "&" in sheet) else sheet
    # Add absolute $ to bare A1 refs for stability against row inserts above
    abs_addr = _absolute(addr)
    refers = f"{qsheet}!{abs_addr}"
    wb.defined_names[name] = DefinedName(name=name, attr_text=refers)
    wb.save(path)


def add_named_ranges_bulk(path: Path, ranges: list[tuple[str, str, str]], *, overwrite: bool = False) -> dict[str, list[str]]:
    """Add multiple named ranges in one pass.

    `ranges` is a list of (name, sheet, addr) tuples.
    Returns {"added": [...], "skipped": [...]}.
    """
    if not all(_VALID_NAME.match(n) for n, _, _ in ranges):
        bad = [n for n, _, _ in ranges if not _VALID_NAME.match(n)]
        raise ValueError(f"Invalid defined names: {bad}")

    wb = load_workbook(path, keep_vba=False)
    added: list[str] = []
    skipped: list[str] = []
    for name, sheet, addr in ranges:
        if sheet not in wb.sheetnames:
            raise ValueError(f"Sheet {sheet!r} not in workbook (for name {name!r})")
        if name in wb.defined_names:
            if not overwrite:
                skipped.append(name)
                continue
            del wb.defined_names[name]
        qsheet = f"'{sheet}'" if (" " in sheet or "&" in sheet) else sheet
        wb.defined_names[name] = DefinedName(name=name, attr_text=f"{qsheet}!{_absolute(addr)}")
        added.append(name)
    wb.save(path)
    return {"added": added, "skipped": skipped}


def list_defined_names(path: Path) -> dict[str, list[str]]:
    """Categorise defined names: junk (CapIQ patterns) vs user."""
    wb = load_workbook(path, read_only=True)
    junk: list[str] = []
    user: list[str] = []
    for name in list(wb.defined_names):
        (junk if _is_capiq_junk(name) else user).append(name)
    return {"junk": sorted(junk), "user": sorted(user)}


# ============================================================================
# Pure-xlwings template surgery
# ----------------------------------------------------------------------------
# openpyxl's save() drops cached formula values AND the calcChain XML part.
# For workbooks with iterative-calc circular references (LBO debt/fees,
# NWC self-references) this leaves Excel unable to bootstrap the recalc:
# even with iteration enabled and a full CalculateFullRebuild, certain cells
# in the cycle never get evaluated. Discovered 2026-05-19 during LBO v2 build.
#
# The fix: do all template surgery via Excel COM (xlwings). The workbook
# stays Excel-native throughout, so the calc chain and cached values are
# preserved exactly as if a human did the edits in the Excel UI.
# ============================================================================

def build_template_v2(
    src: Path,
    dst: Path,
    *,
    delete_name_patterns: tuple = (),
    delete_names_exact: list[str] | None = None,
    remove_sheets: list[str] | None = None,
    add_named_ranges: list[tuple[str, str, str]] | None = None,
    recalc_iterative_max: int = 500,
    recalc_iterative_tol: float = 0.0001,
) -> dict:
    """Build a cleaned + named-range-augmented copy of `src` at `dst`, all via Excel COM.

    Operations (in order):
      1. Copy src → dst.
      2. Open dst in Excel.
      3. Delete every defined name matching any regex in `delete_name_patterns`.
      4. Delete every defined name in `delete_names_exact`.
      5. Delete every sheet named in `remove_sheets` (unless it's the last visible).
      6. Add every (name, sheet, addr) in `add_named_ranges`.
      7. Force-enable iteration at the app level (if workbook wants it).
      8. CalculateFullRebuild + double calculate().
      9. Save.

    Returns a summary dict of what was removed / added / recomputed.
    """
    import re as _re
    import shutil

    try:
        import xlwings as xw  # type: ignore
    except ImportError as e:
        raise RuntimeError(
            "xlwings not installed — build_template_v2 requires Windows + Excel."
        ) from e

    delete_names_exact = list(delete_names_exact or [])
    remove_sheets = list(remove_sheets or [])
    add_named_ranges = list(add_named_ranges or [])
    compiled_patterns = [_re.compile(p) if isinstance(p, str) else p for p in delete_name_patterns]

    # Read workbook iteration intent from openpyxl (no risk — we don't save)
    from openpyxl import load_workbook
    wb_meta = load_workbook(src, read_only=True)
    workbook_wants_iteration = bool(wb_meta.calculation.iterate)
    try:
        wb_meta.close()
    except Exception:
        pass

    # Stage destination
    if dst.exists():
        dst.unlink()
    dst.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy2(src, dst)

    summary: dict[str, object] = {
        "names_removed_by_pattern": 0,
        "names_removed_exact": [],
        "names_exact_not_found": [],
        "sheets_removed": [],
        "sheets_not_found": [],
        "named_ranges_added": [],
        "named_ranges_skipped_existing": [],
        "iterative_calc_used": workbook_wants_iteration,
    }

    app = None
    wb = None
    prior_iter = None
    prior_max_iter = None
    prior_max_change = None

    try:
        app = xw.App(visible=False, add_book=False)
        try:
            app.api.DisplayAlerts = False
            app.api.AskToUpdateLinks = False
            app.api.ScreenUpdating = False
        except Exception:
            pass

        wb = app.books.open(str(dst), update_links=False, read_only=False)

        # --- 3. Delete defined names by regex pattern (workbook-scope + sheet-scope) ---
        # Collect first, delete second — deleting from a COM collection while
        # iterating breaks indexing. We strip from both `Workbook.Names` (workbook-
        # scope) and each `Worksheet.Names` (sheet-scope); S&P CapIQ Pro leaks names
        # at both scopes (e.g. `1PageOutput!CIQWBGuid`).
        wb_to_delete: list[str] = []
        for i in range(1, wb.api.Names.Count + 1):
            try:
                n = wb.api.Names.Item(i).Name
            except Exception:
                continue
            if any(p.match(n) for p in compiled_patterns):
                wb_to_delete.append(n)
        for n in wb_to_delete:
            try:
                wb.api.Names.Item(n).Delete()
            except Exception:
                pass

        # Per-sheet scope. Iterate every sheet's local Names collection.
        # Note: a sheet-scope name's `.Name` property returns the QUALIFIED form
        # ("SheetName!localname") via Workbook.Names, but the UNQUALIFIED form
        # ("localname") via Worksheet.Names.Item(i).Name. We match on the
        # unqualified form.
        sheet_to_delete: list[tuple[str, str]] = []
        for sheet in wb.sheets:
            try:
                sheet_names = sheet.api.Names
                count = sheet_names.Count
            except Exception:
                continue
            for i in range(1, count + 1):
                try:
                    n = sheet_names.Item(i).Name
                    # Strip any "SheetName!" prefix that some Excel versions report
                    local = n.split("!", 1)[1] if "!" in n else n
                except Exception:
                    continue
                if any(p.match(local) for p in compiled_patterns):
                    sheet_to_delete.append((sheet.name, local))
        for sheet_name, local in sheet_to_delete:
            try:
                wb.sheets[sheet_name].api.Names.Item(local).Delete()
            except Exception:
                pass

        summary["names_removed_by_pattern"] = len(wb_to_delete) + len(sheet_to_delete)
        summary["names_removed_workbook_scope"] = len(wb_to_delete)
        summary["names_removed_sheet_scope"] = len(sheet_to_delete)

        # Some defined names — notably Excel's `_xleta.*` LAMBDA sentinels —
        # have dots in their name and Excel's own COM Delete() refuses them
        # with a "syntax not correct" error. Collect any that matched the
        # patterns but survived; we'll surgically remove them via XML edit
        # AFTER closing the workbook below.
        wb_xml_purge_targets: list[str] = []
        try:
            remaining_names = set()
            for i in range(1, wb.api.Names.Count + 1):
                try:
                    remaining_names.add(wb.api.Names.Item(i).Name)
                except Exception:
                    pass
            for n in remaining_names:
                local = n.split("!", 1)[-1]
                if any(p.match(local) for p in compiled_patterns):
                    wb_xml_purge_targets.append(n)
        except Exception:
            pass
        summary["names_pending_xml_purge"] = wb_xml_purge_targets

        # --- 4. Delete specific named ranges ---
        for n in delete_names_exact:
            try:
                wb.api.Names.Item(n).Delete()
                summary["names_removed_exact"].append(n)
            except Exception:
                summary["names_exact_not_found"].append(n)

        # --- 5. Remove sheets ---
        for sheet_name in remove_sheets:
            try:
                # Refuse to delete the last visible sheet
                visible = [s for s in wb.sheets if s.api.Visible == -1]  # xlSheetVisible = -1
                if len(visible) <= 1 and sheet_name in [s.name for s in visible]:
                    raise ValueError(f"Refusing to delete {sheet_name!r} — would leave no visible sheets")
                wb.sheets[sheet_name].delete()
                summary["sheets_removed"].append(sheet_name)
            except KeyError:
                summary["sheets_not_found"].append(sheet_name)

        # --- 6. Add named ranges (workbook scope) ---
        existing_names = set()
        for i in range(1, wb.api.Names.Count + 1):
            try:
                existing_names.add(wb.api.Names.Item(i).Name)
            except Exception:
                pass
        for name, sheet, addr in add_named_ranges:
            if name in existing_names:
                summary["named_ranges_skipped_existing"].append(name)
                continue
            if sheet not in [s.name for s in wb.sheets]:
                raise ValueError(f"Sheet {sheet!r} not in workbook (for name {name!r})")
            qsheet = f"'{sheet}'" if (" " in sheet or "&" in sheet) else sheet
            # Force absolute addressing for stability
            abs_addr = _absolute(addr)
            refers = f"={qsheet}!{abs_addr}"
            wb.api.Names.Add(Name=name, RefersTo=refers)
            summary["named_ranges_added"].append(name)

        # --- 7. Force iteration at the app level (workbook setting only applies on open) ---
        try:
            prior_iter = bool(app.api.Iteration)
            prior_max_iter = int(app.api.MaxIterations)
            prior_max_change = float(app.api.MaxChange)
        except Exception:
            pass

        if workbook_wants_iteration:
            try:
                app.api.Iteration = True
                app.api.MaxIterations = recalc_iterative_max
                app.api.MaxChange = recalc_iterative_tol
            except Exception:
                pass

        # --- 8. Full rebuild + recalc ---
        app.api.CalculateFullRebuild()
        app.calculate()
        app.calculate()

        # --- 9. Save ---
        wb.save()
        summary["recalc_ok"] = True

        # --- 10. Surgical XML purge for dotted names Excel COM refuses to delete ---
        # Must happen AFTER wb.close() so the file isn't locked. We capture the
        # list now and schedule the purge in the finally block once Excel
        # releases the file.
    finally:
        if wb is not None:
            try:
                wb.close()
            except Exception:
                pass
        # Restore the user's prior Excel iteration settings
        if app is not None and prior_iter is not None:
            try:
                app.api.Iteration = prior_iter
                if prior_max_iter is not None:
                    app.api.MaxIterations = prior_max_iter
                if prior_max_change is not None:
                    app.api.MaxChange = prior_max_change
            except Exception:
                pass
        if app is not None:
            try:
                app.quit()
            except Exception:
                try:
                    app.kill()
                except Exception:
                    pass

    # Post-Excel surgical XML purge for names COM refused to delete (dotted names).
    # Runs only if we identified targets in step 7.
    pending = summary.get("names_pending_xml_purge") or []
    if pending:
        xml_result = remove_defined_names_via_xml(dst, list(pending))
        summary["names_purged_via_xml"] = xml_result["removed"]
        summary["names_xml_not_found"] = xml_result["not_found"]
        # Count xml-purged into the headline total
        summary["names_removed_by_pattern"] = (
            int(summary.get("names_removed_by_pattern") or 0) + len(xml_result["removed"])
        )

    return summary


# ====================================================================
# Surgical XML-level removal of stubborn defined names
# --------------------------------------------------------------------
# Excel's COM `Names.Item(...).Delete()` refuses to delete defined names
# whose name contains a dot (e.g. `_xleta.BASE`, `_xleta.N`) — it rejects
# the name's syntax even though Excel itself created and stored it. These
# are Excel's LAMBDA helper sentinels (`_xleta.*`) that survive even when
# no LAMBDA in the workbook actually uses them, and they refer to `=#NAME?`
# (inert but ugly in Name Manager).
#
# Workaround: open the .xlsx as a zip, edit only `xl/workbook.xml` to
# remove the offending `<definedName>` entries, rewrite the zip. Surgical
# — every other zip entry (sheets, calcChain.xml, cached values, etc.) is
# copied through byte-for-byte.
# ====================================================================

def repoint_named_range(path: Path, name: str, sheet: str, addr: str, *, value: object = ...,
                        clear_old: bool = True) -> dict:
    """Repoint a workbook-scope defined name to a new cell.

    Optionally moves the value at the old location to the new one, applies
    orange-input fill at the new cell, and clears the old cell.

    Run via xlwings so calc chain + cached values are preserved.

    Args:
        path: workbook to modify in place
        name: existing workbook-scope defined name
        sheet: target sheet for new location
        addr: target A1 address (e.g. "M50")
        value: if not ...: write this value to the new cell first (use Ellipsis
               as the "no change" sentinel; pass None to clear, 0 to set zero)
        clear_old: if True, find the old cell the name pointed to and clear it
                   (formula AND value)

    Returns: {"old_ref": str, "new_ref": str, "value_at_new": ...}
    """
    try:
        import xlwings as xw  # type: ignore
    except ImportError as e:
        raise RuntimeError("xlwings required for repoint_named_range") from e

    if not path.exists():
        raise FileNotFoundError(path)

    app = None
    wb = None
    result = {"old_ref": None, "new_ref": None, "value_at_new": None}
    try:
        app = xw.App(visible=False, add_book=False)
        try:
            app.api.DisplayAlerts = False
        except Exception:
            pass
        wb = app.books.open(str(path), update_links=False, read_only=False)

        # Resolve the existing name's RefersTo
        try:
            existing = wb.api.Names.Item(name)
            old_refers = existing.RefersTo
            result["old_ref"] = old_refers
        except Exception as e:
            raise RuntimeError(f"Named range {name!r} not found in workbook: {e}") from e

        # Optionally clear the old location (parse "=Sheet!$A$1" form)
        if clear_old:
            try:
                import re as _re
                m = _re.match(r"^=(?:'([^']+)'|([A-Za-z_][A-Za-z0-9_.]*))!(\$?[A-Z]+\$?\d+(?::\$?[A-Z]+\$?\d+)?)$",
                              old_refers)
                if m:
                    old_sheet = m.group(1) or m.group(2)
                    old_addr = m.group(3)
                    try:
                        wb.sheets[old_sheet].range(old_addr).clear_contents()
                    except Exception:
                        pass
            except Exception:
                pass

        # Set value at the new cell (if provided)
        target = wb.sheets[sheet].range(addr)
        if value is not ...:
            target.value = value
        # Capture the value for the return record
        result["value_at_new"] = target.value

        # Repoint the named range
        existing.Delete()  # safe — name doesn't contain a dot
        qsheet = f"'{sheet}'" if (" " in sheet or "&" in sheet) else sheet
        abs_addr = _absolute(addr)
        new_refers = f"={qsheet}!{abs_addr}"
        wb.api.Names.Add(Name=name, RefersTo=new_refers)
        result["new_ref"] = new_refers

        # Re-enable iteration if the workbook wants it (mirrors build_template_v2)
        from openpyxl import load_workbook as _lwb
        meta = _lwb(path, read_only=True)
        if meta.calculation.iterate:
            try:
                app.api.Iteration = True
                app.api.MaxIterations = 500
                app.api.MaxChange = 0.0001
            except Exception:
                pass
        try:
            meta.close()
        except Exception:
            pass

        app.api.CalculateFullRebuild()
        app.calculate(); app.calculate()
        wb.save()
        return result
    finally:
        if wb is not None:
            try:
                wb.close()
            except Exception:
                pass
        if app is not None:
            try:
                app.quit()
            except Exception:
                try:
                    app.kill()
                except Exception:
                    pass


def remove_defined_names_via_xml(path: Path, names: list[str]) -> dict:
    """Remove specific workbook-scope `<definedName>` entries from xl/workbook.xml.

    Bypasses both openpyxl (which strips calcChain on save) and Excel COM
    (which refuses to delete dotted names). Operates in place.

    Args:
        path: .xlsx file to edit.
        names: list of exact name strings to remove. (Workbook-scope only.)

    Returns:
        {"removed": [...], "not_found": [...]}
    """
    import zipfile
    import shutil
    import re as _re
    import tempfile

    if not path.exists():
        raise FileNotFoundError(path)

    target_set = set(names)
    removed: list[str] = []
    not_found: list[str] = list(names)

    # We can't edit a zip in place; write to a sibling tmp then atomic-rename.
    tmp_path = path.with_suffix(path.suffix + ".tmp")
    if tmp_path.exists():
        tmp_path.unlink()

    with zipfile.ZipFile(path, "r") as zin:
        # Sanity: must have workbook.xml
        if "xl/workbook.xml" not in zin.namelist():
            raise ValueError(f"{path} is not a valid xlsx (missing xl/workbook.xml)")

        with zipfile.ZipFile(tmp_path, "w", zipfile.ZIP_DEFLATED, allowZip64=True) as zout:
            for item in zin.infolist():
                if item.filename != "xl/workbook.xml":
                    # Copy through byte-for-byte (preserves calcChain.xml, sheets, etc.)
                    zout.writestr(item, zin.read(item.filename))
                    continue

                # Edit workbook.xml — strip the matching <definedName> entries.
                content = zin.read(item.filename).decode("utf-8")

                # Match <definedName name="X" ...>...</definedName> (with or without attrs)
                # We use a non-greedy match and the name attribute as the anchor.
                def _strip(match):
                    n = match.group("name")
                    if n in target_set:
                        return ""  # drop the whole element
                    return match.group(0)

                pattern = _re.compile(
                    r"<definedName\b[^>]*\bname=\"(?P<name>[^\"]+)\"[^>]*"
                    r"(?:/>|>(?:.*?)</definedName>)",
                    _re.DOTALL,
                )
                before_count = len(_re.findall(pattern, content))
                new_content, _ = _re.subn(pattern, _strip, content)
                after_count = len(_re.findall(pattern, new_content))
                removed_here = before_count - after_count

                # Track which names actually disappeared
                for n in list(not_found):
                    if f'name="{n}"' in content and f'name="{n}"' not in new_content:
                        removed.append(n)
                        not_found.remove(n)

                # Clean up any now-empty <definedNames>...</definedNames> wrapper
                new_content = _re.sub(
                    r"<definedNames>\s*</definedNames>", "", new_content
                )

                # Write the edited workbook.xml — preserve item metadata
                zout.writestr(item, new_content.encode("utf-8"))

    # Atomic replace (rename only — original file is unlinked here)
    path.unlink()
    shutil.move(str(tmp_path), str(path))
    return {"removed": removed, "not_found": not_found}


# --------------------------------------------------------- recalc via xlwings

def recalc_and_save(path: Path, *, keep_excel_open: bool = False, max_iterations: int = 200) -> dict[str, object]:
    """Open the workbook in Excel, force a full recalc (honouring the workbook's
    iterative-calculation settings), save in place.

    Reason: openpyxl strips cached formula values on save. For models with
    circular references resolved via Excel's iterative calc (LBO debt/fees
    loops, NWC self-references, etc.), this leaves the .xlsx in a state that
    requires Excel itself to recompute on first open. If iteration doesn't
    converge cleanly on cold start, Excel may flag the circular ref. Saving
    via Excel populates the cached values so subsequent opens work cleanly.

    Args:
        path: workbook to recalc + save in place.
        keep_excel_open: leave Excel running for inspection (debug only).
        max_iterations: temporarily bump iteration count if iterative calc is
            enabled. None of the workbook's settings are written back.

    Returns:
        Dict with: iterate (bool), iter_count (int), recalc_ok (bool),
        sample_value (the first non-formula numeric cell on the first sheet,
        for sanity).
    """
    try:
        import xlwings as xw  # type: ignore
    except ImportError as e:
        raise RuntimeError(
            "xlwings not installed — recalc_and_save requires Windows + Excel. "
            "Install via `pip install -e .[dev]` from the engine repo."
        ) from e

    # Decide whether iteration should be on. The .xlsx XML carries a per-workbook
    # `calcPr/@iterate` flag, but Excel's recalc honours the *app-level*
    # `Application.Iteration` setting. The Excel default is OFF, so we read the
    # workbook's intent from openpyxl, then force-enable at the app level for
    # the duration of the recalc.
    from openpyxl import load_workbook
    wb_meta = load_workbook(path, read_only=True)
    workbook_wants_iteration = bool(wb_meta.calculation.iterate)
    workbook_iter_count = wb_meta.calculation.iterateCount or 100
    workbook_iter_delta = wb_meta.calculation.iterateDelta or 0.001
    try:
        wb_meta.close()
    except Exception:
        pass

    app = None
    wb = None
    prior_iter = None
    prior_max_iter = None
    prior_max_change = None
    try:
        app = xw.App(visible=False, add_book=False)
        try:
            app.api.DisplayAlerts = False
            app.api.AskToUpdateLinks = False
            app.api.ScreenUpdating = False
        except Exception:
            pass

        # Open the workbook BEFORE touching Iteration — Excel's
        # `Application.Iteration` accessor errors with -2146826246 when no
        # workbook is open. Properties exist but aren't gettable/settable
        # until there's something to apply them to.
        wb = app.books.open(str(path), update_links=False, read_only=False)

        # Now snapshot the user's Excel app-level iteration settings.
        try:
            prior_iter = bool(app.api.Iteration)
            prior_max_iter = int(app.api.MaxIterations)
            prior_max_change = float(app.api.MaxChange)
        except Exception:
            pass

        # If the workbook wants iterative calc, force-enable at the app level.
        # Use the larger of the workbook's `iterateCount` and our `max_iterations`
        # so badly-stuck circular refs still converge.
        if workbook_wants_iteration:
            try:
                app.api.Iteration = True
                app.api.MaxIterations = max(int(workbook_iter_count), int(max_iterations))
                app.api.MaxChange = float(workbook_iter_delta)
            except Exception:
                pass

        # Full rebuild + calc. CalculateFullRebuild() reconstructs the dep graph
        # from scratch and re-iterates all circular refs; the follow-up
        # calculate() ensures cached values are flushed for save.
        wb.app.api.CalculateFullRebuild()
        wb.app.calculate()
        # Belt-and-braces second pass — for deep cycles, the first calculate()
        # can land mid-iteration; a second is cheap and guarantees convergence.
        wb.app.calculate()

        # Pick a sample value for the return record (sanity)
        sample = None
        try:
            first_sheet = wb.sheets[0]
            for row in first_sheet.used_range.value or []:
                if isinstance(row, list):
                    for v in row:
                        if isinstance(v, (int, float)) and v != 0:
                            sample = v
                            break
                if sample is not None:
                    break
        except Exception:
            pass

        wb.save()
        return {
            "iterate": workbook_wants_iteration,
            "iter_count_used": max(int(workbook_iter_count), int(max_iterations)) if workbook_wants_iteration else 0,
            "recalc_ok": True,
            "sample_value": sample,
        }
    finally:
        if wb is not None:
            try:
                wb.close()
            except Exception:
                pass
        # Restore the user's prior Excel app-level iteration settings — we do
        # not want to leave their Excel global preferences mutated.
        if app is not None and prior_iter is not None:
            try:
                app.api.Iteration = prior_iter
                if prior_max_iter is not None:
                    app.api.MaxIterations = prior_max_iter
                if prior_max_change is not None:
                    app.api.MaxChange = prior_max_change
            except Exception:
                pass
        if app is not None and not keep_excel_open:
            try:
                app.quit()
            except Exception:
                try:
                    app.kill()
                except Exception:
                    pass


# ------------------------------------------------------------------ internals

_VALID_NAME = re.compile(r"^[A-Za-z_][A-Za-z0-9_.]*$")
_A1_ABS = re.compile(r"^\$?[A-Z]+\$?\d+(?::\$?[A-Z]+\$?\d+)?$", re.IGNORECASE)


def _absolute(addr: str) -> str:
    """Make A1 / A1:B10 absolute by ensuring `$` on column and row."""
    if not _A1_ABS.match(addr):
        raise ValueError(f"Not a valid A1 address: {addr!r}")
    parts = addr.split(":")
    out: list[str] = []
    for p in parts:
        # Split into col-letters and row-digits
        m = re.match(r"^\$?([A-Z]+)\$?(\d+)$", p, re.IGNORECASE)
        if not m:
            raise ValueError(f"Cannot parse address part {p!r}")
        col, row = m.groups()
        out.append(f"${col.upper()}${row}")
    return ":".join(out)
