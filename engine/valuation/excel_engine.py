"""Core Excel engine — copies a template, populates inputs, triggers recalc,
reads outputs back, runs any post-recalc hardcode steps (with convergence),
validates, and saves the populated workbook.

Uses xlwings (Windows-only) so Excel itself does the recalculation. The
template is *never* modified — every run operates on a fresh copy.

Defensive against the common Excel-COM failure modes:
  - File already open in Excel: bail with a clear error rather than locking.
  - Macros / external-link prompts: suppressed via app.api.AskToUpdateLinks/DisplayAlerts.
  - Dialog popups: visibility off + display_alerts off.
  - Hung Excel: tear-down in a finally block (app.kill() if .quit() doesn't return).
"""
from __future__ import annotations

import logging
import math
import shutil
import sys
import time
import uuid
from datetime import datetime
from pathlib import Path
from typing import Any

from valuation.cell_refs import parse as parse_ref
from valuation.exceptions import (
    CellRefUnresolved,
    ClientFSBlockInvalid,
    ClientFSFormulaCollision,
    ConvergenceFailed,
    ExcelDriverError,
    InputCellMismatch,
    TemplateHashMismatch,
    ValidationFailed,
)
from valuation.models import EngineRun, TemplateSpec

logger = logging.getLogger(__name__)


# --------------------------------------------------------------------- helpers

def _new_output_path(spec: TemplateSpec, output_dir: Path) -> tuple[str, Path]:
    """Default output filename if caller doesn't supply one:
        <template>-<YYYYMMDD-HHMM>-<run_id>.xlsx
    """
    run_id = uuid.uuid4().hex[:8]
    timestamp = datetime.now().strftime("%Y%m%d-%H%M")
    name = f"{spec.name}-{timestamp}-{run_id}.xlsx"
    return run_id, output_dir / name


def _copy_template(src: Path, dst: Path) -> None:
    if not src.exists():
        raise ExcelDriverError(f"Template file not found: {src}")
    dst.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy2(src, dst)


def _verify_hash(spec: TemplateSpec) -> None:
    """Re-hash live file; raise if different from what was registered."""
    from valuation.registry import TemplateRegistry
    live = TemplateRegistry._hash_file(spec.path)
    if live != spec.template_hash:
        raise TemplateHashMismatch(
            f"Template {spec.name!r} hash drift detected.\n"
            f"  registered: {spec.template_hash}\n"
            f"  live:       {live}\n"
            f"Either re-run `engine validate {spec.name} --rehash` after confirming "
            f"the cell map still resolves correctly, or revert the .xlsx edit."
        )


def _check_input_keys(spec: TemplateSpec, given: dict[str, Any]) -> None:
    """Required inputs must all be present; unknowns must match optional_inputs or be rejected."""
    required = set(spec.inputs)
    given_keys = set(given)
    optional = set(spec.optional_inputs)

    missing = required - given_keys
    extra = given_keys - required - optional

    if missing or extra:
        msg = [f"Input mismatch for template {spec.name!r}."]
        if missing:
            msg.append(f"  Missing required: {sorted(missing)}")
        if extra:
            msg.append(f"  Unexpected (not in inputs or optional_inputs): {sorted(extra)}")
        raise InputCellMismatch("\n".join(msg))


# ------------------------------------------------------------------ xlwings IO

def _resolve_range(wb, ref: str):
    """Resolve a cell-map ref to an xlwings Range object.

    Tries: named range first (workbook scope, then sheet scope), then A1.
    """
    parsed = parse_ref(ref)
    # Named range?
    if parsed.is_named:
        # Workbook scope
        for name in wb.names:
            if name.name == parsed.addr or name.name == f"{parsed.sheet}!{parsed.addr}":
                return name.refers_to_range
        # Sheet scope (rare; xlwings exposes via sheet.names)
        if parsed.sheet:
            try:
                sheet = wb.sheets[parsed.sheet]
                for name in sheet.names:
                    if name.name == parsed.addr:
                        return name.refers_to_range
            except Exception:
                pass
        raise CellRefUnresolved(f"Named range not found: {ref}")
    # A1 form
    if not parsed.sheet:
        raise CellRefUnresolved(f"A1 ref must be sheet-qualified: {ref!r}")
    try:
        return wb.sheets[parsed.sheet].range(parsed.addr)
    except Exception as e:
        raise CellRefUnresolved(f"Could not resolve {ref!r}: {e}") from e


def _write_cell(wb, ref: str, value: Any) -> None:
    rng = _resolve_range(wb, ref)
    rng.value = value


def _read_cell(wb, ref: str) -> Any:
    rng = _resolve_range(wb, ref)
    return rng.value


# ----------------------------------------------------------- validation rules

def _evaluate_validation(outputs: dict[str, Any], rules: list[dict[str, Any]]) -> list[str]:
    """Evaluate validation rules against the outputs dict.

    Each rule: {"rule": "outputs.enterprise_value > 0", "message": "..."}
    Returns the list of failed-rule messages (empty if all passed).

    Note: `isinstance` is exposed so rules can guard against non-numeric
    outputs (Excel errors come through as None or '#N/A'/'#REF!' strings).
    """
    failures: list[str] = []
    for rule_spec in rules:
        expr = rule_spec.get("rule")
        if not expr:
            continue
        # Evaluate in a tightly-scoped namespace: only `outputs` and builtins.
        scope = {
            "outputs": _AttrDict(outputs),
            "abs": abs, "min": min, "max": max,
            "isinstance": isinstance, "int": int, "float": float, "str": str,
            "len": len, "all": all, "any": any,
        }
        try:
            ok = bool(eval(expr, {"__builtins__": {}}, scope))
        except Exception as e:
            failures.append(f"rule {expr!r} raised: {e}")
            continue
        if not ok:
            msg = rule_spec.get("message", f"failed: {expr}")
            failures.append(msg)
    return failures


# ----------------------------------------------- F-14: rejected-workbook quarantine

def _rejected_output_path(out_path: Path, run_id: str) -> Path:
    """Quarantine path for a validation-FAILED workbook.

    ``deal_v3.xlsx`` → ``deal_v3.REJECTED.<run_id>.xlsx``. The ``.REJECTED``
    infix sits OUTSIDE the ``vN`` deliverable naming pattern so a rejected
    model can never be mistaken for a real output (Iron Law / F-14); the
    ``run_id`` suffix keeps successive rejections from overwriting one another
    and from colliding with any pre-existing file (codex-5.5 F-14 round 1)."""
    return out_path.with_name(f"{out_path.stem}.REJECTED.{run_id}{out_path.suffix}")


def _quarantine_rejected_workbook(
    wb, out_path: Path, run_id: str, notes: list[str],
) -> None:
    """Persist a validation-failed workbook to its ``*.REJECTED.<run_id>.xlsx``
    quarantine path instead of the live ``vN`` deliverable path (F-14).

    Before this, ``run()`` called ``wb.save()`` (writing to the live ``vN``
    path) on the failure branch BEFORE raising — so a rejected, unsourced model
    was archived as a deliverable. Now the populated-but-invalid workbook is
    saved to the quarantine path (kept for debugging WHY it failed) and the
    caller unlinks the live ``out_path`` so no deliverable-looking artifact
    survives. Best-effort: a quarantine-save failure is noted, never masks the
    ``ValidationFailed`` the caller is about to raise."""
    rejected = _rejected_output_path(out_path, run_id)
    try:
        wb.save(str(rejected))
        notes.append(
            f"validation failed — rejected workbook quarantined at {rejected.name} "
            f"(NOT saved to the deliverable path)"
        )
    except Exception as e:  # noqa: BLE001 — quarantine is best-effort
        notes.append(f"validation failed — could NOT quarantine rejected workbook: {e}")


class _AttrDict(dict):
    """dict that allows attribute access for validation-rule eval."""
    def __getattr__(self, key: str) -> Any:
        try:
            return self[key]
        except KeyError as e:
            raise AttributeError(str(e))


# ------------------------------------------------------- Client_FS pre-write

# Layout contract for the generalized Client_FS operating-model sheet (the
# template's client-financials slot). The period window is FIXED at 10 columns
# J:S; row 4 carries the period-end dates that OpModel_Link joins (SUMIFS)
# against the LBO timeline (the EDATE chain seeded by First_FYE). Partial
# windows are refused — stale values left in unwritten columns would silently
# poison the date-join.
CLIENT_FS_DEFAULT_SHEET = "Client_FS"
CLIENT_FS_DATE_ROW = 4
CLIENT_FS_FIRST_COL = "J"
CLIENT_FS_PERIODS = 10                   # J:S inclusive

# Excel COM calculation modes. Manual is used as a write window so bulk
# Client_FS row writes don't trigger per-write recalcs of the template's
# iterative/circular graph. NEVER set xlCalculationAutomatic (-4105) on this
# template — the 2-var IRR data table then recalcs continuously and Excel
# hangs. The engine only ever RESTORES the workbook's native mode.
XL_CALCULATION_MANUAL = -4135

_CLIENT_FS_ALLOWED_KEYS = {"dates", "rows", "zero_rows", "sheet"}


def _normalize_client_fs(block: dict[str, Any]) -> dict[str, Any]:
    """Validate + normalize a caller-supplied `client_fs` block.

    Expected shape (JSON-friendly — row keys may be strings):
        {"dates":     [10 ISO dates],            # -> Client_FS row 4, J:S
         "rows":      {"6": [10 numbers], ...},  # row number -> J:S values
         "zero_rows": [7, 8, ...],               # rows blanked to 0.0
         "sheet":     "Client_FS"}               # optional, defaults

    UNITS CONTRACT: values are written verbatim — the engine applies NO x1e6
    (or any other) scaling. The Client_FS sheet holds full currency units, so
    the caller sends full values (e.g. 17900000.0 for £17.9m).

    Raises ClientFSBlockInvalid on any shape problem. Returns a normalized
    dict: sheet:str, dates:list[datetime], rows:dict[int, list[float]],
    zero_rows:list[int].
    """
    if not isinstance(block, dict):
        raise ClientFSBlockInvalid(
            f"client_fs must be a dict, got {type(block).__name__}")
    unknown = set(block) - _CLIENT_FS_ALLOWED_KEYS
    if unknown:
        raise ClientFSBlockInvalid(
            f"client_fs has unknown key(s): {sorted(unknown)} "
            f"(allowed: {sorted(_CLIENT_FS_ALLOWED_KEYS)})")

    # ── sheet ────────────────────────────────────────────────────────────
    sheet = block.get("sheet", CLIENT_FS_DEFAULT_SHEET)
    if not isinstance(sheet, str) or not sheet.strip():
        raise ClientFSBlockInvalid(f"client_fs.sheet must be a non-empty string, got {sheet!r}")

    # ── dates ────────────────────────────────────────────────────────────
    raw_dates = block.get("dates")
    if not isinstance(raw_dates, list) or len(raw_dates) != CLIENT_FS_PERIODS:
        raise ClientFSBlockInvalid(
            f"client_fs.dates must be a list of exactly {CLIENT_FS_PERIODS} dates "
            f"(the {CLIENT_FS_FIRST_COL}:S period window), got "
            f"{len(raw_dates) if isinstance(raw_dates, list) else type(raw_dates).__name__}")
    dates: list[datetime] = []
    for i, d in enumerate(raw_dates):
        if isinstance(d, datetime):
            dates.append(d)
            continue
        try:
            dates.append(datetime.fromisoformat(str(d)))
        except ValueError as e:
            raise ClientFSBlockInvalid(
                f"client_fs.dates[{i}] is not an ISO date: {d!r}") from e

    # ── rows ─────────────────────────────────────────────────────────────
    def _row_key(raw: Any, where: str) -> int:
        # JSON object keys arrive as strings — coerce, but refuse non-integers.
        if isinstance(raw, bool) or not isinstance(raw, (int, str)):
            raise ClientFSBlockInvalid(f"client_fs.{where} row key {raw!r} is not an int")
        try:
            row = int(raw)
        except ValueError:
            raise ClientFSBlockInvalid(f"client_fs.{where} row key {raw!r} is not an int")
        if row < 1:
            raise ClientFSBlockInvalid(f"client_fs.{where} row {row} out of range (must be >= 1)")
        if row == CLIENT_FS_DATE_ROW:
            raise ClientFSBlockInvalid(
                f"client_fs.{where} row {row} collides with the date row "
                f"(row {CLIENT_FS_DATE_ROW} is owned by client_fs.dates)")
        return row

    raw_rows = block.get("rows")
    if not isinstance(raw_rows, dict) or not raw_rows:
        raise ClientFSBlockInvalid("client_fs.rows must be a non-empty dict of row -> values")
    rows: dict[int, list[float]] = {}
    for raw_key, values in raw_rows.items():
        row = _row_key(raw_key, "rows")
        if row in rows:
            raise ClientFSBlockInvalid(f"client_fs.rows has duplicate row {row}")
        if not isinstance(values, list) or len(values) != len(dates):
            raise ClientFSBlockInvalid(
                f"client_fs.rows[{raw_key!r}] must be a list of {len(dates)} numbers "
                f"(== dates length), got "
                f"{len(values) if isinstance(values, list) else type(values).__name__}")
        for j, v in enumerate(values):
            if isinstance(v, bool) or not isinstance(v, (int, float)):
                raise ClientFSBlockInvalid(
                    f"client_fs.rows[{raw_key!r}][{j}] is not a number: {v!r}")
            if not math.isfinite(v):
                # json.loads happily parses NaN/Infinity literals — refuse
                # before they poison the workbook (codex review 2026-06-10).
                raise ClientFSBlockInvalid(
                    f"client_fs.rows[{raw_key!r}][{j}] is not finite: {v!r}")
        rows[row] = [float(v) for v in values]

    # ── zero_rows ────────────────────────────────────────────────────────
    raw_zero = block.get("zero_rows", [])
    if not isinstance(raw_zero, list):
        raise ClientFSBlockInvalid("client_fs.zero_rows must be a list of row numbers")
    zero_rows: list[int] = []
    for raw_key in raw_zero:
        row = _row_key(raw_key, "zero_rows")
        if row in rows:
            raise ClientFSBlockInvalid(
                f"client_fs row {row} appears in both rows and zero_rows")
        if row in zero_rows:
            raise ClientFSBlockInvalid(f"client_fs.zero_rows has duplicate row {row}")
        zero_rows.append(row)

    return {"sheet": sheet, "dates": dates, "rows": rows, "zero_rows": zero_rows}


def _write_client_fs(wb, nfs: dict[str, Any]) -> str:
    """Write a normalized client_fs block into the open workbook.

    Refuses (ClientFSFormulaCollision) if ANY target cell holds a formula —
    in the Client_FS layout formula cells are total rows (=SUM(...)); writing
    over one would silently break the template's aggregation. The check runs
    over every target row (date row, data rows, zero rows) BEFORE any write,
    so a refused block leaves the workbook untouched.

    Returns a one-line note for the audit log.
    """
    sheet = nfs["sheet"]
    dates, rows, zero_rows = nfs["dates"], nfs["rows"], nfs["zero_rows"]
    try:
        sht = wb.sheets[sheet]
    except Exception as e:
        raise CellRefUnresolved(f"client_fs sheet {sheet!r} not found in workbook: {e}") from e

    n = len(dates)
    last_col = chr(ord(CLIENT_FS_FIRST_COL) + n - 1)   # J + 10 - 1 = S

    def _row_range(row: int):
        return sht.range(f"{CLIENT_FS_FIRST_COL}{row}:{last_col}{row}")

    # Collision check first — all rows, no writes yet.
    for row in (CLIENT_FS_DATE_ROW, *sorted(rows), *zero_rows):
        # COM Range.HasFormula: True (all formulas), False (none), None (mixed).
        # Anything other than a clean False is a refusal.
        if _row_range(row).api.HasFormula is not False:
            raise ClientFSFormulaCollision(
                f"client_fs refuses to write {sheet}!{CLIENT_FS_FIRST_COL}{row}:{last_col}{row} — "
                f"target holds formula cell(s) (total rows are =SUM() formulas; "
                f"writing values over them would break the template)")

    _row_range(CLIENT_FS_DATE_ROW).value = dates
    for row in sorted(rows):
        _row_range(row).value = rows[row]
    for row in zero_rows:
        _row_range(row).value = [0.0] * n

    return (f"client_fs: wrote {sheet}!{CLIENT_FS_FIRST_COL}{CLIENT_FS_DATE_ROW}:"
            f"{last_col}{CLIENT_FS_DATE_ROW} dates + {len(rows)} data row(s) + "
            f"{len(zero_rows)} zero row(s), values verbatim")


# -------------------------------------------------------- post-recalc hardcode

def _run_post_recalc(wb, app, steps: list[dict[str, Any]]) -> int:
    """Run the post_recalc_hardcode steps.

    Each step copies the current value of `source` into each of `targets`. If
    `converge: true`, we iterate (recalc → re-read source → re-paste) until
    delta < tolerance or max_iters exhausted. Returns total iteration count
    across all steps for the audit log.
    """
    total_iters = 0
    for step in steps:
        source = step["source"]
        targets = step["targets"]
        converge = step["converge"]
        tol = step["tolerance"]
        max_iters = step["max_iters"]

        prev = _read_cell(wb, source)
        for target in targets:
            _write_cell(wb, target, prev)
        app.calculate()
        total_iters += 1

        if not converge:
            continue

        for it in range(max_iters):
            curr = _read_cell(wb, source)
            try:
                delta = abs(float(curr) - float(prev))
            except (TypeError, ValueError):
                # Non-numeric source — can't iterate; one pass and done.
                break
            if delta < tol:
                break
            for target in targets:
                _write_cell(wb, target, curr)
            app.calculate()
            prev = curr
            total_iters += 1
        else:
            raise ConvergenceFailed(
                f"Post-recalc hardcode failed to converge for source={source!r} "
                f"within {max_iters} iterations (last delta={delta:.6g}, tol={tol})"
            )
    return total_iters


# -------------------------------------------------------------------- run

def run(
    spec: TemplateSpec,
    inputs: dict[str, Any],
    output_dir: Path,
    *,
    output_filename: str | None = None,
    keep_excel_open: bool = False,
    verify_hash: bool = True,
    client_fs: dict[str, Any] | None = None,
) -> EngineRun:
    """Execute one template against a set of inputs.

    Steps:
      1. Validate input keys against the spec (required vs optional vs unknown);
         validate the optional client_fs block shape.
      2. Verify the live .xlsx hash still matches what was registered.
      3. Copy template → output path (a fresh file is always produced).
      4. Open via xlwings (Excel COM, headless).
      5. Write the optional client_fs operating-model block, then the inputs,
         into mapped cells — both inside one manual-calc window when client_fs
         is present (restored to the workbook's native mode before recalc).
      6. Trigger full recalc.
      7. Run post-recalc hardcode steps (with optional convergence loop).
      8. Read outputs from mapped cells.
      9. Run validation rules; fail loud if any fail.
     10. Save, close, return EngineRun.

    `client_fs` (optional) is an operating-model block written into the
    template's Client_FS sheet BEFORE the input writes — see
    :func:`_normalize_client_fs` for the schema and the verbatim-units
    contract (the caller sends full currency values; the engine applies no
    scaling).

    Raises:
        InputCellMismatch, TemplateHashMismatch, CellRefUnresolved,
        ClientFSBlockInvalid, ClientFSFormulaCollision,
        ConvergenceFailed, ValidationFailed, ExcelDriverError.
    """
    started_at = datetime.now()
    t0 = time.monotonic()

    _check_input_keys(spec, inputs)
    # Fail fast on a malformed client_fs block — before any file IO / Excel.
    nfs = _normalize_client_fs(client_fs) if client_fs is not None else None
    if verify_hash:
        _verify_hash(spec)

    if output_filename:
        run_id = uuid.uuid4().hex[:8]
        out_path = output_dir / output_filename
    else:
        run_id, out_path = _new_output_path(spec, output_dir)

    _copy_template(spec.path, out_path)

    # Lazy import — xlwings is Windows-only; unit tests on Linux mustn't import it.
    try:
        import xlwings as xw  # type: ignore
    except ImportError as e:
        raise ExcelDriverError(
            "xlwings not installed. Install the engine on Windows with `pip install -e .[dev]`."
        ) from e

    app = None
    wb = None
    convergence_iters = 0
    outputs: dict[str, Any] = {}
    notes: list[str] = []
    status = "ok"
    # F-14: set when validation fails so the live ``vN`` deliverable copy left
    # by ``_copy_template`` is removed after Excel releases it (a rejected model
    # must leave NO artifact at the deliverable path).
    discard_output = False

    try:
        app = xw.App(visible=False, add_book=False)
        # Suppress dialogs / link-update prompts / save-as alerts
        try:
            app.api.DisplayAlerts = False
            app.api.AskToUpdateLinks = False
            app.api.ScreenUpdating = False
        except Exception as e:  # pragma: no cover — Excel COM API drift
            notes.append(f"alert suppression partial: {e}")

        wb = app.books.open(str(out_path), update_links=False, read_only=False)

        # 5. Write phase. With a client_fs block the bulk row writes (and the
        # input writes after them) run inside a manual-calc window so each
        # write doesn't recalc the template's iterative/circular graph. The
        # native mode (typically 2 = automatic-except-tables on this template)
        # is restored before the single full recalc — NEVER forced to
        # xlCalculationAutomatic (-4105), which makes the IRR data table
        # recalc continuously and hangs Excel.
        orig_calc = None
        try:
            if nfs is not None:
                # Entering the window must SUCCEED — proceeding with bulk row
                # writes in the native (semi-automatic) mode would recalc the
                # iterative/circular graph on every write (codex review
                # 2026-06-10: engine error, not a note).
                try:
                    orig_calc = app.api.Calculation
                    app.api.Calculation = XL_CALCULATION_MANUAL
                except Exception as e:
                    orig_calc = None  # mode unchanged — nothing to restore
                    raise ExcelDriverError(
                        f"could not enter manual-calc window for client_fs: {e}") from e
                notes.append(_write_client_fs(wb, nfs))

            # 5b. Write inputs (required + provided optionals)
            for logical, value in inputs.items():
                ref = spec.inputs.get(logical) or spec.optional_inputs.get(logical)
                if ref is None:
                    # already caught by _check_input_keys, but defensive
                    raise InputCellMismatch(f"No cell map entry for input {logical!r}")
                _write_cell(wb, ref, value)
        finally:
            # ALWAYS restore the native calc mode — a collision/COM failure
            # mid-write must never leave the Excel app in manual (codex review
            # 2026-06-10 SEV-1). On an in-flight exception a restore failure
            # must not mask it; on the success path it is fatal (the full
            # recalc below would run in the wrong mode).
            if orig_calc is not None:
                try:
                    app.api.Calculation = orig_calc
                except Exception as e:  # pragma: no cover — Excel COM API drift
                    if sys.exc_info()[0] is None:
                        raise ExcelDriverError(
                            f"failed to restore calc mode after writes: {e}") from e
                    notes.append(f"calc-mode restore failed during error teardown: {e}")

        # 6. Full recalc
        app.calculate()

        # 7. Post-recalc hardcode + convergence loop
        if spec.post_recalc_hardcode:
            convergence_iters = _run_post_recalc(wb, app, spec.post_recalc_hardcode)

        # 8. Read outputs
        for logical, ref in spec.outputs.items():
            outputs[logical] = _read_cell(wb, ref)

        # 9. Validation
        failures = _evaluate_validation(outputs, spec.validation)
        if failures:
            status = "validation_failed"
            # F-14 (Iron Law / HR S-17): do NOT persist a rejected model at the
            # live ``vN`` deliverable path — it could be mistaken for a real
            # output. Quarantine the populated-but-invalid workbook to a
            # ``*.REJECTED.xlsx`` path and flag the live copy for removal.
            _quarantine_rejected_workbook(wb, out_path, run_id, notes)
            discard_output = True
            raise ValidationFailed(
                f"Validation failed for {spec.name!r}:\n  - " + "\n  - ".join(failures)
            )

        # Save
        wb.save()

    except ExcelDriverError:
        status = "excel_error"
        raise
    except Exception as e:
        # Wrap unexpected COM failures
        if not isinstance(e, (InputCellMismatch, TemplateHashMismatch,
                              CellRefUnresolved, ClientFSBlockInvalid,
                              ClientFSFormulaCollision, ConvergenceFailed,
                              ValidationFailed)):
            status = "excel_error"
            raise ExcelDriverError(f"Unexpected Excel-COM failure: {e}") from e
        raise
    finally:
        if wb is not None:
            try:
                wb.close()
            except Exception:
                pass
        if app is not None and not keep_excel_open:
            try:
                app.quit()
            except Exception:
                try:
                    app.kill()
                except Exception:
                    pass
        # F-14: remove the live deliverable copy AFTER Excel has released the
        # file (post wb.close()/app.quit()). The populated-but-invalid workbook
        # survives at its ``*.REJECTED.<run_id>.xlsx`` quarantine path; the
        # ``vN`` path must not. Retry briefly in case Excel is slow to release
        # the handle on Windows, and LOUDLY warn if it still can't be removed —
        # a silently-surviving template copy at the deliverable path defeats the
        # whole point (codex-5.5 F-14 round 1).
        if discard_output:
            removed = False
            for _attempt in range(3):
                try:
                    out_path.unlink()
                    removed = True
                    break
                except FileNotFoundError:
                    removed = True  # already gone (e.g. moved by the SaveAs)
                    break
                except OSError:
                    time.sleep(0.2)
            if not removed:
                logger.warning(
                    "F-14: could not remove the rejected deliverable copy at %s "
                    "(Excel may still hold the file). A non-deliverable template "
                    "copy may remain at the vN path — delete it manually.",
                    out_path,
                )

    duration_ms = int((time.monotonic() - t0) * 1000)
    return EngineRun(
        run_id=run_id,
        template_name=spec.name,
        template_version=spec.version,
        template_hash=spec.template_hash,
        output_path=out_path,
        inputs=dict(inputs),
        outputs=outputs,
        started_at=started_at,
        duration_ms=duration_ms,
        status=status,
        convergence_iters=convergence_iters,
        notes=notes,
    )


# ----------------------------------------------------------------- validate

def validate(spec: TemplateSpec) -> list[str]:
    """Offline validation — check the cell map resolves against the live .xlsx
    without opening Excel. Uses openpyxl (read-only), so this works in CI on
    Linux too.

    Returns a list of issues (empty list = all good).
    """
    issues: list[str] = []
    try:
        from openpyxl import load_workbook
    except ImportError:
        return ["openpyxl not installed — cannot validate offline"]

    try:
        wb = load_workbook(spec.path, data_only=False, read_only=True)
    except Exception as e:
        return [f"could not open {spec.path}: {e}"]

    sheet_names = set(wb.sheetnames)
    # Build set of defined-name keys for fast lookup
    defined_names: set[str] = set()
    try:
        defined_names = set(list(wb.defined_names))
    except Exception:
        pass

    def _check(category: str, refs: dict[str, str]) -> None:
        for logical, ref in refs.items():
            parsed = parse_ref(ref)
            if parsed.is_named:
                if parsed.addr not in defined_names:
                    issues.append(f"[{category}] {logical}={ref!r}: named range not found")
                continue
            if not parsed.sheet:
                issues.append(f"[{category}] {logical}={ref!r}: A1 ref missing sheet prefix")
                continue
            if parsed.sheet not in sheet_names:
                issues.append(f"[{category}] {logical}={ref!r}: sheet {parsed.sheet!r} not in workbook")

    _check("inputs", spec.inputs)
    _check("optional_inputs", spec.optional_inputs)
    _check("outputs", spec.outputs)

    # Post-recalc hardcode refs
    for i, step in enumerate(spec.post_recalc_hardcode):
        _check(f"post_recalc[{i}].source", {"source": step["source"]})
        for j, t in enumerate(step["targets"]):
            _check(f"post_recalc[{i}].target[{j}]", {"target": t})

    try:
        wb.close()
    except Exception:
        pass
    return issues
