"""Openpyxl writer for the earnings tracker workbook.

Mirrors routines.dealtracker.workbook patterns: create-with-header on
first write, idempotent append keyed on (ticker, period_label), datetime
to date normalisation on read-back.
"""

from __future__ import annotations

import logging
from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook

from routines.earnings.schema import COLUMNS, EarningsRecord

log = logging.getLogger(__name__)

SHEET_NAME = "Earnings"


def _norm_date(v: Any) -> date | None:
    if v is None:
        return None
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    return None


def _row_to_record(row: tuple[Any, ...]) -> EarningsRecord:
    """Build an EarningsRecord from a workbook row (length-flexible)."""
    cells = list(row) + [None] * (len(COLUMNS) - len(row))
    return EarningsRecord(
        period_end=_norm_date(cells[0]),
        period_label=str(cells[1] or "").strip(),
        ticker=str(cells[2] or "").strip(),
        company=str(cells[3] or "").strip(),
        currency=str(cells[4] or "").strip(),
        revenue_m=_to_float(cells[5]),
        revenue_yoy=_to_float(cells[6]),
        ebitda_m=_to_float(cells[7]),
        ebitda_margin=_to_float(cells[8]),
        ebit_m=_to_float(cells[9]),
        net_income_m=_to_float(cells[10]),
        eps=_to_float(cells[11]),
        fcf_m=_to_float(cells[12]),
        notes=str(cells[13] or ""),
        source=str(cells[14] or ""),
    )


def _to_float(v: Any) -> float | None:
    if v is None or v == "":
        return None
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def append_earnings(workbook_path: Path, rec: EarningsRecord) -> dict[str, Any]:
    """Append one earnings record. Idempotent on dedupe_key().

    Creates the workbook with COLUMNS header row if it doesn't exist.
    Returns a status dict describing what happened.
    """
    workbook_path = Path(workbook_path)
    workbook_path.parent.mkdir(parents=True, exist_ok=True)

    if workbook_path.exists():
        wb = load_workbook(workbook_path)
        if SHEET_NAME not in wb.sheetnames:
            ws = wb.create_sheet(SHEET_NAME)
            ws.append(COLUMNS)
        else:
            ws = wb[SHEET_NAME]
    else:
        wb = Workbook()
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]
        ws = wb.create_sheet(SHEET_NAME)
        ws.append(COLUMNS)

    # Dedupe scan.
    new_key = rec.dedupe_key()
    for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        existing = _row_to_record(row)
        if existing.dedupe_key() == new_key:
            return {"status": "skipped_duplicate", "existing_row": idx}

    ws.append(rec.to_row())
    wb.save(str(workbook_path))
    return {"status": "appended", "row": ws.max_row}
