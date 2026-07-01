"""CLI for the earnings tracker."""

from __future__ import annotations

import logging
import sys
from datetime import date
from pathlib import Path

import click

from routines.earnings.pull import pull_earnings
from routines.earnings.schema import COLUMNS
from routines.earnings.workbook import append_earnings
from routines.shared import audit
from routines.shared.vault_writer import VaultPaths


DEFAULT_VAULT = Path("/mnt/x/OS AI Vault")
DEFAULT_AUDIT_DIR = Path(__file__).resolve().parent.parent.parent / "runs"
DEFAULT_WORKBOOK_REL = "Projects/_Trackers/Earnings.xlsx"


def _resolve_vault(vault: Path | None) -> Path:
    """Resolve the vault root for the calendar-driven commands.

    The legacy ``pull`` commands hard-default to the WSL path; the new
    calendar-driven ``run`` / ``add-watch`` fire from the cron with no
    ``--vault`` on native Windows, so we fall back to the platform-aware
    bridge default (``routines.api.deps.VAULT``) when nothing is passed."""
    if vault is not None:
        return vault
    from routines.api.deps import VAULT as _BRIDGE_VAULT
    return _BRIDGE_VAULT


@click.group()
@click.option("--debug", is_flag=True)
def main(debug: bool) -> None:
    """Earnings tracker — quarterly results into an Excel workbook."""
    logging.basicConfig(
        level=logging.DEBUG if debug else logging.INFO,
        format="%(asctime)s %(levelname)-7s %(name)s: %(message)s",
        datefmt="%H:%M:%S",
    )


@main.command("pull")
@click.argument("ticker")
@click.option("--periods", type=int, default=4, show_default=True,
              help="Number of fiscal periods to pull")
@click.option("--vault", type=click.Path(exists=True, file_okay=False, path_type=Path),
              default=DEFAULT_VAULT)
@click.option("--workbook",
              help="Path to tracker workbook (default: <vault>/Projects/_Trackers/Earnings.xlsx)")
@click.option("--dry-run", is_flag=True, help="Show records but don't write")
def pull_cmd(ticker: str, periods: int, vault: Path, workbook: str | None, dry_run: bool) -> None:
    """Pull recent earnings for TICKER and append to the tracker workbook.

    Examples:
        earnings-tracker pull WTB.L
        earnings-tracker pull IHG.L --periods 6 --dry-run
    """
    paths = VaultPaths(vault)
    workbook_path = Path(workbook) if workbook else (paths.root / DEFAULT_WORKBOOK_REL)
    DEFAULT_AUDIT_DIR.mkdir(parents=True, exist_ok=True)

    run_id = audit.new_run_id()
    records = pull_earnings(ticker.upper(), periods=periods)

    if not records:
        click.echo(f"No earnings records returned for {ticker}.", err=True)
        sys.exit(1)

    click.echo(f"\nPulled {len(records)} records for {ticker}:\n")
    for r in records:
        rev = f"{r.revenue_m:,.0f}m" if r.revenue_m else "—"
        yoy = f"{r.revenue_yoy * 100:+.1f}%" if r.revenue_yoy is not None else "—"
        margin = f"{r.ebitda_margin * 100:.1f}%" if r.ebitda_margin is not None else "—"
        click.echo(f"  {r.period_label:12s} rev {rev:>12s}  YoY {yoy:>8s}  margin {margin:>7s}")

    if dry_run:
        click.echo("\n(dry-run — not writing)")
        return

    appended = 0
    skipped = 0
    for r in records:
        result = append_earnings(workbook_path, r)
        if result["status"] == "appended":
            appended += 1
        else:
            skipped += 1

    click.echo(f"\nOK  appended {appended} new · skipped {skipped} duplicates")
    click.echo(f"Workbook: {workbook_path}")

    audit.write_structured(
        actor={"type": "system", "id": "routine:earnings"},
        entity_type="vault_note",
        entity_id=str(workbook_path),
        action="run",
        routine="earnings", run_id=run_id, status="ok",
        audit_dir=DEFAULT_AUDIT_DIR,
        semantic_target=str(workbook_path),
        inputs={"ticker": ticker, "periods": periods},
        outputs={
            "workbook": str(workbook_path),
            "appended": appended,
            "skipped": skipped,
        },
    )


@main.command("pull-all")
@click.option("--vault", type=click.Path(exists=True, file_okay=False, path_type=Path),
              default=DEFAULT_VAULT)
@click.option("--periods", type=int, default=4, show_default=True,
              help="Number of fiscal periods to pull per ticker")
@click.option("--workbook",
              help="Path to tracker workbook (default: <vault>/Projects/_Trackers/Earnings.xlsx)")
@click.option("--dry-run", is_flag=True, help="Show records but don't write")
def pull_all_cmd(vault: Path, periods: int, workbook: str | None, dry_run: bool) -> None:
    """Pull earnings for every ticker in `_claude/earnings-watchlist.md`.

    Reads the markdown frontmatter, runs `pull` for each entry, appends
    to the tracker workbook. Idempotent — re-running the same period
    doesn't duplicate rows.
    """
    from routines.earnings.watchlist import load as load_watchlist
    paths = VaultPaths(vault)
    workbook_path = Path(workbook) if workbook else (paths.root / DEFAULT_WORKBOOK_REL)
    DEFAULT_AUDIT_DIR.mkdir(parents=True, exist_ok=True)

    entries = load_watchlist(paths.root)
    click.echo(f"watchlist: {len(entries)} tickers\n")

    run_id = audit.new_run_id()
    total_appended = 0
    total_skipped = 0
    failed: list[str] = []

    for e in entries:
        click.echo(f"-- {e.symbol} {e.name}")
        try:
            records = pull_earnings(e.symbol, periods=periods)
        except Exception as exc:  # noqa: BLE001
            click.echo(f"   FAIL: {exc}", err=True)
            failed.append(e.symbol)
            continue
        if not records:
            click.echo(f"   (no records returned)")
            continue
        for r in records:
            rev = f"{r.revenue_m:,.0f}m" if r.revenue_m else "-"
            click.echo(f"   {r.period_label:10s} rev {rev}")
            if dry_run:
                continue
            result = append_earnings(workbook_path, r)
            if result["status"] == "appended":
                total_appended += 1
            else:
                total_skipped += 1

    click.echo(f"\nOK  watchlist={len(entries)} appended={total_appended} skipped={total_skipped} failed={len(failed)}")
    if not dry_run:
        click.echo(f"Workbook: {workbook_path}")

    audit.write_structured(
        actor={"type": "system", "id": "routine:earnings"},
        entity_type="vault_note",
        entity_id=str(workbook_path),
        action="run",
        routine="earnings", run_id=run_id, status="ok" if not failed else "error",
        audit_dir=DEFAULT_AUDIT_DIR,
        semantic_target=str(workbook_path),
        inputs={"watchlist_size": len(entries), "periods": periods},
        outputs={
            "workbook": str(workbook_path),
            "appended": total_appended,
            "skipped": total_skipped,
            "failed": failed,
        },
    )


@main.command("init")
@click.option("--vault", type=click.Path(exists=True, file_okay=False, path_type=Path),
              default=DEFAULT_VAULT)
@click.option("--workbook",
              help="Path to tracker workbook (default: <vault>/Projects/_Trackers/Earnings.xlsx)")
def init_cmd(vault: Path, workbook: str | None) -> None:
    """Create an empty earnings workbook with the header row, if missing."""
    paths = VaultPaths(vault)
    workbook_path = Path(workbook) if workbook else (paths.root / DEFAULT_WORKBOOK_REL)
    if workbook_path.exists():
        click.echo(f"Workbook already exists: {workbook_path}")
        return

    from openpyxl import Workbook

    workbook_path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]
    ws = wb.create_sheet("Earnings")
    ws.append(COLUMNS)
    wb.save(str(workbook_path))
    click.echo(f"Created: {workbook_path}")
    click.echo(f"Sheet: Earnings, {len(COLUMNS)} columns, 0 data rows")


@main.command("run")
@click.option("--vault", type=click.Path(file_okay=False, path_type=Path), default=None,
              help="Vault root (default: platform-aware bridge default).")
@click.option("--company", default=None,
              help="Restrict the sweep to one company (note stem or name).")
@click.option("--as-of", "as_of", default=None,
              help="Run date override (YYYY-MM-DD); default today. Calendar is "
                   "matched against this date.")
@click.option("--no-overdue", is_flag=True,
              help="Only fire companies whose next-reporting-date == the run date "
                   "(skip catch-up of overdue ones).")
@click.option("--model", default=None, help="Ollama model override (default qwen3:14b).")
def run_cmd(vault: Path | None, company: str | None, as_of: str | None,
            no_overdue: bool, model: str | None) -> None:
    """Run the calendar-driven earnings pipeline (steps 1-11).

    Scans ``Companies/*.md`` for ``type: public-company`` notes due today,
    fetches + extracts each announcement (local Ollama), compares vs consensus +
    prior periods, appends a dated section to the company page, aggregates onto
    the sector page, raises a material-variance proposal, and rolls
    ``next-reporting-date`` forward. Idempotent — a re-fire on the same period is
    a no-op. Catch-up: overdue companies are re-swept until their announcement
    is captured (disable with ``--no-overdue``).

    Examples:
        earnings-tracker run
        earnings-tracker run --company "Whitbread plc"
        earnings-tracker run --as-of 2026-05-07 --no-overdue
    """
    from routines.earnings.pipeline import run_sweep
    from routines.sectornews.firecrawl_client import FirecrawlClient, FirecrawlError
    from routines.shared.ollama_client import OllamaClient

    vault_root = _resolve_vault(vault)
    if not vault_root.is_dir():
        click.echo(f"ERROR: vault not found: {vault_root}", err=True)
        sys.exit(1)

    try:
        run_date = date.fromisoformat(as_of) if as_of else date.today()
    except ValueError:
        click.echo(f"ERROR: --as-of must be YYYY-MM-DD, got {as_of!r}", err=True)
        sys.exit(2)

    DEFAULT_AUDIT_DIR.mkdir(parents=True, exist_ok=True)
    run_id = audit.new_run_id()

    try:
        fc_client = FirecrawlClient()
    except FirecrawlError as e:
        click.echo(f"ERROR: Firecrawl unavailable: {e}", err=True)
        sys.exit(1)
    ollama_client = OllamaClient()

    extract_model = model or "qwen3:14b"
    result = run_sweep(
        vault_root=vault_root, today=run_date,
        fc_client=fc_client, ollama_client=ollama_client,
        only=company, include_overdue=not no_overdue,
        run_id=run_id, model=extract_model,
    )

    counts = result.counts()
    click.echo(f"\nearnings sweep {result.run_date} — {len(result.outcomes)} due\n")
    for o in result.outcomes:
        extra = f" → {o.proposal_path}" if o.proposal_path else ""
        mat = " ⚠material" if o.material else ""
        click.echo(f"  {o.status:14s} {o.company}{mat}{extra}")
    click.echo(
        f"\nOK  captured={result.captured} skipped={counts['skipped_exists']} "
        f"not-published={counts['not_published']} fetch-error={counts['fetch_error']} "
        f"extract-failed={counts['extract_failed']} wrong-issuer={counts['wrong_issuer']} "
        f"errors={counts['error']} proposals={result.proposals}"
    )

    # Audit "partial" for ANY non-terminal-success outcome — not just hard errors.
    # extract_failed / wrong_issuer / missing_page / contended all mean a due
    # company did NOT cleanly capture this sweep (it stays due, the cron re-fires),
    # so the run is partial and must be visible as such, not silently "ok" (#44
    # Codex SEV-2). captured / skipped_exists / not_published are the clean
    # terminal-success outcomes — BUT even a "captured" outcome is only partial-
    # success if a durable side-effect (sector point / proposal) failed and left
    # the date unrolled (``side_effects_ok=False``); that too must audit partial.
    _TERMINAL_OK = {"captured", "skipped_exists", "not_published"}
    status = "ok"
    if any(o.status not in _TERMINAL_OK or not o.side_effects_ok for o in result.outcomes):
        status = "partial"
    audit.write_structured(
        actor={"type": "system", "id": "routine:earnings"},
        entity_type="vault_note",
        entity_id=str(vault_root / "Companies"),
        action="run",
        routine="earnings", run_id=run_id, status=status,
        audit_dir=DEFAULT_AUDIT_DIR,
        semantic_target=str(vault_root / "Companies"),
        inputs={"run_date": result.run_date, "company": company, "include_overdue": not no_overdue},
        outputs={
            "due": len(result.outcomes),
            "captured": result.captured,
            "proposals": result.proposals,
            "counts": counts,
        },
    )


@main.command("add-watch")
@click.argument("company")
@click.option("--vault", type=click.Path(file_okay=False, path_type=Path), default=None,
              help="Vault root (default: platform-aware bridge default).")
@click.option("--ticker", default="", help="Ticker symbol, e.g. WTB.L")
@click.option("--reporting-date", "reporting_date", required=True,
              help="Next reporting date (YYYY-MM-DD) — the calendar trigger.")
@click.option("--sector", default="", help="Sector slug or wikilink for sector aggregation.")
@click.option("--consensus-source", "consensus_source", default="",
              help="Provenance label for consensus figures (e.g. 'Bloomberg consensus').")
@click.option("--cadence", type=click.Choice(["quarterly", "semi-annual", "annual"]),
              default="quarterly", show_default=True,
              help="Reporting cadence — how far to roll next-reporting-date forward.")
@click.option("--source-url", "source_url", default="",
              help="IR / RNS results URL to scrape on the reporting date (optional).")
def add_watch_cmd(company: str, vault: Path | None, ticker: str, reporting_date: str,
                  sector: str, consensus_source: str, cadence: str, source_url: str) -> None:
    """Scaffold a new watched public company at ``Companies/<company>.md``.

    Writes a minimal ``type: public-company`` note carrying the
    ``next-reporting-date`` calendar trigger so the daily ``run`` sweep picks it
    up on that date. Refuses to overwrite an existing note (append-only).

    Example:
        earnings-tracker add-watch "Whitbread plc" --ticker WTB.L \\
            --reporting-date 2026-04-29 --sector hospitality \\
            --consensus-source "Bloomberg consensus"
    """
    import frontmatter

    from routines.earnings.calendar import PUBLIC_COMPANY_TYPE
    from routines.shared.vault_writer import atomic_write

    vault_root = _resolve_vault(vault)
    companies_dir = vault_root / "Companies"
    if not companies_dir.is_dir():
        click.echo(f"ERROR: Companies/ not found under vault: {vault_root}", err=True)
        sys.exit(1)

    try:
        next_date = date.fromisoformat(reporting_date)
    except ValueError:
        click.echo(f"ERROR: --reporting-date must be YYYY-MM-DD, got {reporting_date!r}", err=True)
        sys.exit(2)

    # Keep the note stem a single safe path segment.
    stem = company.strip()
    if not stem or any(c in stem for c in ("/", "\\", ":")) or ".." in stem:
        click.echo(f"ERROR: invalid company name {company!r}", err=True)
        sys.exit(2)

    path = companies_dir / f"{stem}.md"
    if path.exists():
        click.echo(f"Company note already exists: {path}\n"
                   f"  (edit its frontmatter directly to set type/next-reporting-date)")
        return

    metadata: dict[str, object] = {
        "type": PUBLIC_COMPANY_TYPE,
        "memory_kind": "semantic",
        "name": stem,
        "ticker": ticker.strip(),
        "sector": sector.strip(),
        "next-reporting-date": next_date.isoformat(),
        "reporting-cadence": cadence,
        "consensus-source": consensus_source.strip(),
        "sensitivity": "public",
        "tags": ["company", "public-company", "earnings-watch"],
    }
    if source_url.strip():
        metadata["earnings-source-url"] = source_url.strip()

    body = (
        f"# {stem}\n\n"
        f"## Snapshot\n\n"
        f"## Earnings history\n\n"
        f"*(populated automatically by the earnings tracker on each reporting date)*\n"
    )
    post = frontmatter.Post(body)
    post.metadata.update(metadata)
    atomic_write(path, frontmatter.dumps(post) + "\n", vault_root=vault_root)

    click.echo(f"Created watched company: {path}")
    click.echo(f"  type: {PUBLIC_COMPANY_TYPE} · next-reporting-date: {next_date.isoformat()} "
               f"· cadence: {cadence}")


if __name__ == "__main__":
    main()
