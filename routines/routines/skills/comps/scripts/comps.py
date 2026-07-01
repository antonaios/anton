"""Comps skill — Stage 0-3 operator-gated pipeline (#21-comps, per COMPS-REDESIGN-2026-06-01).

Greenfield deliverable skill. NOT to be confused with
``routines.markets.comps.build_comps`` (the ticker-multiples leaf, shared by
#21-equity-research as a sub-leaf — explicitly out of scope to rename).

Pipeline shape:

  * **Stage 0** (``run_stage_0``) — read target from the vault, propose
    subsectors from ``profile.md sector_sub_lens.<parent-sector>``, return an
    ``approval_pending`` payload. Operator approves the subset; the approval
    token is the entry condition for Stage 1.
  * **Stage 1** (``run_stage_1``) — per approved subsector: propose CoCo peers
    via orchestrated cloud skills (``equity-research:screen`` /
    ``investment-banking:buyer-list``) + the markets provider's ``get_peers``;
    propose CoTrans deals via the canonical precedent tracker + ``deep-research``
    gap-fill. WRITE NEW DEALS BACK to the tracker via
    ``routines.dealtracker.workbook.append_deal``. Return ``approval_pending``.
  * **Stage 2** (``run_stage_2``) — acquire each populated figure WITH A
    SOURCE: provider for CoCo trading data; IR / Firecrawl scrape for net debt;
    deep-research (or connector when licensed) for LFY+1 + CoTrans gaps.
    Flag every ccy mismatch + every unsourced LFY+1 — never silently fill.
    Return ``approval_pending`` for any unsourced figure.
  * **Stage 3** (``run_stage_3``) — stamp the v2 template (block-flex: one
    CoCo block per approved subsector + CoTrans grouped by subsector),
    save to the deal Valuation folder per workspace-write-policy, archive
    prior version, fire the #76 capture, refresh the Sectors mirror.

The route (``routines.api.routes.comps``) wraps each stage with the central
``tool_call_hooks`` stack — the ``enforce_skill_sensitivity`` guard fires on
the ``@before_tool_call`` path and refuses non-project workspaces / MNPI
inputs.

ANTHROPIC CLOUD SKILL ORCHESTRATION. Stages 1-2 conceptually call
``equity-research:screen``, ``investment-banking:buyer-list``, ``deep-research``,
``anthropic-skills:xlsx``. These live as ENGINE-skill SHIMS at the bottom of
this module so tests can mock them cleanly (and so the no-MNPI-to-cloud check
fires BEFORE any shim is invoked, per the route's sensitivity gate).
"""

from __future__ import annotations

import hashlib
import hmac
import json
import logging
import os
import copy
import re
import unicodedata
import secrets
import shutil
import statistics
import threading
import time
import uuid
from collections import Counter
from dataclasses import dataclass, field
from datetime import date, datetime, timezone
from pathlib import Path
from typing import Any, Literal, Optional

from pydantic import BaseModel, Field

log = logging.getLogger(__name__)


# ─────────────────────────────────────────────────────────────────────────────
# WS-B: HMAC approval-token gate (#21-comps-hardening).
#
# Each stage approval token is HMAC-signed over (deal_name, stage_label,
# canonical-hash(approved-payload)) keyed by a per-process secret. The route
# verifies the token on re-fire: forged / missing / stale / cross-stage-
# replayed tokens are rejected with HTTP 422.
#
# Secret location: read from env `AGENTIC_COMPS_HMAC_SECRET` if set; otherwise
# generated-on-boot via `secrets.token_hex(32)` and stashed in module state.
# Rotate-on-restart is fine for the loopback-only single-operator case
# (today's deployment). MULTI-OPERATOR / EXTERNAL UPGRADE PATH: set
# AGENTIC_COMPS_HMAC_SECRET to a persistent 32-byte hex string shared across
# the operator's bridge processes (e.g. via vault / process-manager env) so
# tokens issued on one process remain verifiable on another. Without a
# persistent secret, restarting the bridge invalidates all in-flight Stage 0/
# 1/2 approvals — the operator must re-run from Stage 0.
# ─────────────────────────────────────────────────────────────────────────────

_HMAC_SECRET_ENV = "AGENTIC_COMPS_HMAC_SECRET"
_HMAC_SECRET_FALLBACK: Optional[bytes] = None
_HMAC_SECRET_LOCK = threading.Lock()

# Stage labels — load-bearing for cross-stage replay defence. A Stage-0-issued
# token signs `(deal, "stage_1_subsectors", h(payload))`. If a caller tries to
# present that exact token at Stage 2 verification (which expects
# `"stage_2_peers"` / `"stage_2_deals"` in the HMAC payload), the recomputed
# HMAC will mismatch — replay rejected.
STAGE_LABEL_SUBSECTORS = "stage_1_subsectors"
STAGE_LABEL_PEERS = "stage_2_peers"
STAGE_LABEL_DEALS = "stage_2_deals"
STAGE_LABEL_STAGE_2_BLOCKS = "stage_3_data"
STAGE_LABEL_ASSUMPTIONS = "stage_3_assumptions"


# ─────────────────────────────────────────────────────────────────────────────
# Canonical-precedent-tracker URL form (#21-comps CoTrans Source hyperlink v2,
# operator clarification 2026-06-02). The CoTrans Source cell hyperlinks to the
# operator's curated tracker xlsx (which carries their own source link + curated
# context) — NOT to a Mergermarket URL the operator can't access. The path is
# the dealtracker module's CANONICAL_WORKBOOK_PATH; the URL form is the same
# path file:// encoded so Excel renders a clickable cross-workbook link.
# ─────────────────────────────────────────────────────────────────────────────

_CANONICAL_TRACKER_SHEET = "Precedent transactions"


def _canonical_tracker_file_url() -> Optional[str]:
    """Return the file:// URL pointing at the dealtracker module's canonical
    workbook path, URL-encoded so spaces / ampersands render cleanly in Excel.
    Returns None when the dealtracker module is unavailable (test env without
    it on the path) — callers fall back to no-hyperlink in that case."""
    try:
        from routines.dealtracker.workbook import CANONICAL_WORKBOOK_PATH
    except Exception:  # noqa: BLE001
        return None
    return _tracker_path_to_file_url(CANONICAL_WORKBOOK_PATH)


def _tracker_path_to_file_url(p: Path) -> str:
    """Convert a Windows-style tracker xlsx Path into a file:// URL with the
    path component URL-encoded (so spaces + `&` + other special chars in
    `Corporate Finance/4. Research & data/...` survive Excel's URL parser).

    Implementation note: ``urllib.parse.quote`` with ``safe="/"`` matches
    Excel's accepted form — preserves slashes, escapes spaces as ``%20`` and
    ampersand as ``%26``. We normalise backslashes → forward slashes first so
    the same routine works for both Windows + POSIX path inputs (and for the
    tmp_path fixtures the tests use)."""
    from urllib.parse import quote

    s = str(p).replace("\\", "/")
    # On Windows: "X:/..." → "file:///X:/..." (three slashes; the first two are
    # the file:// scheme separator, the third is the absolute-path lead).
    # On POSIX: "/x/..." → "file:///x/..." (same shape, lead slash preserved).
    if len(s) >= 2 and s[1] == ":":      # drive-letter form
        encoded = quote(s, safe="/:")
        return f"file:///{encoded}"
    if s.startswith("/"):
        encoded = quote(s, safe="/")
        return f"file://{encoded}"
    encoded = quote(s, safe="/")
    return f"file:///{encoded}"


def _get_hmac_secret() -> bytes:
    """Return the per-process HMAC secret. Reads env first; falls back to a
    generated-on-first-call 32-byte token cached in module state."""
    global _HMAC_SECRET_FALLBACK
    env = os.environ.get(_HMAC_SECRET_ENV)
    if env:
        return env.encode("utf-8")
    with _HMAC_SECRET_LOCK:
        if _HMAC_SECRET_FALLBACK is None:
            _HMAC_SECRET_FALLBACK = secrets.token_hex(32).encode("utf-8")
            log.info(
                "comps: HMAC secret generated-on-boot (rotate-on-restart). "
                "Set %s for multi-process / persistent verification.",
                _HMAC_SECRET_ENV,
            )
        return _HMAC_SECRET_FALLBACK


def _canonical_hash(payload: Any) -> str:
    """Deterministic sha256 of a JSON-serialisable payload.

    Uses ``sort_keys=True`` + ``separators=(",", ":")`` so the same logical
    payload produces the same hash across re-fires (independent of dict
    insertion order or whitespace). Default JSON-encoder fallback handles
    objects via ``default=str`` for the rare path where Stage outputs carry
    non-JSON-native types (e.g. dates).
    """
    return hashlib.sha256(
        json.dumps(payload, sort_keys=True, separators=(",", ":"),
                   default=str).encode("utf-8"),
    ).hexdigest()


def _sign_token(deal_name: str, stage_label: str, payload: Any) -> str:
    """HMAC-sign ``(deal_name, stage_label, canonical-hash(payload))``.
    Returns hex digest. Uses sha256.

    The signed message is a CANONICAL JSON encoding of the 3-tuple — NOT a
    ``"|"``-joined string. A ``"|"`` in ``deal_name`` / ``stage_label`` would
    otherwise shift field boundaries so two DIFFERENT logical tuples produce the
    same message (e.g. ``("a|b","c")`` vs ``("a","b|c")`` both → ``"a|b|c|…"``),
    enabling a cross-deal / cross-stage collision. JSON quotes + escapes each
    element, so the encoding is injective. (codex-5.5 — low reachability since
    ``deal_name`` is operator-controlled, but cheap to make non-forgeable.)"""
    msg = json.dumps(
        [deal_name, stage_label, _canonical_hash(payload)],
        separators=(",", ":"),
    ).encode("utf-8")
    return hmac.new(_get_hmac_secret(), msg, hashlib.sha256).hexdigest()


def _verify_token(
    deal_name: str, stage_label: str, payload: Any, presented_token: str,
) -> bool:
    """Recompute the HMAC over (deal_name, stage_label, h(payload)) and
    compare against the presented token in constant time. Mismatch / empty
    presented → False."""
    if not presented_token:
        return False
    expected = _sign_token(deal_name, stage_label, payload)
    return hmac.compare_digest(expected, presented_token)


def _verify_subset_approval(
    *,
    deal_name: str,
    stage_label: str,
    approved: Any,
    proposed: Any,
    token: str,
) -> tuple[bool, Optional[str]]:
    """HMAC + subset verification for #21-comps-step-3 narrow-without-refire.

    When ``proposed`` is ``None`` → back-compat exact-match: the token is
    verified over ``approved`` directly (i.e. the operator echoed the bridge's
    proposal verbatim — pre-Step-3 contract, never breaks).

    When ``proposed`` is supplied → the token is verified over ``proposed``
    (the bridge's signed full universe), then ``approved`` MUST be a SUBSET
    of ``proposed``. The operator can DROP items but never ADD — additions
    would let a caller smuggle unsigned data past the gate, defeating the
    HMAC's anti-forgery guarantee. Subset preserves every existing safety
    property (HMAC anti-forgery, cross-stage replay defence via
    ``stage_label`` isolation).

    Supported shapes: flat list (subsectors) and dict-of-lists
    (peers_by_subsector / deals_by_subsector).

    Returns ``(ok, error_message)``. ``error_message`` is ``None`` on success
    and a short human-readable string on failure (surfaced to the caller via
    the route's 422).
    """
    universe = proposed if proposed is not None else approved
    if not _verify_token(deal_name, stage_label, universe, token):
        return False, "HMAC mismatch against signed payload"
    if proposed is None:
        return True, None
    # Subset enforcement — shape-aware so peers/deals (dict-of-lists) and
    # subsectors (flat list) compose under the same helper. MULTISET semantics
    # (Counter, not set) so "drop-only" is STRICT: approving MORE copies of an
    # item than were proposed is rejected, not silently collapsed (codex-5.5).
    # Unhashable / nested items (the contract is flat-list-of-str /
    # dict-of-lists-of-str) fail CLOSED with a clean message — never a 500.
    try:
        if isinstance(approved, list) and isinstance(proposed, list):
            extras = Counter(approved) - Counter(proposed)
            if extras:
                return False, (
                    f"approved contains items not in proposed (or more copies): "
                    f"{sorted(extras)}"
                )
        elif isinstance(approved, dict) and isinstance(proposed, dict):
            for key, items in approved.items():
                if key not in proposed:
                    return False, f"approved key {key!r} not in proposed"
                extras = Counter(items) - Counter(proposed[key])
                if extras:
                    return False, (
                        f"approved[{key!r}] contains items not in proposed: "
                        f"{sorted(extras)}"
                    )
        else:
            return False, (
                f"approved/proposed shape mismatch: "
                f"{type(approved).__name__} vs {type(proposed).__name__}"
            )
    except TypeError as e:
        return False, f"unsupported approved/proposed item shape: {e}"
    return True, None


def _assumption_identity(assumptions: list[dict[str, Any]]) -> list[list[str]]:
    """Stable identity of an LFY+1 assumption set: the sorted
    ``(subsector_slug, ticker, field)`` triples.

    The assumptions token signs THIS (not the full dicts) so the operator can
    add their decision (``choice`` / ``growth_rate`` / ``justification``) to
    each proposed assumption WITHOUT invalidating the token — the token only
    guarantees the operator is deciding the SAME set of (subsector, ticker,
    field) assumptions Stage 2 surfaced. The DECISION content is validated
    separately by ``_apply_approved_assumptions`` (Iron Law: un-justified
    operator growth is refused). Lists (not tuples) so the JSON canonical-hash
    is stable. #21-comps Q2-token follow-up, 2026-06-02.
    """
    return sorted(
        [str(a.get("subsector_slug") or ""), str(a.get("ticker") or ""),
         str(a.get("field") or "")]
        for a in assumptions
    )


# ─────────────────────────────────────────────────────────────────────────────
# WS-C: Stage-2 blocks cache (#21-comps-hardening).
#
# Composition with WS-B: the cache key IS the HMAC approval token signed over
# (deal_name, STAGE_LABEL_STAGE_2_BLOCKS, h(stage_2_blocks)). A tampered
# cache entry would mismatch on the existing HMAC verify path (defence in
# depth — the integrity binding is baked into the key itself).
#
# In-process dict with TTL eviction on read. Re-fire of Stage 3 with the
# Stage-2 token => look up here => NO provider re-call between Stage 2
# approval and Stage 3 stamp. Cache miss (eviction / process restart) => 422
# with "stage-2 cache miss; re-run Stage 2" message (the route surfaces).
# ─────────────────────────────────────────────────────────────────────────────

_STAGE_2_CACHE_TTL_SECONDS = 3600  # 1h — sensible balance between operator
                                    # review time + bounded growth.
_STAGE_2_CACHE_MAX_ENTRIES = 64     # hard cap so abandoned tokens can't grow
                                    # the cache for the process lifetime.
_STAGE_2_CACHE: dict[str, tuple[float, list[dict[str, Any]]]] = {}
_STAGE_2_CACHE_LOCK = threading.Lock()


def _cache_stage_2_blocks(token: str, blocks: list[dict[str, Any]]) -> None:
    """Store ``blocks`` in the in-process cache keyed by the HMAC approval
    ``token`` with current-time stamp.

    Hardening (codex-5.5): (1) DEEP-COPY on write — the read path already
    deep-copies, but storing ``blocks`` by reference let a caller mutating the
    original AFTER caching corrupt the signed entry; (2) prune TTL-expired
    entries + (3) cap the cache at ``_STAGE_2_CACHE_MAX_ENTRIES`` (evict oldest)
    so abandoned tokens — never re-read, so never evicted on the read path —
    can't grow the cache unbounded for the process lifetime."""
    now = time.monotonic()
    with _STAGE_2_CACHE_LOCK:
        # Prune anything past its TTL on the way in.
        for stale in [
            t for t, (ts, _) in _STAGE_2_CACHE.items()
            if now - ts > _STAGE_2_CACHE_TTL_SECONDS
        ]:
            _STAGE_2_CACHE.pop(stale, None)
        # Cap size — evict the oldest live entry if still at the ceiling.
        if len(_STAGE_2_CACHE) >= _STAGE_2_CACHE_MAX_ENTRIES and token not in _STAGE_2_CACHE:
            oldest = min(_STAGE_2_CACHE.items(), key=lambda kv: kv[1][0])[0]
            _STAGE_2_CACHE.pop(oldest, None)
        _STAGE_2_CACHE[token] = (now, copy.deepcopy(blocks))


def _get_cached_stage_2_blocks(token: str) -> Optional[list[dict[str, Any]]]:
    """Return a DEEP COPY of the cached blocks for ``token`` if present +
    not TTL-expired, else ``None``. TTL expiry triggers eviction of the
    entry on the read path so the cache doesn't grow unbounded.

    Deep-copy is load-bearing for cache integrity (codex-review finding
    SEV-2, 2026-06-03): Stage 3's ``_apply_approved_assumptions`` mutates
    the returned ``blocks`` in place (sets ``revenue_lfy1_m``,
    ``ebitdaal_source`` etc. on rows). Without the deep-copy, the cache
    entry itself was being mutated — so a Stage-3 retry or any subsequent
    HMAC verify of the cached blocks against the stage_2_blocks_token
    would fail with "cache integrity violation" because
    ``canonical_hash(mutated_blocks) != canonical_hash(original_blocks)``.
    The cache entry now stays byte-identical to what Stage 2 signed."""
    with _STAGE_2_CACHE_LOCK:
        entry = _STAGE_2_CACHE.get(token)
        if entry is None:
            return None
        stored_at, blocks = entry
        if time.monotonic() - stored_at > _STAGE_2_CACHE_TTL_SECONDS:
            # TTL expired — evict + return None.
            _STAGE_2_CACHE.pop(token, None)
            return None
        # Deep-copy so the caller can freely mutate the returned structure
        # without invalidating the cache's HMAC integrity guarantee.
        return copy.deepcopy(blocks)


def _clear_stage_2_cache() -> None:
    """Test-only helper: wipe the cache (simulate process restart)."""
    with _STAGE_2_CACHE_LOCK:
        _STAGE_2_CACHE.clear()


# ─────────────────────────────────────────────────────────────────────────────
# Exceptions — the route maps these to HTTP codes.
# ─────────────────────────────────────────────────────────────────────────────


class CompsSkillError(Exception):
    """Base for all comps-skill failures."""


class MissingApprovalToken(CompsSkillError):
    """Stage 1+ called without the prior-stage approval token (or with a
    stale/forged token). The route maps to HTTP 422 — the gate IS the contract."""


class UnsourcedFigureError(CompsSkillError):
    """A populated row entered Stage 3 without a Source. The Iron Law's
    pre-stamp guard refuses the workbook write. Route maps to HTTP 422."""


class TargetBriefMissing(CompsSkillError):
    """Stage 0 couldn't read the target brief / profile to propose subsectors.
    Route maps to HTTP 422."""


class TemplateStampFailed(CompsSkillError):
    """The xlsx stamp engine refused (template-shape drift, missing block-flex
    helper). Route maps to HTTP 502."""


# ─────────────────────────────────────────────────────────────────────────────
# Pydantic IO models
# ─────────────────────────────────────────────────────────────────────────────


# A literal taxonomy of currency-mismatch flag kinds — keeps the Source cell
# guard fully deterministic in the assertion.
CcyFlagKind = Literal["fs_vs_trading", "unit_p_vs_pound", "period_mismatch"]
# A literal taxonomy of the source-form prefixes the Iron Law's mechanical
# test recognises as "real" Sources. Anything outside this set (or a blank)
# is treated as unsourced.
_ACCEPTED_SOURCE_PREFIXES = (
    "http://", "https://",
    "tracker:",
    "src:",
    "operator-approved:",
)


def _looks_sourced(s: Optional[str]) -> bool:
    """A source string is "real" if it starts with a URL, a tracker/src
    reference, OR an operator-approved marker, OR matches ``<provider>:<as_of>``
    (any non-empty provider tag followed by ':' and a date-ish suffix)."""
    if not s or not isinstance(s, str):
        return False
    s2 = s.strip()
    if not s2:
        return False
    if s2.startswith(_ACCEPTED_SOURCE_PREFIXES):
        return True
    # `<provider>:<as_of_date>` form — e.g. `openbb-yfinance:2026-06-01`.
    # Require a colon + at least 4 chars on each side (provider name + date).
    if ":" in s2:
        left, right = s2.split(":", 1)
        if left.strip() and right.strip() and len(left.strip()) >= 3 and len(right.strip()) >= 4:
            return True
    return False


class CompsBuildInput(BaseModel):
    """Inputs to the comps-build skill.

    Each stage may be called individually; ``stage`` selects which Stage to
    run (0/1/2/3). Approval tokens from prior stages must be supplied for the
    later stages — the route refuses without them."""

    deal_name: str = Field(..., min_length=1, max_length=64, pattern=r"^[A-Za-z0-9_][A-Za-z0-9 _-]*$")
    target: str = Field(..., min_length=1, max_length=128)
    parent_sector: str = Field(..., min_length=1, max_length=64,
                               description="Lowercased-hyphenated sector slug, e.g. 'hospitality'")

    stage: Literal[0, 1, 2, 3] = 0

    # Routing + workspace context (NOT comps inputs; consumed by the central guards).
    workspace_type: Literal["project", "bd", "general"] = "project"
    workspace_name: str = "default"
    workspace_sensitivity: Literal["public", "internal", "confidential", "MNPI"] = "internal"

    # Approval payloads — set by the operator on re-fire.
    approved_subsectors: Optional[list[str]] = None              # Stage 1 entry — the operator-signed subsector list
    approved_peers_by_subsector: Optional[dict[str, list[str]]] = None    # Stage 2 entry — {subsector_slug: [ticker, ...]}
    approved_deals_by_subsector: Optional[dict[str, list[str]]] = None    # Stage 2 entry — {subsector_slug: [deal_id, ...]}
    approved_assumptions: Optional[list[dict[str, Any]]] = None  # Stage 3 entry — list of {"field": "...", "approved_value": "..."}

    # Path-A orchestration — operator-submitted candidate universe (#21-comps
    # orchestration, 2026-06-02). The Claude session that orchestrates the
    # comps build runs the Anthropic Skills (equity-research:screen /
    # investment-banking:buyer-list / deep-research) itself, then SUBMITS the
    # resulting candidates into Stage 1. ``run_stage_1`` merges these with the
    # in-process shim/provider output (deduped by ticker / deal-key) BEFORE it
    # proposes + HMAC-signs, so the operator approves a universe that includes
    # the attended session's research. Empty / absent → today's behaviour (the
    # shims return [] and only provider peers + tracker deals are proposed).
    # NOT consumed by any later stage — purely a Stage-1 candidate input.
    #   coco shape:    {subsector_slug: [{"ticker","name","country","why"}, ...]}
    #   cotrans shape: {subsector_slug: [{"announced_date","target","acquirer",
    #                   "country","currency","ev_m","ev_revenue_x","ev_ebitda_x",
    #                   "source", ...}, ...]} — flows through the same tracker
    #                   write-back + dedup path as deep-research output.
    submitted_coco_candidates_by_subsector: Optional[dict[str, list[dict[str, Any]]]] = None
    submitted_cotrans_candidates_by_subsector: Optional[dict[str, list[dict[str, Any]]]] = None

    # #21-comps-step-3 (2026-06-03) — Subset-approval support. The HMAC token
    # signs the bridge's FULL proposal at each stage. By default the operator
    # must approve verbatim. To NARROW WITHOUT RE-FIRING the prior stage, the
    # operator echoes the bridge's full proposed_* set + a strict subset of it
    # via the existing approved_* fields. Verification:
    #   • token must HMAC-match proposed_* (the bridge's signed universe)
    #   • approved_* must be ⊆ proposed_* (operator can DROP, never ADD —
    #     additions would smuggle unsigned data past the gate)
    # Back-compat: when proposed_* is absent, verification falls back to
    # exact-match against approved_* (existing single-shot callers unchanged).
    proposed_subsectors: Optional[list[str]] = None
    proposed_peers_by_subsector: Optional[dict[str, list[str]]] = None
    proposed_deals_by_subsector: Optional[dict[str, list[str]]] = None

    # WS-B: HMAC approval tokens — load-bearing on re-fire.
    # Each token is signed by the prior stage over (deal_name, stage_label,
    # canonical-hash(approved-payload)). Verification recomputes the HMAC
    # from the SUPPLIED approved payload + presented token; mismatch / missing
    # / cross-stage replay → MissingApprovalToken → HTTP 422.
    subsectors_approval_token: Optional[str] = None              # Stage 1 entry — signs approved_subsectors
    peers_approval_token: Optional[str] = None                   # Stage 2 entry — signs approved_peers_by_subsector
    deals_approval_token: Optional[str] = None                   # Stage 2 entry — signs approved_deals_by_subsector
    stage_2_blocks_approval_token: Optional[str] = None          # Stage 3 entry — signs Stage-2 blocks (cache key)
    assumptions_approval_token: Optional[str] = None             # Stage 3 entry — signs approved_assumptions (optional)

    # Operator overrides (small N — most operator decisions land via the
    # approval-pending re-fire loop).
    today: Optional[str] = None       # ISO date; tests pin this for determinism


# ── Stage outputs ───────────────────────────────────────────────────────────


class ApprovalPayload(BaseModel):
    """The blob the operator signs to advance to the next stage."""
    kind: Literal["subsectors", "peers_and_deals", "assumptions"]
    proposed: Any        # shape depends on kind — see references/ for the schemas
    rationale: Optional[dict[str, Any]] = None
    tracker_writes_planned: Optional[list[dict[str, Any]]] = None


class StageResult(BaseModel):
    """What every Stage returns. ``stage='complete'`` only on a successful
    Stage 3; otherwise ``stage='approval_pending'`` carrying the payload the
    operator signs to advance."""
    ok: bool = True
    stage: Literal["approval_pending", "complete"]
    stage_just_completed: Optional[int] = None
    deal_name: str
    target: Optional[str] = None
    run_id: str
    approval_payload: Optional[ApprovalPayload] = None
    approval_token_to_sign: Optional[str] = None
    warnings: list[str] = Field(default_factory=list)

    # WS-B: per-token surfacing — the bridge UI extracts these to pass back
    # at the next stage. ``approval_token_to_sign`` stays as the combined /
    # primary token for backward-compat (Stage 0 / Stage 2 nothing-to-approve);
    # Stage 1 + Stage 2-with-assumptions surface multiple tokens.
    subsectors_approval_token: Optional[str] = None      # issued by Stage 0
    peers_approval_token: Optional[str] = None           # issued by Stage 1
    deals_approval_token: Optional[str] = None           # issued by Stage 1
    stage_2_blocks_approval_token: Optional[str] = None  # issued by Stage 2 (the cache key)
    assumptions_approval_token: Optional[str] = None     # issued by Stage 2 (when assumptions pending)

    # Stage 3 deliverable fields (only populated when stage='complete'):
    approved_subsectors: Optional[list[str]] = None
    blocks: Optional[list[dict[str, Any]]] = None
    headline_ev_ebitda_median: Optional[float] = None
    headline_ev_revenue_median: Optional[float] = None
    peer_count: Optional[int] = None
    deal_count: Optional[int] = None
    as_of: Optional[str] = None
    provider: Optional[str] = None
    template_path: Optional[str] = None
    prior_archived_path: Optional[str] = None
    mirror_refresh_path: Optional[str] = None
    tracker_writes: list[dict[str, Any]] = Field(default_factory=list)
    iron_law_assertion: Optional[dict[str, Any]] = None


# ─────────────────────────────────────────────────────────────────────────────
# Anthropic-cloud-skill orchestration shims (mockable in tests).
#
# Each shim returns a STRUCTURED dict that the stage code consumes. Real-world
# implementations bind to the Anthropic cloud skill APIs; tests monkeypatch
# these to return deterministic payloads. The sensitivity check fires BEFORE
# any shim is invoked (route layer + this module's _sensitivity_check helper).
# ─────────────────────────────────────────────────────────────────────────────


def _sensitivity_check(sensitivity: str) -> None:
    """Hard guard before any cloud-skill orchestration. MNPI must never reach
    a cloud skill (CLAUDE.md §5.2). Internal/public/confidential pass — the
    route's central guard handles workspace-scope refusal upstream of this.

    Confidential is permitted here for cloud-skill orchestration only when
    the comps-build skill itself is internal (the declared sensitivity in
    the SKILL.md frontmatter) and the WORKSPACE happens to be confidential
    — that's the cross-skill case handled by the central guard, which would
    have refused earlier. By the time we reach this shim, the effective
    sensitivity is the SKILL's (internal), not the workspace's."""
    if sensitivity == "MNPI":
        raise CompsSkillError(
            f"Sensitivity={sensitivity!r}: MNPI inputs are never permitted "
            f"on comps-build (cloud-skill orchestration is on the call path)."
        )


def _shim_equity_research_screen(
    subsector_slug: str, target: str, *, sensitivity: str,
) -> list[dict[str, Any]]:
    """Orchestrate Anthropic's `equity-research:screen` cloud skill. Returns
    candidate CoCo peers per subsector with (ticker, name, country, why).
    Tests mock this to return deterministic candidates."""
    _sensitivity_check(sensitivity)
    # Real impl: HTTP call to the Anthropic Agent SDK; out of scope for the
    # initial wiring. The skill returns a STRUCTURED candidate list; we leave
    # this empty so tests must monkeypatch to exercise the Stage 1 path.
    log.info("comps: equity-research:screen orchestrated for %s/%s (sensitivity=%s)",
             subsector_slug, target, sensitivity)
    return []


def _shim_investment_banking_buyer_list(
    subsector_slug: str, target: str, *, sensitivity: str,
) -> list[dict[str, Any]]:
    """Orchestrate Anthropic's `investment-banking:buyer-list` cloud skill.
    Returns candidate strategic + sponsor buyers; useful as a peer
    cross-check. Tests mock this."""
    _sensitivity_check(sensitivity)
    log.info("comps: investment-banking:buyer-list orchestrated for %s/%s",
             subsector_slug, target)
    return []


def _shim_deep_research_cotrans(
    subsector_slug: str, lookback_years: int, *, sensitivity: str,
) -> list[dict[str, Any]]:
    """Orchestrate Anthropic's `deep-research` cloud skill for CoTrans gap-
    fill. Returns deals not already in the tracker; the caller writes them
    back via append_deal()."""
    _sensitivity_check(sensitivity)
    log.info("comps: deep-research orchestrated for cotrans %s (last %dy)",
             subsector_slug, lookback_years)
    return []


# ─── Block-flex stamping (openpyxl) ──────────────────────────────────────────
#
# The v2 template (operator-controlled at os-templates/Project_x_Comps_v2.xlsx)
# ships with ONE CoCo block + ONE CoTrans block as a unit. The skill must stamp
# ONE CoCo block PER APPROVED SUBSECTOR. We can't pre-compute this from the
# template's range-formula structure at engine-startup time (the operator
# adjusts the template layout outside our control), so block-flex is
# implemented at the openpyxl layer: each approved-subsector block is written
# as a STANDALONE rectangle of cells with its own per-block Mean/Median/75th/
# 25th/Min/Max range formulas referencing only that block's data rows.
#
# Layout (one CoCo block on a single "CoCo" worksheet, blocks stacked
# vertically with a 2-row gutter):
#
#   row 1:    "<subsector_slug>" merged-A1 banner
#   row 2:    column headers (Ticker, Company, CCY, Market Cap, Net Debt,
#             EV, Revenue LFY, EBITDA LFY, EV/Revenue, EV/EBITDA, Source)
#   row 3..N: one data row per CoCo peer (N varies per subsector)
#   row N+1:  Mean    | (formulas re-based to this block's data rows)
#   row N+2:  Median  | (formulas re-based to this block's data rows)
#   row N+3:  75th    | (...)
#   row N+4:  25th    | (...)
#   row N+5:  Min     | (...)
#   row N+6:  Max     | (...)
#   row N+8:  next block starts (2-row gutter)
#
# CoTrans rows from each block stack on a second "CoTrans" worksheet using
# the same banner-then-rows pattern but without the stats footer (deals are
# point-in-time observations).
#
# Output: a real .xlsx the operator can open in Excel. Tests inspect via
# openpyxl to count blocks (per subsector) + assert per-block stats formulas
# reference only that block's range.


def _coco_block_columns() -> tuple[str, ...]:
    """The column headers of a CoCo block. Order is load-bearing — the stats
    row formulas address fixed column letters."""
    return (
        "Ticker", "Company", "CCY",
        "Market Cap (m)", "Net Debt (m)", "EV (m)",
        "Revenue LFY (m)", "EBITDA LFY (m)",
        "EV / Revenue", "EV / EBITDA",
        "Source",
    )


def _cotrans_block_columns() -> tuple[str, ...]:
    """The column headers of a CoTrans block."""
    return (
        "Announced", "Target", "Acquirer", "Country", "Description",
        "EV (m)", "EV / Revenue", "EV / EBITDA",
        "Strategic Commentary", "Source",
    )


def _write_coco_block(ws, start_row: int, block: dict[str, Any]) -> int:
    """Write one CoCo block at ``start_row``. Returns the row index AFTER
    the last stats row (so the next block knows where to start).

    Block structure:
      banner row → header row → data rows → 6 stats rows.

    The stats row formulas reference the block's data range only — re-based
    per block so the 6 medians/quartiles reflect the subsector, NOT the
    full sheet (the bug the operator review flagged).
    """
    from openpyxl.utils import get_column_letter

    ss = str(block.get("subsector_slug") or "unknown")
    coco_rows = list(block.get("coco_rows") or [])
    cols = _coco_block_columns()

    # ── 1. banner row ─────────────────────────────────────────────────────
    ws.cell(row=start_row, column=1, value=f"CoCo: {ss}")

    # ── 2. header row ─────────────────────────────────────────────────────
    header_row = start_row + 1
    for ci, name in enumerate(cols, start=1):
        ws.cell(row=header_row, column=ci, value=name)

    # ── 3. data rows ──────────────────────────────────────────────────────
    first_data = header_row + 1
    for ri, row in enumerate(coco_rows):
        rnum = first_data + ri
        mc = row.get("market_cap_m")
        nd = row.get("net_debt_m")
        ev = (mc + nd) if (mc is not None and nd is not None) else None
        rev = row.get("revenue_lfy_m")
        ebitda = row.get("ebitda_lfy_m")
        ev_rev = (ev / rev) if (ev is not None and rev) else None
        ev_eb = (ev / ebitda) if (ev is not None and ebitda) else None
        ws.cell(row=rnum, column=1, value=row.get("ticker"))
        ws.cell(row=rnum, column=2, value=row.get("name"))
        ws.cell(row=rnum, column=3, value=row.get("currency"))
        ws.cell(row=rnum, column=4, value=mc)
        ws.cell(row=rnum, column=5, value=nd)
        ws.cell(row=rnum, column=6, value=ev)
        ws.cell(row=rnum, column=7, value=rev)
        ws.cell(row=rnum, column=8, value=ebitda)
        ws.cell(row=rnum, column=9, value=ev_rev)
        ws.cell(row=rnum, column=10, value=ev_eb)
        ws.cell(row=rnum, column=11, value=row.get("source"))
    last_data = first_data + len(coco_rows) - 1 if coco_rows else first_data

    # ── 4. stats rows (re-based per block) ────────────────────────────────
    # Only EV/Revenue (col I) and EV/EBITDA (col J) get stats — the columns
    # that drive the median headline. If a block has zero data rows, leave
    # the stats cells blank (avoid #DIV/0! noise).
    stats_first = last_data + 1
    stats_labels = (
        ("Mean", "AVERAGE"),
        ("Median", "MEDIAN"),
        ("75th",   "PERCENTILE"),
        ("25th",   "PERCENTILE"),
        ("Min",    "MIN"),
        ("Max",    "MAX"),
    )

    # Range arguments for the data block (columns I + J = EV/Revenue + EV/EBITDA).
    ev_rev_range = (
        f"{get_column_letter(9)}{first_data}:{get_column_letter(9)}{last_data}"
        if coco_rows else ""
    )
    ev_eb_range = (
        f"{get_column_letter(10)}{first_data}:{get_column_letter(10)}{last_data}"
        if coco_rows else ""
    )

    for i, (label, fn) in enumerate(stats_labels):
        rnum = stats_first + i
        ws.cell(row=rnum, column=2, value=label)
        if coco_rows:
            if fn == "PERCENTILE":
                pct = 0.75 if label == "75th" else 0.25
                ws.cell(row=rnum, column=9, value=f"=PERCENTILE({ev_rev_range},{pct})")
                ws.cell(row=rnum, column=10, value=f"=PERCENTILE({ev_eb_range},{pct})")
            else:
                ws.cell(row=rnum, column=9, value=f"={fn}({ev_rev_range})")
                ws.cell(row=rnum, column=10, value=f"={fn}({ev_eb_range})")

    # Next block starts 2 rows later (1-row gutter).
    return stats_first + len(stats_labels) + 1


def _write_cotrans_block(ws, start_row: int, block: dict[str, Any]) -> int:
    """Write one CoTrans block at ``start_row``. Returns the next free row.

    No stats footer — CoTrans are point-in-time observations, not population.
    """
    ss = str(block.get("subsector_slug") or "unknown")
    rows = list(block.get("cotrans_rows") or [])
    cols = _cotrans_block_columns()
    ws.cell(row=start_row, column=1, value=f"CoTrans: {ss}")
    header_row = start_row + 1
    for ci, name in enumerate(cols, start=1):
        ws.cell(row=header_row, column=ci, value=name)
    for ri, row in enumerate(rows):
        rnum = header_row + 1 + ri
        ws.cell(row=rnum, column=1, value=row.get("announced_date"))
        ws.cell(row=rnum, column=2, value=row.get("target"))
        ws.cell(row=rnum, column=3, value=row.get("acquirer"))
        ws.cell(row=rnum, column=4, value=row.get("country"))
        ws.cell(row=rnum, column=5, value=row.get("description"))
        ws.cell(row=rnum, column=6, value=row.get("ev_m"))
        ws.cell(row=rnum, column=7, value=row.get("ev_revenue"))
        ws.cell(row=rnum, column=8, value=row.get("ev_ebitda"))
        ws.cell(row=rnum, column=9, value=row.get("strategic_commentary"))
        ws.cell(row=rnum, column=10, value=row.get("source"))
    last_data = header_row + len(rows) if rows else header_row
    return last_data + 2   # 1-row gutter before the next block


# ─── Template block geometry (v2 layout, operator-restructured 2026-06-02) ──
#
# The v2 template ships ONE CoCo block + ONE CoTrans block per sheet, with
# notes/methodology rows at the bottom. Block geometry (NEW layout):
#
#   row 1:       "Prices as of:" label at A1; date stamp at B1 of each block
#                (B1, B23, B44, ... after row-offset replication).
#   row 2:       banner "COMPARABLE COMPANIES (CoCo) — [Subsector]" at B2
#   row 3:       sub-heads (Shr. / Revenue / EBITDA / EV/EBITDA / EV/Revenue / P/E)
#   row 4:       period headers (Ticker / Company / CCY (FS) / price /
#                outstanding / Mkt Cap / Net Debt / EV / YE / LFY / LFY+1 / ...)
#   rows 5-12:   8 pre-formatted data rows. Formula cells:
#                  G = =+E*F           (Mkt Cap)
#                  I = =IF(COUNT(G:H)=0,"",SUM(G:H))   (EV)
#                  S/T/U = EV / EBITDA(LFY-1/LFY/LFY+1)
#                  W/X   = EV / Revenue(LFY/LFY+1)
#                  AD    = FX? flag
#   row 13:      gap (default; flexes when N peers > 8)
#   rows 14-19:  stats rows (Mean / Median / 75th / 25th / Min / Max) on the
#                multiple cols S/T/U/W/X/Z — re-targeted per block.
#   rows 22-28:  NOTES & METHODOLOGY (preserved once at the bottom).
#
# Re-stamping N blocks: keep block 1 in place (rows 1-19 modified in place);
# copy rows 1-19 (cached BEFORE block-1 mutations) per additional block at
# offset (BLOCK_HEIGHT + GUTTER) * i; then re-stamp the notes/methodology rows
# at the very bottom of the sheet. When a block has N peers > 8, additional
# data rows are inserted AFTER row 12 (relative) and the stats footer formula
# ranges are extended to cover them — see ``_flex_block_for_peers``.

_TPL_BLOCK_FIRST_ROW = 1
_TPL_BLOCK_LAST_ROW = 19
_TPL_DATA_FIRST_ROW = 5
_TPL_DATA_LAST_ROW = 12
_TPL_STATS_FIRST_ROW = 14
_TPL_STATS_LAST_ROW = 19
_TPL_BANNER_ROW = 2
_TPL_DATE_ROW = 1   # B1 gets the date stamp (column 2)
_TPL_DATE_COL = 2
_TPL_NOTES_FIRST_ROW = 22
_TPL_NOTES_LAST_ROW = 28
_BLOCK_HEIGHT = _TPL_BLOCK_LAST_ROW - _TPL_BLOCK_FIRST_ROW + 1   # = 19
_BLOCK_GUTTER = 2
_BLOCK_STRIDE = _BLOCK_HEIGHT + _BLOCK_GUTTER                    # = 21 (default)

# Banner substitution marker: the template uses literal "[Subsector]" so the
# stamper finds-and-replaces with the actual slug per block.
_BANNER_SUBSECTOR_MARKER = "[Subsector]"

# ─── Column resolver (header-name-aware) ─────────────────────────────────────
#
# Reads the block's row-3 sub-head + row-4 period labels and returns a mapping
# of logical names → 1-indexed column numbers. The mapping is stable across
# minor operator edits to the template layout (re-shuffled columns / added
# spacers): the stamping code looks up positions by logical name, never by
# hardcoded letter.
#
# Sub-head row (row 3 relative to the block) groups the periods:
#   E/F  -> "Shr." (the share price + shares outstanding pair)
#   L    -> "Revenue"  (covers L=LFY, M=LFY+1)
#   P    -> "EBITDA"   (covers O=LFY-1, P=LFY, Q=LFY+1)
#   S    -> "EV / EBITDA" (covers S=LFY-1, T=LFY, U=LFY+1)
#   W    -> "EV / Revenue" (covers W=LFY, X=LFY+1)
#   Z    -> "P / E"    (covers Z=LTM)
#
# Period-head row (row 4) carries:
#   Ticker / Company / CCY (FS) / price / outstanding / Mkt Cap / Net Debt /
#   EV / YE / [Rev: LFY, LFY+1] / [EBITDA: LFY-1, LFY, LFY+1] /
#   [EV/EBITDA: LFY-1, LFY, LFY+1] / [EV/Revenue: LFY, LFY+1] / LTM /
#   Source / CCY (Px) / FX?


def _resolve_coco_columns(ws, block_row_offset: int = 0) -> dict[str, int]:
    """Resolve logical CoCo column names → 1-indexed column numbers from the
    BLOCK'S header rows. ``block_row_offset`` is added to the template's
    relative rows 3+4 (so block 2 reads rows 24+25 etc.).

    The resolver is forgiving: it walks every cell in the two header rows and
    builds the mapping by tagging each period-head with the most recent
    sub-head group seen on row 3. This handles spacer columns (gap K=11, N=14,
    R=18, etc.) gracefully — they simply produce no entry.
    """
    row_subhead = _TPL_BLOCK_FIRST_ROW + 2 + block_row_offset   # = 3
    row_period = _TPL_BLOCK_FIRST_ROW + 3 + block_row_offset    # = 4
    max_col = ws.max_column

    # 1. Locate each group's anchor column in row 3. Each anchor's group
    # covers row-4 period labels in a small window around the anchor — the
    # operator's restructured layout sometimes places "LFY-1" ONE COLUMN
    # LEFT of an EBITDA / EV/EBITDA anchor (e.g. O is LFY-1 EBITDA but the
    # anchor "EBITDA" sits at P). The window-based assignment handles this
    # robustly without depending on forward-fill ordering.
    #
    # Each entry: (anchor_col, group_name). Windows are computed by sorting
    # anchors L→R and assigning the half-open span [previous_anchor+0,
    # next_anchor) to each anchor — anchors are at the LEFTMOST OR INSIDE
    # the group's span. To handle "LFY-1 sits LEFT of anchor", we extend each
    # anchor's window 2 columns LEFT (capped by the prior anchor's window).
    anchors: list[tuple[int, str]] = []
    for c in range(1, max_col + 1):
        v = ws.cell(row=row_subhead, column=c).value
        if not isinstance(v, str) or not v.strip():
            continue
        s = v.strip().lower()
        if "/" in s:
            # NB: check "ebitdaal" before "ebitda" — the EV/EBITDAaL group is
            # distinct from EV/EBITDA (lease-adjusted vs IFRS-16 basis).
            if "ebitdaal" in s:
                anchors.append((c, "ev_ebitdaal"))
            elif "ebitda" in s:
                anchors.append((c, "ev_ebitda"))
            elif "revenue" in s:
                anchors.append((c, "ev_rev"))
            elif s.split("/")[0].strip() == "p":
                anchors.append((c, "pe"))
        elif s.startswith("shr"):
            anchors.append((c, "shr"))
        elif s == "ye" or s.startswith("year") or s == "fye":
            anchors.append((c, "ye"))            # v2: row-3 "YE" + row-4 "date"
        elif s.startswith("net debt"):
            anchors.append((c, "net_debt"))     # v2 lease split: H/I (excl/incl)
        elif s == "ev":
            anchors.append((c, "ev"))            # v2 EV split: J/K (excl/incl) — formula cols
        elif s.startswith("ebitdaal"):
            anchors.append((c, "ebitdaal"))      # check before "ebitda"
        elif s == "revenue" or s.startswith("revenue"):
            anchors.append((c, "rev"))
        elif s == "ebitda" or s.startswith("ebitda"):
            anchors.append((c, "ebitda"))
    anchors.sort()
    # Build group[col] mapping. Each anchor's primary span is [acol,
    # next_anchor-1]. Additionally, if a row-4 period label "LFY-1" appears
    # in the 1-2 cols IMMEDIATELY LEFT of the anchor (and that left col
    # isn't claimed by the prior anchor's primary span), extend ownership
    # to it — this catches the EBITDA / EV-EBITDA case where "LFY-1" sits
    # left of the anchor in the operator's restructured layout.
    group: dict[int, str] = {}
    for i, (acol, gname) in enumerate(anchors):
        prev_acol = anchors[i - 1][0] if i > 0 else 0
        next_start = anchors[i + 1][0] if i + 1 < len(anchors) else max_col + 1
        # Primary span: [acol, next_start - 1].
        for c in range(acol, next_start):
            group[c] = gname
        # Look back up to 2 cols for a "LFY-1" period label that belongs to
        # this anchor's group. Stop at the prior anchor (don't poach).
        for look in range(1, 3):
            c = acol - look
            if c <= prev_acol:
                break
            v = ws.cell(row=row_period, column=c).value
            if isinstance(v, str) and v.strip().lower() in ("lfy-1", "lfy -1"):
                group[c] = gname

    # 2. Walk row 4 and emit logical names per (group, period) pair.
    out: dict[str, int] = {}
    for c in range(1, max_col + 1):
        v = ws.cell(row=row_period, column=c).value
        if not isinstance(v, str):
            continue
        label = v.strip()
        if not label:
            continue
        lbl_low = label.lower()
        grp = group.get(c, "")
        # Stand-alone period-head labels (no group on row 3).
        if lbl_low == "ticker":
            out["ticker"] = c
        elif lbl_low == "company":
            out["company"] = c
        elif lbl_low == "ccy (fs)" or lbl_low == "ccy(fs)":
            out["ccy_fs"] = c
        elif lbl_low == "mkt cap":
            out["mkt_cap"] = c
        elif lbl_low == "net debt":
            out["net_debt"] = c
        elif lbl_low == "ev":
            out["ev"] = c
        elif lbl_low in ("ye", "fye", "fy end", "year end"):
            out["fye"] = c
        elif lbl_low == "source":
            out["source"] = c
        elif lbl_low == "ccy (px)" or lbl_low == "ccy(px)":
            out["ccy_px"] = c
        elif lbl_low in ("fx?", "fx"):
            out["fx_flag"] = c
        # Grouped period-head labels.
        elif grp == "shr" and lbl_low == "price":
            out["shr_price"] = c
        elif grp == "shr" and lbl_low == "outstanding":
            out["shr_out"] = c
        elif grp == "rev" and lbl_low == "lfy":
            out["rev_lfy"] = c
        elif grp == "rev" and lbl_low == "lfy+1":
            out["rev_lfy1"] = c
        elif grp == "ebitda" and lbl_low == "lfy-1":
            out["ebitda_lfym1"] = c
        elif grp == "ebitda" and lbl_low == "lfy":
            out["ebitda_lfy"] = c
        elif grp == "ebitda" and lbl_low == "lfy+1":
            out["ebitda_lfy1"] = c
        elif grp == "ev_ebitda" and lbl_low == "lfy-1":
            out["ev_ebitda_lfym1"] = c
        elif grp == "ev_ebitda" and lbl_low == "lfy":
            out["ev_ebitda_lfy"] = c
        elif grp == "ev_ebitda" and lbl_low == "lfy+1":
            out["ev_ebitda_lfy1"] = c
        elif grp == "ev_rev" and lbl_low == "lfy":
            out["ev_rev_lfy"] = c
        elif grp == "ev_rev" and lbl_low == "lfy+1":
            out["ev_rev_lfy1"] = c
        elif grp == "pe" and lbl_low == "ltm":
            out["pe_ltm"] = c
        # v2 lease split — Net Debt + EV each carry "(excl. leases)" / "(incl.
        # leases)" period labels under their row-3 group. EV cols are formulas
        # (resolved for completeness; never written).
        elif grp == "net_debt" and _is_excl_leases(lbl_low):
            out["net_debt_excl"] = c
        elif grp == "net_debt" and _is_incl_leases(lbl_low):
            out["net_debt_incl"] = c
        elif grp == "ev" and _is_excl_leases(lbl_low):
            out["ev_excl"] = c
        elif grp == "ev" and _is_incl_leases(lbl_low):
            out["ev_incl"] = c
        # v2 EBITDAaL group — distinct from EBITDA (lease-adjusted earnings).
        elif grp == "ebitdaal" and lbl_low == "lfy-1":
            out["ebitdaal_lfym1"] = c
        elif grp == "ebitdaal" and lbl_low == "lfy":
            out["ebitdaal_lfy"] = c
        elif grp == "ebitdaal" and lbl_low == "lfy+1":
            out["ebitdaal_lfy1"] = c
        # v2 YE group: row-3 "YE", row-4 "date" → fiscal year-end.
        elif grp == "ye" and lbl_low in ("date", "ye", "fye", "fy end", "year end"):
            out["fye"] = c
    return out


def _is_excl_leases(lbl_low: str) -> bool:
    s = lbl_low.strip("() ").replace(".", "")
    return s in ("excl leases", "excl lease", "ex leases", "ex-leases", "excl. leases")


def _is_incl_leases(lbl_low: str) -> bool:
    s = lbl_low.strip("() ").replace(".", "")
    return s in ("incl leases", "incl lease", "inc leases", "incl. leases")


def _probe_template_safe(wb) -> list[str]:
    """Probe an openpyxl-loaded workbook for features openpyxl can't safely
    round-trip on save. Returns a list of human-readable descriptions; empty
    list = safe.

    Covers:
      * threaded comments (Excel 2016+ collab comments — openpyxl strips them)
      * legacy cell comments
      * conditional formatting rules (openpyxl preserves SOME but mutates the
        XML — we flag any presence so an operator can sanity-check)
      * data validations (openpyxl handles many but not all edge cases)
      * charts (CompletedChart-only support; not all chart types round-trip)
      * images

    This is the safety guard at stamp time. The brief's re-verification on
    2026-06-01 confirmed the live template at os-templates/Project_x_Comps_v2.xlsx
    has ZERO of these. If a FUTURE operator edit adds any of them,
    _shim_xlsx_stamp fails loud with the offending descriptions so the
    operator can either remove the feature or move to the
    anthropic-skills:xlsx cloud-side drop-in (documented as the future seam).
    """
    offending: list[str] = []
    for sn in wb.sheetnames:
        ws = wb[sn]
        # Threaded comments — openpyxl exposes them via ws._comments or
        # ws._threadedComments depending on version; cover both.
        try:
            for c in (getattr(ws, "_comments", None) or []):
                offending.append(
                    f"sheet {sn!r}: legacy comment at {getattr(c, 'parent', '?')}: {getattr(c, 'text', '?')!r}"
                )
        except Exception:  # noqa: BLE001
            pass
        try:
            tc = getattr(ws, "_threadedComments", None) or []
            for c in tc:
                offending.append(f"sheet {sn!r}: threaded comment {c!r}")
        except Exception:  # noqa: BLE001
            pass
        # Conditional formatting
        try:
            cf_count = len(list(ws.conditional_formatting))
            if cf_count:
                offending.append(
                    f"sheet {sn!r}: {cf_count} conditional-formatting rule(s) — "
                    f"openpyxl may not preserve all rule types"
                )
        except Exception:  # noqa: BLE001
            pass
        # Data validations
        try:
            dv = list(ws.data_validations.dataValidation)
            if dv:
                offending.append(
                    f"sheet {sn!r}: {len(dv)} data-validation rule(s) — "
                    f"openpyxl may not preserve list-source validations"
                )
        except Exception:  # noqa: BLE001
            pass
        # Charts
        try:
            charts = getattr(ws, "_charts", None) or []
            if charts:
                offending.append(
                    f"sheet {sn!r}: {len(charts)} chart(s) — openpyxl chart "
                    f"round-trip is partial"
                )
        except Exception:  # noqa: BLE001
            pass
        # Images
        try:
            imgs = getattr(ws, "_images", None) or []
            if imgs:
                offending.append(
                    f"sheet {sn!r}: {len(imgs)} embedded image(s) — openpyxl "
                    f"may drop them on save"
                )
        except Exception:  # noqa: BLE001
            pass
    return offending


# Pre-compiled regex for re-targeting formula cell-references onto a different
# block's row range. Matches "<col_letters><row_number>" (e.g. "E5", "AA12").
# Used to shift formulas from the template's block (rows 5-12, 14-19) onto a
# replicated block at a row offset.
_CELL_REF_RE = re.compile(r"(?<![A-Z])([A-Z]+)(\d+)")


def _shift_formula(formula: str, row_offset: int) -> str:
    """Shift all row numbers in a cell formula by ``row_offset``. Preserves
    column letters + non-cell-reference content. Skips $-absolute refs."""
    if not isinstance(formula, str) or not formula.startswith("="):
        return formula
    def _sub(m: re.Match) -> str:
        col, rownum = m.group(1), int(m.group(2))
        return f"{col}{rownum + row_offset}"
    return _CELL_REF_RE.sub(_sub, formula)


def _unmerge_in_block_region(ws, first_row: int, last_row: int) -> None:
    """Unmerge any merged ranges entirely inside [first_row..last_row].
    Required before writing to merged-cell regions (MergedCell.value is
    read-only)."""
    for mr in list(ws.merged_cells.ranges):
        if mr.min_row >= first_row and mr.max_row <= last_row:
            ws.unmerge_cells(str(mr))


def _stamp_one_coco_block(
    ws, *, first_row: int, subsector_slug: str, coco_rows: list[dict[str, Any]],
    template_cells: list[list[Any]], template_styles: list[list[Any]],
    template_merged: list[Any],
    today_iso: Optional[str] = None,
) -> int:
    """Stamp ONE CoCo block onto ``ws`` starting at ``first_row``.

    ``template_cells`` / ``template_styles`` carry the cached
    (value, style) of every cell in the template's block region (rows 1-19)
    BEFORE any in-place mutation — so block 2..N replicate the template's
    original layout, NOT the mutated block-1 contents.

    ``today_iso`` (when supplied) is stamped at the block's B1 (the operator's
    "Prices as of:" slot). Caller passes the same date for every block in a
    run so the date strip is consistent.

    Row-flex: when ``len(coco_rows) > 8`` (the template's pre-formatted data
    rows), additional data rows are inserted AFTER row 12 (relative to block)
    BEFORE the stats footer. Stats range formulas (col S/T/U/W/X/Z, rows
    14-19 of the block) are re-written to cover the extended range.

    Returns the next free row (after the block + gutter + flex insertions).
    """
    row_offset = first_row - _TPL_BLOCK_FIRST_ROW
    dst_first = _TPL_BLOCK_FIRST_ROW + row_offset
    dst_last = _TPL_BLOCK_LAST_ROW + row_offset
    extra = max(0, len(coco_rows) - (_TPL_DATA_LAST_ROW - _TPL_DATA_FIRST_ROW + 1))

    # Unmerge any pre-existing merges in the destination region (block 1's
    # destination is the live template merges; block 2..N's destination is
    # below the template's merges, so nothing to unmerge usually — but we
    # cover both with a single helper). Required because MergedCell.value
    # is read-only.
    _unmerge_in_block_region(ws, dst_first, dst_last)

    # ── 1. Replicate the template block region cell-by-cell (values + styles
    # + formulas, with formulas row-shifted onto the new block).
    for ri, row_vals in enumerate(template_cells):
        row_styles = template_styles[ri]
        tpl_row = _TPL_BLOCK_FIRST_ROW + ri
        dst_row = tpl_row + row_offset
        # When flexing (extra > 0), rows from tpl_row >= 13 (the stats footer +
        # gap) shift further down to accommodate the inserted data rows.
        if extra > 0 and tpl_row > _TPL_DATA_LAST_ROW:
            dst_row += extra
        for ci, val in enumerate(row_vals):
            col = ci + 1
            dst = ws.cell(row=dst_row, column=col)
            if isinstance(val, str) and val.startswith("="):
                # Re-target formula row references onto the new block. When
                # flexing, the stats footer formulas (rows 14-19) get a wider
                # range — handled separately below; here we just shift the
                # original.
                dst.value = _shift_formula(val, dst_row - tpl_row)
            elif isinstance(val, str) and _BANNER_SUBSECTOR_MARKER in val:
                # Banner: substitute the [Subsector] placeholder.
                dst.value = val.replace(_BANNER_SUBSECTOR_MARKER, subsector_slug)
            else:
                dst.value = val
            # Style copy (font/fill/alignment/border/number_format).
            src_style = row_styles[ci]
            if src_style is not None:
                _apply_cached_style(dst, src_style)

    # ── 1b. For row-flex: replicate the LAST template data row (row 12
    # relative) styles + formulas into the inserted rows 13..(12+extra).
    if extra > 0:
        tpl_last_data_idx = _TPL_DATA_LAST_ROW - _TPL_BLOCK_FIRST_ROW   # = 11
        last_data_vals = template_cells[tpl_last_data_idx]
        last_data_styles = template_styles[tpl_last_data_idx]
        for j in range(extra):
            src_row_no = _TPL_DATA_LAST_ROW + row_offset
            dst_row = src_row_no + 1 + j   # row 13, 14, ... (relative-shifted)
            for ci, val in enumerate(last_data_vals):
                col = ci + 1
                dst = ws.cell(row=dst_row, column=col)
                if isinstance(val, str) and val.startswith("="):
                    dst.value = _shift_formula(val, dst_row - _TPL_DATA_LAST_ROW)
                else:
                    dst.value = val
                src_style = last_data_styles[ci]
                if src_style is not None:
                    _apply_cached_style(dst, src_style)

    # ── 2. Re-create merged ranges shifted onto the new block's row offset.
    for mr in template_merged:
        ws.merge_cells(
            start_row=mr.min_row + row_offset,
            end_row=mr.max_row + row_offset,
            start_column=mr.min_col,
            end_column=mr.max_col,
        )

    # ── 3. Date stamp at B1 of this block (only if caller supplied today_iso).
    if today_iso:
        ws.cell(row=_TPL_DATE_ROW + row_offset, column=_TPL_DATE_COL, value=today_iso)

    # ── 4. Resolve column positions header-name-aware — never hardcoded.
    cols = _resolve_coco_columns(ws, block_row_offset=row_offset)

    # ── 5. Populate the block's data rows with the actual peer data. Only
    # INPUT cells are written; formula cells (G=Mkt Cap, I=EV, S/T/U/W/X
    # multiples, AD=FX flag) stay as the row-shifted template formulas.
    data_first_dst = _TPL_DATA_FIRST_ROW + row_offset
    n_data_slots = (_TPL_DATA_LAST_ROW - _TPL_DATA_FIRST_ROW + 1) + extra
    for ri, row in enumerate(coco_rows[:n_data_slots]):
        rnum = data_first_dst + ri
        _write_input(ws, rnum, cols.get("ticker"),     row.get("ticker"))
        _write_input(ws, rnum, cols.get("company"),    row.get("name"))
        _write_input(ws, rnum, cols.get("ccy_fs"),     row.get("fs_currency") or row.get("currency"))
        _write_input(ws, rnum, cols.get("shr_price"),  row.get("shr_price"))
        _write_input(ws, rnum, cols.get("shr_out"),    row.get("shares_out"))
        # Mkt Cap: write the provider's direct market cap into col G when present
        # (overrides the template's =E*F formula — avoids the GBp/GBP pence
        # mis-scale). Falls through to the formula when market_cap_m is None.
        _write_input(ws, rnum, cols.get("mkt_cap"),    row.get("market_cap_m"))
        # Net debt: single col (old layout) OR split excl/incl leases (v2). Each
        # write no-ops when its column isn't resolved, so both layouts work.
        _write_input(ws, rnum, cols.get("net_debt"),      row.get("net_debt_m"))
        _write_input(ws, rnum, cols.get("net_debt_excl"), row.get("net_debt_excl_m"))
        _write_input(ws, rnum, cols.get("net_debt_incl"), row.get("net_debt_incl_m"))
        # EBITDAaL (operator input — None here → cells left blank for the operator
        # to enter from the IFRS-16 note; the EV/EBITDAaL formulas then compute).
        _write_input(ws, rnum, cols.get("ebitdaal_lfym1"), row.get("ebitdaal_lfym1_m"))
        _write_input(ws, rnum, cols.get("ebitdaal_lfy"),   row.get("ebitdaal_lfy_m"))
        _write_input(ws, rnum, cols.get("ebitdaal_lfy1"),  row.get("ebitdaal_lfy1_m"))
        _write_input(ws, rnum, cols.get("fye"),        row.get("fye"))
        _write_input(ws, rnum, cols.get("rev_lfy"),    row.get("revenue_lfy_m"))
        _write_input(ws, rnum, cols.get("rev_lfy1"),   row.get("revenue_lfy1_m"))
        _write_input(ws, rnum, cols.get("ebitda_lfym1"), row.get("ebitda_lfym1_m"))
        _write_input(ws, rnum, cols.get("ebitda_lfy"),   row.get("ebitda_lfy_m"))
        _write_input(ws, rnum, cols.get("ebitda_lfy1"),  row.get("ebitda_lfy1_m"))
        _write_input(ws, rnum, cols.get("pe_ltm"),       row.get("pe_ltm"))
        _write_input(ws, rnum, cols.get("source"),       row.get("source"))
        _write_input(ws, rnum, cols.get("ccy_px"),
                     row.get("ccy_px") or row.get("currency"))
    # Note: stats footer extension when row-flexing — re-write the stat row
    # formulas with the extended data range.
    if extra > 0:
        _extend_stats_footer(ws, row_offset=row_offset, extra=extra, cols=cols)

    # Stamp a parsable "CoCo: <slug>" locator marker (post-AD gutter, col 31).
    # Placed in column 31 so it's clearly outside the operator's printable
    # range but findable by tests + downstream parsers.
    ws.cell(row=_TPL_BANNER_ROW + row_offset, column=31,
            value=f"CoCo: {subsector_slug}")

    # Next block begins after this block's full height (with row-flex) + gutter.
    return first_row + _BLOCK_HEIGHT + extra + _BLOCK_GUTTER


def _write_input(ws, row: int, col: Optional[int], value: Any) -> None:
    """Write an input value at ``(row, col)`` only when both col is resolved
    AND value is not None. Leaves the cell as-is for unset values — preserves
    the template's formatting / formulas for empty inputs."""
    if col is None or value is None:
        return
    ws.cell(row=row, column=col, value=value)


def _write_source_cell(
    ws, row: int, col: Optional[int], *, deal_id: str, source: str,
    tracker_row: Optional[int] = None,
) -> None:
    """Write the CoTrans Source cell with tracker-first hyperlink priority.

    Priority (operator clarification 2026-06-02 — the operator doesn't have
    a Mergermarket account so a raw web URL is useless; they want the cell
    to jump them into their own curated precedent tracker xlsx, which itself
    carries the underlying source link + curated context):

      1. ``tracker_row`` is set (deal IS in the canonical tracker) →
         hyperlink target is the canonical tracker xlsx file:// URL +
         location ``'Precedent transactions'!A<tracker_row>`` (Excel honours
         the location for cross-workbook external links when it can; harmless
         if it can't).
      2. ``tracker_row`` is None AND ``source`` is URL-shaped (rare fallback —
         deal_id approved but never landed in the tracker) → hyperlink target
         is the raw URL.
      3. Neither → write the display text as plain text, no hyperlink.

    Display text is ``tracker:<deal_id>`` whenever ``deal_id`` is populated
    (so the operator sees a stable, recognisable cross-sheet reference);
    falls back to ``source`` when ``deal_id`` is empty.

    No-ops if ``col`` is unresolved (mirrors ``_write_input``'s contract).
    """
    if col is None:
        return
    cell = ws.cell(row=row, column=col)
    display_text = f"tracker:{deal_id}" if deal_id else (source or "")

    hyperlink_target: Optional[str] = None
    hyperlink_location: Optional[str] = None

    if tracker_row is not None:
        # Priority 1: tracker-first. Point at the canonical tracker xlsx.
        # Location anchors the click at the deal's specific row inside the
        # 'Precedent transactions' sheet (Excel honours this where it can
        # for cross-workbook external links; harmless if it ignores).
        tracker_url = _canonical_tracker_file_url()
        if tracker_url:
            hyperlink_target = tracker_url
            hyperlink_location = f"'{_CANONICAL_TRACKER_SHEET}'!A{tracker_row}"
    elif isinstance(source, str):
        # Priority 2: URL-shaped source fallback. Only fires when the deal
        # isn't in the tracker (tracker_row is None) — preserves operator
        # access to SOMETHING clickable for orphan rows.
        src_stripped = source.strip()
        if src_stripped.startswith(("http://", "https://")):
            hyperlink_target = src_stripped

    if hyperlink_target is None:
        # Priority 3: plain text. No hyperlink.
        cell.value = display_text
        return

    # openpyxl doesn't auto-style hyperlinked cells (Excel does that via the
    # theme); explicitly set color + underline so the cell renders as a
    # recognisable hyperlink without depending on the template carrying
    # hyperlink-style formatting.
    from openpyxl.styles import Font
    from openpyxl.worksheet.hyperlink import Hyperlink

    cell.value = display_text
    cell.hyperlink = Hyperlink(
        ref=cell.coordinate,
        target=hyperlink_target,
        location=hyperlink_location,
        display=display_text,
    )
    cur_font = cell.font
    cell.font = Font(
        name=cur_font.name, size=cur_font.size,
        bold=cur_font.bold, italic=cur_font.italic,
        color="0563C1",   # standard Excel hyperlink blue
        underline="single",
    )


def _extend_stats_footer(
    ws, *, row_offset: int, extra: int, cols: dict[str, int],
) -> None:
    """When a block flexed (extra > 0 data rows inserted), rewrite the 6 stats
    rows' formulas to cover ``rows 5..(12+extra)`` of the block. openpyxl's
    ``insert_rows`` would have shifted them, but we never called it — the
    extra rows were stamped in place from the cached last-data-row template.
    Here we re-write each stat cell's formula to the extended range.
    """
    from openpyxl.utils import get_column_letter

    stats_first_dst = _TPL_STATS_FIRST_ROW + row_offset + extra
    data_first_dst = _TPL_DATA_FIRST_ROW + row_offset
    data_last_dst = _TPL_DATA_LAST_ROW + row_offset + extra
    stat_funcs = [
        ("AVERAGE", None),    # Mean (row 14 of block)
        ("MEDIAN",  None),    # Median (row 15)
        ("QUARTILE", 3),      # 75th (row 16)
        ("QUARTILE", 1),      # 25th (row 17)
        ("MIN",     None),    # Min (row 18)
        ("MAX",     None),    # Max (row 19)
    ]
    stat_keys = (
        "ev_ebitda_lfym1", "ev_ebitda_lfy", "ev_ebitda_lfy1",
        "ev_ebitdaal_lfym1", "ev_ebitdaal_lfy", "ev_ebitdaal_lfy1",
        "ev_rev_lfy", "ev_rev_lfy1", "pe_ltm",
    )
    for j, (fn, q) in enumerate(stat_funcs):
        rnum = stats_first_dst + j
        for k in stat_keys:
            c = cols.get(k)
            if c is None:
                continue
            letter = get_column_letter(c)
            rng = f"{letter}{data_first_dst}:{letter}{data_last_dst}"
            if fn == "QUARTILE":
                formula = f'=IFERROR(QUARTILE({rng},{q}),"")'
            else:
                formula = f'=IFERROR({fn}({rng}),"")'
            ws.cell(row=rnum, column=c, value=formula)


def _resolve_cotrans_columns(ws, block_row_offset: int = 0) -> dict[str, int]:
    """Resolve logical CoTrans column names → 1-indexed column numbers from
    the block's row-4 header row. ``block_row_offset`` shifts the read row
    for blocks 2..N (which sit below block 1)."""
    row_header = _TPL_BLOCK_FIRST_ROW + 3 + block_row_offset   # = 4
    max_col = ws.max_column
    out: dict[str, int] = {}
    for c in range(1, max_col + 1):
        v = ws.cell(row=row_header, column=c).value
        if not isinstance(v, str):
            continue
        s = v.strip().lower()
        if s == "date":
            out["date"] = c
        elif s == "target":
            out["target"] = c
        elif s == "acquirer":
            out["acquirer"] = c
        elif s == "country":
            out["country"] = c
        elif s in ("target description", "description"):
            out["description"] = c
        elif s in ("ccy", "currency"):
            out["currency"] = c
        elif s in ("ev (m)", "ev"):
            out["ev_m"] = c
        elif s in ("revenue (m)", "revenue"):
            out["revenue_m"] = c
        elif s in ("ebitda (m)", "ebitda"):
            out["ebitda_m"] = c
        elif s in ("ev/revenue", "ev / revenue"):
            out["ev_revenue"] = c
        elif s in ("ev/ebitda", "ev / ebitda"):
            out["ev_ebitda"] = c
        elif s in ("acquirer type", "buyer type", "type"):
            out["buyer_type"] = c
        elif s in ("strategic commentary", "commentary"):
            out["strategic_commentary"] = c
        elif s == "source":
            out["source"] = c
    return out


def _stamp_one_cotrans_block(
    ws, *, first_row: int, subsector_slug: str,
    cotrans_rows: list[dict[str, Any]],
    template_cells: list[list[Any]], template_styles: list[list[Any]],
    template_merged: list[Any],
    today_iso: Optional[str] = None,
) -> int:
    """Stamp ONE CoTrans block with ROW-FLEX: the block grows to fit ANY number
    of deals (no 8-row cap, so the operator is never limited). When n_deals
    exceeds the template's pre-formatted data rows, extra rows are inserted
    after the last data row (replicating its styles + the EV/Revenue, EV/EBITDA
    formulas), the stats footer + notes shift down by ``extra``, and the stats
    ranges extend to cover all deals. Mirrors ``_stamp_one_coco_block``."""
    row_offset = first_row - _TPL_BLOCK_FIRST_ROW
    n_deals = len(cotrans_rows)
    n_tpl_data = _TPL_DATA_LAST_ROW - _TPL_DATA_FIRST_ROW + 1
    extra = max(0, n_deals - n_tpl_data)
    dst_first = _TPL_BLOCK_FIRST_ROW + row_offset
    dst_last = _TPL_BLOCK_LAST_ROW + row_offset + extra

    # Unmerge pre-existing merges in the destination region (see CoCo
    # stamper for rationale).
    _unmerge_in_block_region(ws, dst_first, dst_last)

    # Replicate the template block. Rows below the data region (stats + notes,
    # tpl_row > 12) shift down by `extra` to make room for the inserted rows.
    for ri, row_vals in enumerate(template_cells):
        row_styles = template_styles[ri]
        tpl_row = _TPL_BLOCK_FIRST_ROW + ri
        dst_row = tpl_row + row_offset
        if extra > 0 and tpl_row > _TPL_DATA_LAST_ROW:
            dst_row += extra
        for ci, val in enumerate(row_vals):
            col = ci + 1
            dst = ws.cell(row=dst_row, column=col)
            if isinstance(val, str) and val.startswith("="):
                dst.value = _shift_formula(val, dst_row - tpl_row)
            elif isinstance(val, str) and _BANNER_SUBSECTOR_MARKER in val:
                dst.value = val.replace(_BANNER_SUBSECTOR_MARKER, subsector_slug)
            else:
                dst.value = val
            src_style = row_styles[ci]
            if src_style is not None:
                _apply_cached_style(dst, src_style)

    # Row-flex: replicate the LAST template data row into the inserted rows.
    if extra > 0:
        tpl_last_data_idx = _TPL_DATA_LAST_ROW - _TPL_BLOCK_FIRST_ROW
        last_vals = template_cells[tpl_last_data_idx]
        last_styles = template_styles[tpl_last_data_idx]
        for j in range(extra):
            dst_row = _TPL_DATA_LAST_ROW + row_offset + 1 + j
            for ci, val in enumerate(last_vals):
                col = ci + 1
                dst = ws.cell(row=dst_row, column=col)
                if isinstance(val, str) and val.startswith("="):
                    dst.value = _shift_formula(val, dst_row - _TPL_DATA_LAST_ROW)
                else:
                    dst.value = val
                src_style = last_styles[ci]
                if src_style is not None:
                    _apply_cached_style(dst, src_style)

    for mr in template_merged:
        top = mr.min_row + row_offset + (extra if mr.min_row > _TPL_DATA_LAST_ROW else 0)
        bot = mr.max_row + row_offset + (extra if mr.max_row > _TPL_DATA_LAST_ROW else 0)
        ws.merge_cells(start_row=top, end_row=bot,
                       start_column=mr.min_col, end_column=mr.max_col)

    # Header-name-aware column resolver (mirrors CoCo's approach).
    cols = _resolve_cotrans_columns(ws, block_row_offset=row_offset)

    # CoTrans data rows — ALL of them (row-flex grew the block to fit).
    data_first_dst = _TPL_DATA_FIRST_ROW + row_offset
    n_slots = n_tpl_data + extra
    for ri, row in enumerate(cotrans_rows[:n_slots]):
        rnum = data_first_dst + ri
        _write_input(ws, rnum, cols.get("date"),         row.get("announced_date"))
        _write_input(ws, rnum, cols.get("target"),       row.get("target"))
        _write_input(ws, rnum, cols.get("acquirer"),     row.get("acquirer"))
        _write_input(ws, rnum, cols.get("buyer_type"),   row.get("buyer_type"))
        _write_input(ws, rnum, cols.get("country"),      row.get("country"))
        _write_input(ws, rnum, cols.get("description"),
                     row.get("description") or row.get("strategic_commentary"))
        _write_input(ws, rnum, cols.get("currency"),     row.get("currency"))
        _write_input(ws, rnum, cols.get("ev_m"),         row.get("ev_m"))
        _write_input(ws, rnum, cols.get("revenue_m"),    row.get("revenue_m"))
        _write_input(ws, rnum, cols.get("ebitda_m"),     row.get("ebitda_m"))
        # Direct multiples from the source (PDF / tracker) → override the
        # template's =EV/Revenue, =EV/EBITDA formulas when supplied. Precedent
        # deals usually disclose the multiple, not the underlying rev/EBITDA,
        # so without this the operator's sourced multiples (e.g. 11.7x) would
        # be silently dropped (the formula computes blank with no rev/EBITDA).
        _write_input(ws, rnum, cols.get("ev_revenue"),   row.get("ev_revenue_x"))
        _write_input(ws, rnum, cols.get("ev_ebitda"),    row.get("ev_ebitda_x"))
        _write_input(ws, rnum, cols.get("strategic_commentary"),
                     row.get("strategic_commentary"))
        _write_source_cell(
            ws, rnum, cols.get("source"),
            deal_id=str(row.get("deal_id") or ""),
            source=str(row.get("source") or ""),
            tracker_row=row.get("_tracker_row"),
        )
    # Stats footer (EV/Revenue + EV/EBITDA, Mean/Median/quartiles/Min/Max) —
    # re-base the ranges to cover all deals when the block flexed.
    if extra > 0:
        _extend_cotrans_stats_footer(ws, row_offset=row_offset, extra=extra, cols=cols)

    # Banner-locator marker — far gutter col (col 31).
    ws.cell(row=_TPL_BANNER_ROW + row_offset, column=31,
            value=f"CoTrans: {subsector_slug}")

    return first_row + _BLOCK_HEIGHT + extra + _BLOCK_GUTTER


def _extend_cotrans_stats_footer(
    ws, *, row_offset: int, extra: int, cols: dict[str, int],
) -> None:
    """Re-base the CoTrans stats rows (Mean/Median/75th/25th/Min/Max) on the
    EV/Revenue + EV/EBITDA columns to cover ``rows 5..(12+extra)`` after a
    row-flex. Mirrors ``_extend_stats_footer`` (CoCo)."""
    from openpyxl.utils import get_column_letter

    stats_first_dst = _TPL_STATS_FIRST_ROW + row_offset + extra
    data_first_dst = _TPL_DATA_FIRST_ROW + row_offset
    data_last_dst = _TPL_DATA_LAST_ROW + row_offset + extra
    stat_funcs = [
        ("AVERAGE", None), ("MEDIAN", None), ("QUARTILE", 3),
        ("QUARTILE", 1), ("MIN", None), ("MAX", None),
    ]
    for j, (fn, q) in enumerate(stat_funcs):
        rnum = stats_first_dst + j
        for k in ("ev_revenue", "ev_ebitda"):
            c = cols.get(k)
            if c is None:
                continue
            letter = get_column_letter(c)
            rng = f"{letter}{data_first_dst}:{letter}{data_last_dst}"
            if fn == "QUARTILE":
                formula = f'=IFERROR(QUARTILE({rng},{q}),"")'
            else:
                formula = f'=IFERROR({fn}({rng}),"")'
            ws.cell(row=rnum, column=c, value=formula)


def _cache_template_block(ws) -> tuple[list[list[Any]], list[list[Any]], list[Any]]:
    """Cache the template's block region (rows 1-19, all columns) as
    (values, styles, merged-ranges-in-region) BEFORE any in-place mutation.

    Block 2..N replicate from THIS cache, so they get the template's
    original layout — not block-1's data-populated state.
    """
    from copy import copy as _copy

    n_cols = ws.max_column
    cells: list[list[Any]] = []
    styles: list[list[Any]] = []
    for r in range(_TPL_BLOCK_FIRST_ROW, _TPL_BLOCK_LAST_ROW + 1):
        row_vals: list[Any] = []
        row_styles: list[Any] = []
        for c in range(1, n_cols + 1):
            cell = ws.cell(row=r, column=c)
            row_vals.append(cell.value)
            if cell.has_style:
                row_styles.append({
                    "font": _copy(cell.font),
                    "fill": _copy(cell.fill),
                    "alignment": _copy(cell.alignment),
                    "border": _copy(cell.border),
                    "number_format": cell.number_format,
                    "protection": _copy(cell.protection),
                })
            else:
                row_styles.append(None)
        cells.append(row_vals)
        styles.append(row_styles)
    # Merged ranges intersecting the block region.
    merged: list[Any] = []
    for mr in list(ws.merged_cells.ranges):
        if mr.min_row >= _TPL_BLOCK_FIRST_ROW and mr.max_row <= _TPL_BLOCK_LAST_ROW:
            # Snapshot the bounds (the live MergedCellRange object mutates).
            from openpyxl.worksheet.cell_range import CellRange
            merged.append(CellRange(
                min_row=mr.min_row, max_row=mr.max_row,
                min_col=mr.min_col, max_col=mr.max_col,
            ))
    return cells, styles, merged


def _apply_cached_style(dst_cell, cached: dict[str, Any]) -> None:
    """Apply a cached style dict (built by ``_cache_template_block``) onto
    ``dst_cell``."""
    dst_cell.font = cached["font"]
    dst_cell.fill = cached["fill"]
    dst_cell.alignment = cached["alignment"]
    dst_cell.border = cached["border"]
    dst_cell.number_format = cached["number_format"]
    dst_cell.protection = cached["protection"]


def _cache_template_notes(ws) -> tuple[list[list[Any]], list[list[Any]], list[Any]]:
    """Cache the notes/methodology rows (22-28) so we can re-stamp them at
    the bottom of the output, AFTER all replicated blocks."""
    from copy import copy as _copy

    n_cols = ws.max_column
    cells: list[list[Any]] = []
    styles: list[list[Any]] = []
    for r in range(_TPL_NOTES_FIRST_ROW, _TPL_NOTES_LAST_ROW + 1):
        row_vals: list[Any] = []
        row_styles: list[Any] = []
        for c in range(1, n_cols + 1):
            cell = ws.cell(row=r, column=c)
            row_vals.append(cell.value)
            if cell.has_style:
                row_styles.append({
                    "font": _copy(cell.font),
                    "fill": _copy(cell.fill),
                    "alignment": _copy(cell.alignment),
                    "border": _copy(cell.border),
                    "number_format": cell.number_format,
                    "protection": _copy(cell.protection),
                })
            else:
                row_styles.append(None)
        cells.append(row_vals)
        styles.append(row_styles)
    merged: list[Any] = []
    for mr in list(ws.merged_cells.ranges):
        if mr.min_row >= _TPL_NOTES_FIRST_ROW and mr.max_row <= _TPL_NOTES_LAST_ROW:
            from openpyxl.worksheet.cell_range import CellRange
            merged.append(CellRange(
                min_row=mr.min_row, max_row=mr.max_row,
                min_col=mr.min_col, max_col=mr.max_col,
            ))
    return cells, styles, merged


def _clear_template_notes(ws) -> None:
    """Wipe the notes/methodology rows (we re-stamp them at the bottom).

    UNMERGE first — a MergedCell's ``.value`` is read-only, so any merged
    cell in the notes region must be unmerged before we can blank it.
    """
    # Unmerge any merged ranges in the notes region so we can blank the
    # cells + so the re-stamped notes at the bottom can re-create their
    # merges without conflict.
    for mr in list(ws.merged_cells.ranges):
        if mr.min_row >= _TPL_NOTES_FIRST_ROW and mr.max_row <= _TPL_NOTES_LAST_ROW:
            ws.unmerge_cells(str(mr))
    for r in range(_TPL_NOTES_FIRST_ROW, _TPL_NOTES_LAST_ROW + 1):
        for c in range(1, ws.max_column + 1):
            ws.cell(row=r, column=c).value = None


def _stamp_notes_at_bottom(
    ws, *, first_row: int,
    template_cells: list[list[Any]], template_styles: list[list[Any]],
    template_merged: list[Any],
) -> None:
    """Stamp the cached notes/methodology rows starting at ``first_row``."""
    row_offset = first_row - _TPL_NOTES_FIRST_ROW
    for ri, row_vals in enumerate(template_cells):
        row_styles = template_styles[ri]
        tpl_row = _TPL_NOTES_FIRST_ROW + ri
        dst_row = tpl_row + row_offset
        for ci, val in enumerate(row_vals):
            col = ci + 1
            dst = ws.cell(row=dst_row, column=col)
            dst.value = val
            src_style = row_styles[ci]
            if src_style is not None:
                _apply_cached_style(dst, src_style)
    for mr in template_merged:
        ws.merge_cells(
            start_row=mr.min_row + row_offset,
            end_row=mr.max_row + row_offset,
            start_column=mr.min_col,
            end_column=mr.max_col,
        )


def _shim_xlsx_stamp(
    template_src: Path, output_path: Path, blocks: list[dict[str, Any]],
    *, sensitivity: str,
) -> None:
    """Block-flex stamp of the v2 comps template — TEMPLATE-COPY variant.

    This is the WS-A implementation (per SESSION-COMPS-FOLLOWUP brief +
    re-verified 2026-06-01 template-safety probe). The flow:

      1. ``shutil.copy2`` the v2 template into the deal Valuation folder
         (preserves operator's formatting / formulas / fills bit-for-bit).
      2. Probe the loaded copy for openpyxl-unsafe features (threaded
         comments / conditional formatting / data validations / charts /
         images). If ANY appear (a future operator template edit), fail
         loud with a clear message — DO NOT silently strip them. The
         documented future-seam drop-in for that case is
         ``anthropic-skills:xlsx`` (cloud-side); not a dependency today.
      3. Cache the template's block region (rows 1-19) + notes region
         (rows 22-28) cell-by-cell with their styles BEFORE any mutation.
      4. For block 1: stamp in place over rows 1-19 (substitute the
         banner's [Subsector] marker, populate data rows, formulas in the
         template stay since they reference block-1's data range).
      5. For block 2..N: copy the cached block region (with styles +
         re-targeted formulas) at the next free row offset.
      6. Re-stamp the cached notes region at the very bottom (after all
         replicated blocks) so the methodology text appears ONCE.
      7. Save the populated workbook to ``output_path``.

    The per-block stats rows (Mean / Median / 75th / 25th / Min / Max for
    EV/Revenue + EV/EBITDA) are re-based per block — block i's formula
    references only block i's data range, NOT a union across blocks.

    Args:
        template_src: Path to the v2 template (typically
            ``os-templates/Project_x_Comps_v2.xlsx``; tests pass a
            synthetic v2-shaped fixture).
        output_path: Where the populated XLSX lands. Parent is created.
        blocks: One entry per approved subsector, carrying ``coco_rows`` +
            ``cotrans_rows`` populated by Stage 2.
        sensitivity: Workspace sensitivity (gating happens at the route).

    Raises:
        TemplateStampFailed: openpyxl-unsafe features detected in the template,
            OR a stamp-time openpyxl error. The detailed offending-features
            description is included in the error message for the operator.
    """
    import openpyxl

    _sensitivity_check(sensitivity)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    # ── 1. Copy the template bit-for-bit into the output path. The copy is
    # the canvas we mutate; the source file is never touched.
    if not template_src.exists():
        raise TemplateStampFailed(
            f"v2 template not found at {template_src!r}; cannot stamp"
        )
    shutil.copy2(str(template_src), str(output_path))

    # ── 2. Load the copy + safety-probe BEFORE any mutation. If the operator
    # has added any openpyxl-unsafe features since the 2026-06-01
    # re-verification, fail loud — silently dropping them would corrupt the
    # operator's expectations. Future-seam: anthropic-skills:xlsx cloud-side
    # drop-in handles the unsafe-template case (not a dependency today).
    wb = openpyxl.load_workbook(str(output_path))
    offending = _probe_template_safe(wb)
    if offending:
        raise TemplateStampFailed(
            f"v2 template at {template_src!r} carries openpyxl-unsafe features "
            f"(template-edit drift since the 2026-06-01 re-verification). "
            f"Refusing to stamp to avoid silently dropping operator content. "
            f"Either remove the features from the template OR switch this "
            f"shim to the anthropic-skills:xlsx cloud-side drop-in. "
            f"Offending:\n  - " + "\n  - ".join(offending)
        )

    # Today's stamp (ISO YYYY-MM-DD) lands at B1 of each block on each sheet.
    today_iso = datetime.now(timezone.utc).date().isoformat()

    # ── 3-6. Per sheet: cache template, stamp blocks, re-stamp notes.
    for sheet_name, get_rows_key in (("CoCo", "coco_rows"), ("CoTrans", "cotrans_rows")):
        if sheet_name not in wb.sheetnames:
            raise TemplateStampFailed(
                f"v2 template missing required sheet {sheet_name!r}; got "
                f"{wb.sheetnames!r}"
            )
        ws = wb[sheet_name]
        # Cache the template's BLOCK + NOTES regions before any mutation.
        tpl_block_cells, tpl_block_styles, tpl_block_merged = _cache_template_block(ws)
        tpl_notes_cells, tpl_notes_styles, tpl_notes_merged = _cache_template_notes(ws)
        # Wipe the template's notes region — we re-stamp at the bottom.
        _clear_template_notes(ws)
        # Stamp each block at its row offset.
        next_row = _TPL_BLOCK_FIRST_ROW
        for block in blocks:
            ss = str(block.get("subsector_slug") or "unknown")
            rows = list(block.get(get_rows_key) or [])
            if sheet_name == "CoCo":
                next_row = _stamp_one_coco_block(
                    ws, first_row=next_row, subsector_slug=ss,
                    coco_rows=rows,
                    template_cells=tpl_block_cells, template_styles=tpl_block_styles,
                    template_merged=tpl_block_merged,
                    today_iso=today_iso,
                )
            else:
                next_row = _stamp_one_cotrans_block(
                    ws, first_row=next_row, subsector_slug=ss,
                    cotrans_rows=rows,
                    template_cells=tpl_block_cells, template_styles=tpl_block_styles,
                    template_merged=tpl_block_merged,
                    today_iso=today_iso,
                )
        # Re-stamp notes at the very bottom (1-row gap).
        _stamp_notes_at_bottom(
            ws, first_row=next_row,
            template_cells=tpl_notes_cells, template_styles=tpl_notes_styles,
            template_merged=tpl_notes_merged,
        )

    # ── 7. Save.
    wb.save(str(output_path))
    log.info(
        "comps: xlsx-stamp -> %s (%d blocks; template-copy; sensitivity=%s; src=%s)",
        output_path, len(blocks), sensitivity, template_src,
    )


# ─────────────────────────────────────────────────────────────────────────────
# Stage 0 — Understand + scope
# ─────────────────────────────────────────────────────────────────────────────


# Optional leading/trailing double-quote tolerance — `- "Sector: descriptor"`
# is the YAML-conventional "list of strings" form; `- Sector: descriptor`
# (unquoted) is the keyed-mapping form. Both parse the same way here so
# operators can write whichever feels natural in profile.md.
#
# Sector key allows Unicode letters + hyphens (`\w` is Unicode-aware in
# Py3) so non-English operator deployments work (`São Paulo: ...`,
# `Telecoms-Süd: ...`) — codex-review finding P3, 2026-06-03.
#
# Trailing inline `# comment` after the descriptor is stripped, but
# REQUIRES whitespace before the `#` (`\s#`) so `#` inside a descriptor
# (e.g. `"Sector: foo (a#b / c)"`) is NOT mistaken for a comment marker
# — codex-review finding P3, 2026-06-03 (previously `(?:#.*)?` matched
# any `#` and silently dropped descriptor content after the first one).
_SECTOR_SUB_LENS_LINE = re.compile(
    r'^\s*-\s+"?(?P<sector>[\w &.\-]+):\s*(?P<descriptors>.+?)"?\s*(?:\s#.*)?\s*$',
    re.UNICODE,
)


def _to_slug(text: str) -> str:
    """Reduce arbitrary text to a kebab-case slug — Unicode-aware ASCII fold
    (São → sao) via NFD normalisation + combining-mark strip, then any
    non-alphanumeric run collapses to a single ``-`` with leading/trailing
    dashes stripped.

    Replaces the prior ``.lower().replace(" ", "-")`` which left ``/``,
    ``&``, commas, periods etc. as-is and produced invalid slugs like
    ``tower-/-passive-infrastructure`` for entries containing slashes.

    The NFD step (codex-review finding P3, 2026-06-03) preserves Unicode
    sector names for new-operator deployments — without it, ``São Paulo``
    became ``s-o-paulo`` (the ``ã`` was a non-alphanumeric, dropped).
    """
    nfd = unicodedata.normalize("NFD", text.lower())
    folded = "".join(c for c in nfd if not unicodedata.combining(c))
    return re.sub(r"[^a-z0-9]+", "-", folded).strip("-")


def _parse_sector_sub_lens(profile_text: str) -> dict[str, list[str]]:
    """Parse the ``sector_sub_lens:`` block out of profile.md.

    The block is a YAML-like mapping but profile.md is also free-form markdown
    in some sections, so this is a tolerant line-by-line parser. Returns
    ``{sector-slug: [subsector-slug, ...]}``."""
    out: dict[str, list[str]] = {}
    in_block = False
    for line in profile_text.splitlines():
        stripped = line.strip()
        if stripped.startswith("sector_sub_lens:"):
            in_block = True
            continue
        if in_block:
            # Block ends at the next top-level YAML key (no indent, no '-')
            # or at a blank line followed by a non-indented line.
            if stripped and not (line.startswith(" ") or line.startswith("\t") or stripped.startswith("-")):
                break
            m = _SECTOR_SUB_LENS_LINE.match(line)
            if not m:
                continue
            parent = _to_slug(m.group("sector"))
            descriptors = m.group("descriptors").strip()
            # descriptors look like "Hotels (full-service / limited-service / lifestyle / boutique)"
            # Extract the parenthesised list + split on / + slugify each.
            subs: list[str] = []
            paren = re.search(r"\(([^)]+)\)", descriptors)
            if paren:
                root = _to_slug(descriptors.split("(", 1)[0])
                for piece in paren.group(1).split("/"):
                    slug = _to_slug(piece)
                    if slug:
                        subs.append(f"{root}-{slug}" if root else slug)
            else:
                # No parentheses → take the whole descriptor as the single subsector.
                slug = _to_slug(descriptors)
                if slug:
                    subs.append(slug)
            # Accumulate (with order-preserving dedup) so multiple keyed lines for
            # the same parent sector compose rather than overwrite. Lets profile.md
            # group a sector's subsectors across multiple readable lines:
            #   - Hospitality: Hotels (full-service / limited-service / lifestyle)
            #   - Hospitality: Holiday parks
            #   - Hospitality: Pubs (managed / tenanted / freehold-heavy)
            # Stage 0 sees all 7 hospitality subsectors instead of just the last line's.
            existing = out.setdefault(parent, [])
            existing.extend(s for s in subs if s not in existing)
    return out


def run_stage_0(
    inputs: CompsBuildInput,
    *,
    vault_root: Path,
    run_id: str,
) -> StageResult:
    """Read the target + propose subsectors. Returns an ``approval_pending``
    payload; the operator approves the subsector list to advance to Stage 1."""
    profile_path = vault_root / "_claude" / "profile.md"
    if not profile_path.is_file():
        raise TargetBriefMissing(
            f"profile.md not found at {profile_path}; cannot propose subsectors"
        )
    sub_lens = _parse_sector_sub_lens(profile_path.read_text(encoding="utf-8"))
    parent_slug = inputs.parent_sector.strip().lower().replace(" ", "-")
    candidates = sub_lens.get(parent_slug, [])
    if not candidates:
        # Surface as a warning + empty proposed list; operator decides whether
        # to update profile.md or accept the single-bucket default.
        warnings = [
            f"profile.md sector_sub_lens has no entry for parent sector "
            f"{parent_slug!r}; proposing the parent as the single subsector"
        ]
        proposed = [parent_slug]
    else:
        warnings = []
        proposed = list(candidates)

    # One-line rationale per candidate (templated; the real propose loop adds
    # judgment via the brief / CIM, mocked in tests).
    rationale = {slug: f"candidate from profile.md sector_sub_lens.{parent_slug}"
                 for slug in proposed}

    payload = ApprovalPayload(
        kind="subsectors", proposed=proposed, rationale=rationale,
    )
    # WS-B: HMAC-sign the token over (deal_name, STAGE_LABEL_SUBSECTORS,
    # h(proposed)). On Stage 1 re-fire the route recomputes from the
    # caller-supplied approved_subsectors + presented token.
    token = _sign_token(inputs.deal_name, STAGE_LABEL_SUBSECTORS, proposed)
    return StageResult(
        ok=True,
        stage="approval_pending",
        stage_just_completed=0,
        deal_name=inputs.deal_name,
        target=inputs.target,
        run_id=run_id,
        approval_payload=payload,
        approval_token_to_sign=token,
        subsectors_approval_token=token,
        warnings=warnings,
    )


# ─────────────────────────────────────────────────────────────────────────────
# Stage 1 — Identify peers + deals (+ tracker write-back)
# ─────────────────────────────────────────────────────────────────────────────


def run_stage_1(
    inputs: CompsBuildInput,
    *,
    vault_root: Path,
    run_id: str,
    tracker_workbook: Optional[Path] = None,
    today: Optional[date] = None,
) -> StageResult:
    """Per approved subsector: propose CoCo peers + CoTrans deals. Write any
    deep-research-surfaced deals BACK to the tracker. Return approval_pending.

    The ``approved_subsectors`` on the input must be a non-empty list; absent
    → MissingApprovalToken (route maps to 422)."""
    if not inputs.approved_subsectors:
        raise MissingApprovalToken(
            "Stage 1 requires approved_subsectors (from operator-signed Stage 0)"
        )
    # WS-B + #21-comps-step-3: HMAC-verify the Stage-0-issued
    # subsectors_approval_token. Two modes:
    #   (a) proposed_subsectors absent → exact-match against approved_subsectors
    #       (pre-Step-3 behaviour; the operator echoed the proposal verbatim).
    #   (b) proposed_subsectors supplied → verify token over proposed_subsectors
    #       (the bridge's full Stage-0 universe) + enforce approved ⊆ proposed
    #       so the operator can NARROW the subsector list without re-firing
    #       Stage 0. Operator can only DROP items, never ADD (a subset add
    #       would let a caller smuggle unsigned data past the gate).
    # Forged / missing / stale / cross-stage-replayed all mismatch the HMAC.
    ok, err = _verify_subset_approval(
        deal_name=inputs.deal_name,
        stage_label=STAGE_LABEL_SUBSECTORS,
        approved=list(inputs.approved_subsectors),
        proposed=(
            list(inputs.proposed_subsectors)
            if inputs.proposed_subsectors is not None
            else None
        ),
        token=inputs.subsectors_approval_token or "",
    )
    if not ok:
        raise MissingApprovalToken(
            f"Stage 1: invalid or missing subsectors_approval_token — {err}. "
            "Either echo the Stage-0 proposal verbatim as approved_subsectors, "
            "OR send proposed_subsectors (the full Stage-0 proposal) + "
            "approved_subsectors (any subset of it) to narrow without re-firing."
        )
    today = today or datetime.now(timezone.utc).date()

    coco_proposed_by_subsector: dict[str, list[dict[str, Any]]] = {}
    cotrans_proposed_by_subsector: dict[str, list[dict[str, Any]]] = {}
    tracker_writes_planned: list[dict[str, Any]] = []
    tracker_writes_done: list[dict[str, Any]] = []
    warnings: list[str] = []

    # Lazy import — tests don't need openpyxl installed unless they exercise
    # the write-back path.
    try:
        from routines.dealtracker.schema import DealRecord, build_deal_id
        from routines.dealtracker.workbook import (
            CANONICAL_WORKBOOK_PATH, append_deal,
        )
    except Exception as e:  # noqa: BLE001
        warnings.append(f"tracker write-back unavailable ({e}); proposed deals not persisted")
        DealRecord = None
        append_deal = None
        CANONICAL_WORKBOOK_PATH = None
        build_deal_id = None

    workbook_path = tracker_workbook or CANONICAL_WORKBOOK_PATH

    for subsector_slug in inputs.approved_subsectors:
        # ── CoCo peers ────────────────────────────────────────────────────
        coco_candidates = _shim_equity_research_screen(
            subsector_slug, inputs.target, sensitivity=inputs.workspace_sensitivity,
        )
        bl_candidates = _shim_investment_banking_buyer_list(
            subsector_slug, inputs.target, sensitivity=inputs.workspace_sensitivity,
        )
        # Path-A: operator-submitted CoCo candidates for this subsector (the
        # attended orchestration session's equity-research:screen output).
        # Listed FIRST in the dedup so a submitted candidate establishes the
        # canonical row (name / country / why); the shims/provider just append
        # to its ``found_by``.
        submitted_coco = list(
            (inputs.submitted_coco_candidates_by_subsector or {}).get(subsector_slug, [])
        )
        # Provider peers — sanity-check; route maps result to the dict shape.
        try:
            from routines.markets import get_provider
            provider = get_provider()
            # Seed off the first available ticker: shim screen output first,
            # else an operator-submitted candidate (the common Path-A case
            # where the shims are unwired and the session supplies the peers).
            seed = None
            if coco_candidates:
                seed = coco_candidates[0].get("ticker")
            elif submitted_coco:
                seed = submitted_coco[0].get("ticker")
            provider_peers: list[dict[str, Any]] = []
            if seed:
                pr = provider.get_peers(seed, limit=8)
                provider_peers = [
                    {"ticker": p.symbol, "name": p.name, "country": p.country,
                     "why": "provider peer (sanity check)",
                     "found_by": ["provider.get_peers"]}
                    for p in pr.peers
                ]
        except Exception as e:  # noqa: BLE001
            warnings.append(f"provider.get_peers failed for {subsector_slug}: {e}")
            provider_peers = []

        # Dedup by ticker across the sources, preserving order of discovery +
        # accumulating the `found_by` list. Operator-submitted first.
        seen: dict[str, dict[str, Any]] = {}
        for source_name, batch in (
            ("operator-submitted", submitted_coco),
            ("equity-research:screen", coco_candidates),
            ("investment-banking:buyer-list", bl_candidates),
            ("provider.get_peers", provider_peers),
        ):
            for cand in batch:
                tk = (cand.get("ticker") or "").upper()
                if not tk or tk == inputs.target.upper():
                    continue  # never include the target as its own peer
                if tk in seen:
                    seen[tk].setdefault("found_by", []).append(source_name)
                else:
                    row = dict(cand)
                    row["ticker"] = tk
                    row["found_by"] = list(row.get("found_by") or [source_name])
                    seen[tk] = row
        coco_proposed_by_subsector[subsector_slug] = list(seen.values())

        # ── CoTrans: tracker query + deep-research / operator-submitted gap-fill ──
        tracker_deals = _query_tracker(workbook_path, subsector_slug, lookback_years=5)
        dr_deals = _shim_deep_research_cotrans(
            subsector_slug, 5, sensitivity=inputs.workspace_sensitivity,
        )
        # Path-A: operator-submitted CoTrans candidates flow through the SAME
        # dedup + tracker write-back path as deep-research output (the redesign
        # SSOT rule — sourced deals get written BACK into the tracker, never a
        # second source of truth). Submitted first so they win the dedup key.
        submitted_cotrans = list(
            (inputs.submitted_cotrans_candidates_by_subsector or {}).get(subsector_slug, [])
        )
        research_deals = submitted_cotrans + list(dr_deals)
        # Dedup research deals vs tracker AND vs each other by (announced_date,
        # target) — submitted + deep-research can overlap on the same deal.
        existing_keys = {
            (d.get("announced_date") or "", (d.get("target") or "").strip().lower())
            for d in tracker_deals
        }
        gap_fills: list[dict[str, Any]] = []
        for d in research_deals:
            key = (d.get("announced_date") or "", (d.get("target") or "").strip().lower())
            if key in existing_keys:
                continue
            existing_keys.add(key)
            gap_fills.append(d)

        # Write back the gap-fills to the canonical tracker (per Stage 1
        # contract — the tracker is the single source of truth).
        if gap_fills and append_deal is not None and workbook_path is not None and DealRecord is not None:
            for d in gap_fills:
                ann_iso = d.get("announced_date") or ""
                try:
                    ann = date.fromisoformat(ann_iso) if ann_iso else None
                except (ValueError, TypeError):
                    ann = None
                rec = DealRecord(
                    announced_date=ann,
                    target_company=str(d.get("target") or ""),
                    target_description=str(d.get("target_description") or ""),
                    target_sector=inputs.parent_sector,
                    subsector_slug=subsector_slug,
                    target_country=str(d.get("country") or ""),
                    bidder_company=str(d.get("acquirer") or ""),
                    currency=str(d.get("currency") or ""),
                    enterprise_value_m=_safe_float(d.get("ev_m")),
                    reported_revenue_m_y1=_safe_float(d.get("revenue_m")),
                    reported_ebitda_m_y1=_safe_float(d.get("ebitda_m")),
                    reported_revenue_multiple_y1=_safe_float(d.get("ev_revenue_x")),
                    reported_ebitda_multiple_y1=_safe_float(d.get("ev_ebitda_x")),
                    deal_description=str(d.get("deal_description") or ""),
                    source=str(d.get("source") or d.get("source_url") or ""),
                    source_url=str(d.get("source") or d.get("source_url") or ""),
                    deal_id=build_deal_id(ann, str(d.get("target") or "")),
                )
                # Propagate the canonical deal_id back onto the proposed dict so
                # the operator approves an ID that Stage 2 can hydrate from the
                # tracker (the row was just written there). Without this, the
                # proposed entry carried no deal_id and Stage 2 fell back to a
                # bare back-reference — fixes that for both submitted +
                # deep-research deals.
                d["deal_id"] = d.get("deal_id") or rec.deal_id
                planned = {
                    "deal_id": rec.deal_id,
                    "subsector_slug": subsector_slug,
                    "announced_date": ann_iso,
                    "target": rec.target_company,
                    "source": rec.source,
                }
                tracker_writes_planned.append(planned)
                try:
                    res = append_deal(workbook_path, rec)
                    tracker_writes_done.append({**planned, **res})
                except Exception as e:  # noqa: BLE001
                    warnings.append(
                        f"tracker append_deal failed for {rec.deal_id}: {e}"
                    )

        cotrans_proposed_by_subsector[subsector_slug] = tracker_deals + [
            {**d, "source_note": "sourced (operator-submitted / deep-research); written back to tracker"}
            for d in gap_fills
        ]

    payload = ApprovalPayload(
        kind="peers_and_deals",
        proposed={
            "coco_by_subsector": coco_proposed_by_subsector,
            "cotrans_by_subsector": cotrans_proposed_by_subsector,
        },
        tracker_writes_planned=tracker_writes_planned,
    )
    # WS-B: TWO HMAC tokens — one over the approved-peers shape the operator
    # will signal, one over the approved-deals shape. Stage 2 verifies both
    # independently so mutating one payload doesn't sneak past with the
    # other's still-valid token. The proposed→approved shape is the same
    # {subsector_slug: [ticker_or_deal_id, ...]} mapping the operator picks
    # from; we sign on the FULL proposed mapping as the "what would be
    # approved if the operator accepts the proposal as-is" payload. The
    # operator may down-select — in which case they re-call Stage 1 with the
    # narrowed scope, NOT mutate Stage 2's payload behind a stale token.
    proposed_peers = {ss: [c.get("ticker") for c in coco_proposed_by_subsector.get(ss, [])]
                      for ss in inputs.approved_subsectors}
    proposed_deals = {ss: [d.get("deal_id") or d.get("target") or ""
                           for d in cotrans_proposed_by_subsector.get(ss, [])]
                      for ss in inputs.approved_subsectors}
    peers_token = _sign_token(inputs.deal_name, STAGE_LABEL_PEERS, proposed_peers)
    deals_token = _sign_token(inputs.deal_name, STAGE_LABEL_DEALS, proposed_deals)

    # Concatenate the two HMACs into ``approval_token_to_sign`` for backward
    # compatibility with the wire shape (single field). The operator's
    # bridge UI splits on ":" and presents peers_approval_token +
    # deals_approval_token separately on Stage 2 re-fire.
    combined = f"peers={peers_token}|deals={deals_token}"

    result = StageResult(
        ok=True,
        stage="approval_pending",
        stage_just_completed=1,
        deal_name=inputs.deal_name,
        target=inputs.target,
        run_id=run_id,
        approval_payload=payload,
        approval_token_to_sign=combined,
        warnings=warnings,
        approved_subsectors=list(inputs.approved_subsectors),
        peers_approval_token=peers_token,
        deals_approval_token=deals_token,
    )
    # Stash the writes done so the route's audit row carries them.
    result.tracker_writes = tracker_writes_done
    return result


# ── Deal-status filter (#21-comps Q6, operator decision 2026-06-02) ──────────
# Default policy "announced_or_closed": include announced + closed/completed
# deals; exclude terminated / withdrawn / rumoured. Deep-research-sourced deals
# carry NO explicit status → treated as ANNOUNCED (included). A Mergermarket
# import (later) supplies a closed/completed status + closed date. The tracker
# has no Status column today, so this is include-by-construction with the
# exclusion hook already in place for when status data arrives.
_DEAL_STATUS_EXCLUDED: tuple[str, ...] = (
    "terminated", "withdrawn", "rumoured", "rumored",
    "lapsed", "abandoned", "failed", "cancelled", "canceled",
)


def _deal_status_included(status: str) -> bool:
    """True if a deal with this status enters the CoTrans pool under the
    ``announced_or_closed`` default. Blank/unknown status → included (deep-
    research deals are announced by nature). Explicit terminated / withdrawn /
    rumoured / lapsed → excluded."""
    s = (status or "").strip().lower()
    if not s:
        return True
    return not any(x in s for x in _DEAL_STATUS_EXCLUDED)


def _query_tracker(
    workbook_path: Optional[Path], subsector_slug: str, *, lookback_years: int,
) -> list[dict[str, Any]]:
    """Query the canonical precedent tracker by subsector_slug + lookback,
    applying the ``announced_or_closed`` deal-status filter (#21-comps Q6).
    Returns a list of dicts with the lean-schema fields populated. Missing
    workbook → empty list (the caller's deep-research path fills the gap).
    """
    if not workbook_path or not Path(workbook_path).is_file():
        return []
    try:
        import openpyxl
        wb = openpyxl.load_workbook(str(workbook_path), read_only=True, data_only=True)
        # Use the canonical sheet name; tolerate operator-renamed first sheet.
        sheet_name = "Precedent transactions"
        ws = wb[sheet_name] if sheet_name in wb.sheetnames else wb.worksheets[0]
    except Exception as e:  # noqa: BLE001
        log.warning("comps: tracker read failed (%s); proceeding with no tracker rows", e)
        return []

    cutoff = date.today().year - lookback_years
    out: list[dict[str, Any]] = []
    headers: Optional[list[str]] = None
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            headers = [str(c) if c is not None else "" for c in row]
            continue
        if not row or not headers:
            continue
        rec = {headers[j]: row[j] for j in range(min(len(headers), len(row)))}
        ss = str(rec.get("Subsector (slug)") or "").strip().lower()
        if ss != subsector_slug.strip().lower():
            continue
        ann = rec.get("Announced Date")
        ann_iso = ""
        if isinstance(ann, datetime):
            ann_iso = ann.date().isoformat()
            if ann.year < cutoff:
                continue
        elif isinstance(ann, date):
            ann_iso = ann.isoformat()
            if ann.year < cutoff:
                continue
        # Q6 deal-status filter (announced_or_closed). Tracker carries no Status
        # column today → blank → included; the hook honours a future
        # "Status"/"Deal Status" column without a code change.
        if not _deal_status_included(
            str(rec.get("Status") or rec.get("Deal Status") or "")
        ):
            continue
        out.append({
            "announced_date": ann_iso,
            "target": str(rec.get("Target") or ""),
            "acquirer": str(rec.get("Acquirer") or ""),
            "buyer_type": str(rec.get("Acquirer type") or ""),
            "country": str(rec.get("Country") or ""),
            "currency": str(rec.get("Currency") or ""),
            "ev_m": _safe_float(rec.get("EV (m)")),
            "revenue_m": _safe_float(rec.get("Revenue (m)")),
            "ebitda_m": _safe_float(rec.get("EBITDA (m)")),
            "ev_revenue_x": _safe_float(rec.get("EV/Revenue")),
            "ev_ebitda_x": _safe_float(rec.get("EV/EBITDA")),
            "description": str(rec.get("Target Description") or rec.get("Deal Description") or ""),
            "strategic_commentary": str(rec.get("Strategic Commentary") or ""),
            "source": str(rec.get("Source") or ""),
            "deal_id": str(rec.get("Deal ID") or ""),
        })
    return out


def _load_tracker_rows_by_id(
    workbook_path: Optional[Path], deal_ids: list[str],
) -> dict[str, dict[str, Any]]:
    """Return a ``{deal_id: row-dict}`` lookup for the supplied ``deal_ids``
    from the canonical precedent tracker.

    Independent of subsector / lookback filters — Stage 2 already has the
    operator-approved deal_ids and just needs to hydrate the full row data
    (announced_date, target, acquirer, country, currency, ev_m, revenue_m,
    ebitda_m, description, strategic_commentary, source) for each. Missing
    workbook → empty dict (caller falls back to the back-reference source).
    Deal IDs not found in the tracker → absent from the dict (caller falls
    back to back-reference for that ID).

    Each returned dict also carries a ``_tracker_row`` key with the deal's
    1-indexed row number in the source xlsx — Stage 3's CoTrans Source-cell
    stamper uses it to construct a cross-workbook hyperlink location
    (``'Precedent transactions'!A<row>``) so clicking the cell in Excel jumps
    straight to the curated tracker row. ``_tracker_row`` is None when the
    workbook is missing or the deal_id isn't found (the latter case omits the
    deal from the dict entirely, but the key is documented for callers
    iterating the row dict shape).
    """
    if not workbook_path or not Path(workbook_path).is_file() or not deal_ids:
        return {}
    wanted = {str(d) for d in deal_ids if d}
    if not wanted:
        return {}
    try:
        import openpyxl
        wb = openpyxl.load_workbook(str(workbook_path), read_only=True, data_only=True)
        sheet_name = "Precedent transactions"
        ws = wb[sheet_name] if sheet_name in wb.sheetnames else wb.worksheets[0]
    except Exception as e:  # noqa: BLE001
        log.warning("comps: tracker read-by-id failed (%s); using back-references only", e)
        return {}

    out: dict[str, dict[str, Any]] = {}
    headers: Optional[list[str]] = None
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            headers = [str(c) if c is not None else "" for c in row]
            continue
        if not row or not headers:
            continue
        rec = {headers[j]: row[j] for j in range(min(len(headers), len(row)))}
        did = str(rec.get("Deal ID") or "").strip()
        if not did or did not in wanted:
            continue
        ann = rec.get("Announced Date")
        ann_iso = ""
        if isinstance(ann, datetime):
            ann_iso = ann.date().isoformat()
        elif isinstance(ann, date):
            ann_iso = ann.isoformat()
        # _tracker_row is the deal's 1-indexed row number in the xlsx
        # (header is row 1, first data row is row 2 — i+1 matches that).
        out[did] = {
            "deal_id": did,
            "announced_date": ann_iso,
            "target": str(rec.get("Target") or ""),
            "acquirer": str(rec.get("Acquirer") or ""),
            "buyer_type": str(rec.get("Acquirer type") or ""),
            "country": str(rec.get("Country") or ""),
            "currency": str(rec.get("Currency") or ""),
            "ev_m": _safe_float(rec.get("EV (m)")),
            "revenue_m": _safe_float(rec.get("Revenue (m)")),
            "ebitda_m": _safe_float(rec.get("EBITDA (m)")),
            "ev_revenue_x": _safe_float(rec.get("EV/Revenue")),
            "ev_ebitda_x": _safe_float(rec.get("EV/EBITDA")),
            "description": str(
                rec.get("Target Description") or rec.get("Deal Description") or ""
            ),
            "strategic_commentary": str(rec.get("Strategic Commentary") or ""),
            "source": str(rec.get("Source") or ""),
            "_tracker_row": i + 1,
        }
        if len(out) == len(wanted):
            break
    return out


def _safe_float(v: Any) -> Optional[float]:
    """Tolerant float-cast; '' / None / non-numeric → None."""
    if v is None or v == "":
        return None
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


# ─────────────────────────────────────────────────────────────────────────────
# Stage 2 — Acquire data, sourced
# ─────────────────────────────────────────────────────────────────────────────


@dataclass
class _AcquiredCoCo:
    """One CoCo row's acquired state (post operator's 2026-06-02 v2 restructure).

    Note: ``market_cap_m`` + ``ev`` are NO LONGER inputs — the template's
    column G (=E*F) computes Mkt Cap from shr_price * shares_out, and col I
    computes EV. The acquisition layer feeds the upstream inputs instead.
    The legacy fields are retained for backward compat with the headline-
    median computation (which falls back to deriving EV from Mkt Cap +
    Net Debt when the row's formula-derived EV isn't on the dict yet)."""
    subsector_slug: str
    ticker: str
    name: str
    currency: Optional[str]
    fs_currency: Optional[str]
    ccy_px: Optional[str]
    shr_price: Optional[float]
    shares_out: Optional[float]
    market_cap_m: Optional[float]      # legacy / derived; not stamped
    net_debt_m: Optional[float]
    fye: Optional[str]
    revenue_lfy_m: Optional[float]
    revenue_lfy1_m: Optional[float]
    ebitda_lfym1_m: Optional[float]
    ebitda_lfy_m: Optional[float]
    ebitda_lfy1_m: Optional[float]
    pe_ltm: Optional[float]
    source: str
    lfy1_source: str
    ccy_flags: list[dict[str, Any]] = field(default_factory=list)


def run_stage_2(
    inputs: CompsBuildInput,
    *,
    vault_root: Path,
    run_id: str,
    today: Optional[date] = None,
    tracker_workbook: Optional[Path] = None,
) -> StageResult:
    """Acquire CoCo + CoTrans data with sources. Surface every ccy mismatch +
    every unsourced LFY+1 as an approval-pending payload.

    The ``tracker_workbook`` arg lets the route point at the canonical
    precedent tracker (or a test fixture) so CoTrans rows are hydrated with
    the full deal data (announced_date / target / acquirer / etc.) plus the
    underlying URL when the tracker row carries one. Absent / unresolved →
    falls back to the back-reference source string and otherwise-empty deal
    rows (preserves pre-2026-06-02 behaviour for callers that haven't
    threaded the path through yet)."""
    if not inputs.approved_subsectors:
        raise MissingApprovalToken("Stage 2 requires approved_subsectors")
    if not inputs.approved_peers_by_subsector or not inputs.approved_deals_by_subsector:
        raise MissingApprovalToken(
            "Stage 2 requires approved_peers_by_subsector + approved_deals_by_subsector "
            "(from operator-signed Stage 1)"
        )
    # WS-B + #21-comps-step-3: verify BOTH Stage-1-issued tokens — peers +
    # deals are signed independently so a caller can't mutate one payload
    # while presenting the other's valid token. Cross-stage replay defence:
    # presenting a Stage-0 subsectors token here would mismatch (different
    # stage_label in the HMAC payload). Step-3 narrow-without-refire: when
    # the operator supplies proposed_peers_by_subsector /
    # proposed_deals_by_subsector, the token is verified over the FULL
    # proposed dict and approved must be a subset (DROPs OK; ADDs refused).
    ok_peers, err_peers = _verify_subset_approval(
        deal_name=inputs.deal_name,
        stage_label=STAGE_LABEL_PEERS,
        approved=inputs.approved_peers_by_subsector,
        proposed=inputs.proposed_peers_by_subsector,
        token=inputs.peers_approval_token or "",
    )
    if not ok_peers:
        raise MissingApprovalToken(
            f"Stage 2: invalid or missing peers_approval_token — {err_peers}. "
            "Either echo the Stage-1 proposal verbatim as "
            "approved_peers_by_subsector, OR send proposed_peers_by_subsector "
            "(the full Stage-1 proposal) + approved_peers_by_subsector (any "
            "subset of it) to narrow without re-firing."
        )
    ok_deals, err_deals = _verify_subset_approval(
        deal_name=inputs.deal_name,
        stage_label=STAGE_LABEL_DEALS,
        approved=inputs.approved_deals_by_subsector,
        proposed=inputs.proposed_deals_by_subsector,
        token=inputs.deals_approval_token or "",
    )
    if not ok_deals:
        raise MissingApprovalToken(
            f"Stage 2: invalid or missing deals_approval_token — {err_deals}. "
            "Either echo the Stage-1 proposal verbatim as "
            "approved_deals_by_subsector, OR send proposed_deals_by_subsector "
            "(the full Stage-1 proposal) + approved_deals_by_subsector (any "
            "subset of it) to narrow without re-firing."
        )
    today = today or datetime.now(timezone.utc).date()

    warnings: list[str] = []
    blocks: list[dict[str, Any]] = []
    unsourced_assumptions: list[dict[str, Any]] = []

    try:
        from routines.markets import get_provider
        provider = get_provider()
        provider_name = getattr(provider, "name", "provider")
    except Exception as e:  # noqa: BLE001
        warnings.append(f"markets provider unavailable ({e})")
        provider = None
        provider_name = "no-provider"

    # Resolve the tracker path: prefer the explicit arg (route / tests pass
    # tmp_path), else fall back to the canonical workbook so production
    # Stage 2 hydrates CoTrans rows from the real tracker. Missing module
    # (test env without dealtracker on path) → None; the lookup helper
    # gracefully degrades to back-reference-only.
    tracker_path: Optional[Path] = tracker_workbook
    if tracker_path is None:
        try:
            from routines.dealtracker.workbook import CANONICAL_WORKBOOK_PATH
            tracker_path = CANONICAL_WORKBOOK_PATH
        except Exception:  # noqa: BLE001
            tracker_path = None

    # Hydrate ALL approved deal_ids across subsectors in one read.
    all_deal_ids: list[str] = []
    for ss in inputs.approved_subsectors:
        all_deal_ids.extend((inputs.approved_deals_by_subsector or {}).get(ss, []))
    tracker_by_id = _load_tracker_rows_by_id(tracker_path, all_deal_ids)

    for subsector_slug in inputs.approved_subsectors:
        tickers = list((inputs.approved_peers_by_subsector or {}).get(subsector_slug, []))
        deal_ids = list((inputs.approved_deals_by_subsector or {}).get(subsector_slug, []))
        coco_rows: list[dict[str, Any]] = []
        ccy_flags_block: list[dict[str, Any]] = []

        # CoCo acquisition — per peer. `assumptions` is a list of 0-2
        # operator-gated assumptions per row (LFY+1 and/or EBITDAaL); both
        # flow into the Stage-2 unsourced_assumptions list signed by the
        # assumptions_approval_token via `_assumption_identity`.
        for tk in tickers:
            row, flags, assumptions_for_row = _acquire_coco_row(
                tk, subsector_slug, provider=provider, provider_name=provider_name,
                today=today,
            )
            coco_rows.append(row)
            ccy_flags_block.extend(flags)
            unsourced_assumptions.extend(assumptions_for_row)

        # CoTrans acquisition — per approved deal. Hydrate FULL deal data
        # from the tracker lookup so the stamper writes populated cells
        # (announced_date / target / acquirer / etc.) — NOT just deal_id +
        # source.
        #
        # Source-cell hyperlink semantics (operator clarification
        # 2026-06-02, tracker-first priority — see _write_source_cell):
        #   * Deal IS in the tracker (the common case) → propagate
        #     ``_tracker_row`` (the deal's 1-indexed row number in the
        #     canonical xlsx) through to the stamper. The Stage 3 stamper
        #     hyperlinks the cell to the canonical tracker file at that
        #     row — the operator's curated tracker carries their own
        #     source links + context, so the hyperlink lands them inside
        #     their tool not on a 3rd-party paywall (e.g. Mergermarket).
        #   * Deal NOT in the tracker (orphan approved deal_id) → no
        #     ``_tracker_row``; the source field's URL (if any) is the
        #     only signal the stamper can use for the rare-fallback web
        #     URL hyperlink path.
        cotrans_rows: list[dict[str, Any]] = []
        for did in deal_ids:
            tracker_row = tracker_by_id.get(did)
            if tracker_row is not None:
                # ``dict(tracker_row)`` carries the ``_tracker_row`` key set
                # by ``_load_tracker_rows_by_id`` — the stamper reads it to
                # build the tracker-first hyperlink.
                row_dict = dict(tracker_row)
                src = (row_dict.get("source") or "").strip()
                if not src.startswith(("http://", "https://")):
                    row_dict["source"] = f"tracker:{did}"
                row_dict["stamped"] = True
                cotrans_rows.append(row_dict)
            else:
                # Deal ID approved but not found in tracker (cache miss /
                # operator-entered ID without a tracker row yet) — keep the
                # back-reference so the stamper still writes a recognisable
                # Source cell. ``_tracker_row`` is absent so the stamper
                # falls to priority 2 (URL fallback) or 3 (plain text).
                cotrans_rows.append({
                    "deal_id": did,
                    "source": f"tracker:{did}",
                    "stamped": True,
                })

        blocks.append({
            "subsector_slug": subsector_slug,
            "coco_rows": coco_rows,
            "cotrans_rows": cotrans_rows,
            "ccy_flags": ccy_flags_block,
            "unsourced_lfy1": [a for a in unsourced_assumptions
                               if a.get("subsector_slug") == subsector_slug],
        })

    # WS-B + WS-C: Stage 2 ALWAYS issues the stage_2_blocks_approval_token
    # signed over the acquired blocks (this IS the cache key per the brief's
    # composition). WS-C writes the blocks into the in-process cache keyed by
    # this token so Stage 3 reads them back WITHOUT a provider re-call.
    stage_2_blocks_token = _sign_token(
        inputs.deal_name, STAGE_LABEL_STAGE_2_BLOCKS, blocks,
    )
    _cache_stage_2_blocks(stage_2_blocks_token, blocks)

    if unsourced_assumptions:
        payload = ApprovalPayload(
            kind="assumptions",
            proposed=unsourced_assumptions,
        )
        # Assumptions token — signed over the proposed-assumptions payload so
        # mutation between Stage 2 → Stage 3 is detected. The stage_2_blocks
        # token is independent (binds to the row data, not the assumptions).
        assumptions_token = _sign_token(
            inputs.deal_name, STAGE_LABEL_ASSUMPTIONS,
            _assumption_identity(unsourced_assumptions),
        )
        combined = f"assumptions={assumptions_token}|stage_2_blocks={stage_2_blocks_token}"
        return StageResult(
            ok=True,
            stage="approval_pending",
            stage_just_completed=2,
            deal_name=inputs.deal_name,
            target=inputs.target,
            run_id=run_id,
            approval_payload=payload,
            approval_token_to_sign=combined,
            warnings=warnings,
            approved_subsectors=list(inputs.approved_subsectors),
            blocks=blocks,
            assumptions_approval_token=assumptions_token,
            stage_2_blocks_approval_token=stage_2_blocks_token,
        )

    # Nothing to approve → Stage 2 complete, ready for Stage 3 with just the
    # stage_2_blocks token (the cache key).
    return StageResult(
        ok=True,
        stage="approval_pending",
        stage_just_completed=2,
        deal_name=inputs.deal_name,
        target=inputs.target,
        run_id=run_id,
        approval_token_to_sign=f"stage_2_blocks={stage_2_blocks_token}",
        warnings=warnings,
        approved_subsectors=list(inputs.approved_subsectors),
        blocks=blocks,
        stage_2_blocks_approval_token=stage_2_blocks_token,
    )


def _self_historical_growth(
    lfy: Optional[float], lfym1: Optional[float],
) -> Optional[float]:
    """The company's own realised LFY-over-LFY-1 growth: ``lfy/lfym1 - 1``.

    Returns None when either period is missing or the prior year is
    non-positive (a negative/zero base makes the growth rate meaningless).
    Used by the LFY+1 self-historical proxy (#21-comps Q2, decision
    2026-06-02) — keeps each metric's proxy traceable to its own filings.
    """
    if lfy is None or lfym1 is None or lfym1 <= 0:
        return None
    return (lfy / lfym1) - 1.0


def _acquire_coco_row(
    ticker: str, subsector_slug: str, *,
    provider: Any, provider_name: str, today: date,
) -> tuple[dict[str, Any], list[dict[str, Any]], list[dict[str, Any]]]:
    """Fetch one CoCo row with sources. Returns ``(row, ccy_flags, assumptions)``.

    ``assumptions`` is a list of 0-2 operator-gated assumptions per row
    (#21-comps-step-4, 2026-06-03 — signature changed from a single
    ``Optional[dict]`` to support both LFY+1 and EBITDAaL surfaces
    simultaneously). Possible entries:
      * LFY+1 (field ``"revenue_lfy1_m + ebitda_lfy1_m"``) — surfaced when
        the row has LFY revenue or EBITDA data.
      * EBITDAaL (field ``"ebitdaal_lfym1_m + ebitdaal_lfy_m +
        ebitdaal_lfy1_m"``) — surfaced when the provider reports
        ``capital_lease_obligations > 0`` AND any EBITDA value.

    Post operator's 2026-06-02 v2 template restructure:

      Acquired inputs (paste into template):
        * shr_price        ← provider.info.currentPrice (best-effort)
        * shares_out       ← provider.info.sharesOutstanding (best-effort)
        * net_debt_m       ← IR scrape (follow-on; left None for now)
        * fye              ← provider.info.lastFiscalYearEnd ISO date
        * revenue_lfy_m    ← Fundamentals.years[0].revenue
        * ebitda_lfym1_m   ← Fundamentals.years[1].ebitda (prior year)
        * ebitda_lfy_m     ← Fundamentals.years[0].ebitda
        * pe_ltm           ← FundamentalsRatios.pe (Yahoo trailingPE direct)
        * ccy_px           ← Quote.currency (price provider's currency)
        * ccy_fs           ← Fundamentals.currency (FS reporting currency)

      Formula-derived in the template (NOT acquired):
        * Mkt Cap (G = E*F)
        * EV (I = SUM(G:H))
        * EV/EBITDA, EV/Revenue (S/T/U/W/X)
        * FX? (AD)

      Surfaced for operator approval (not silently filled):
        * revenue_lfy1_m   ← LFY+1 consensus; operator-approves
        * ebitda_lfy1_m    ← LFY+1 consensus; operator-approves

    The optional provider fields (currentPrice / sharesOutstanding / trailingPE /
    lastFiscalYearEnd) are accessed via ``getattr`` so that providers exposing
    only the canonical ``Fundamentals`` Protocol still work — those fields land
    as None and the operator pastes the values directly.
    """
    ccy_flags: list[dict[str, Any]] = []
    assumption: Optional[dict[str, Any]] = None
    trading_ccy: Optional[str] = None
    fs_ccy: Optional[str] = None
    shr_price: Optional[float] = None
    shares_out: Optional[float] = None
    market_cap: Optional[float] = None
    total_debt_bs: Optional[float] = None
    lease_bs: Optional[float] = None
    cash_bs: Optional[float] = None
    fye_iso: Optional[str] = None
    pe_ltm: Optional[float] = None
    rev_lfy: Optional[float] = None
    rev_lfym1: Optional[float] = None
    rev_lfy1: Optional[float] = None
    ebitda_lfym1: Optional[float] = None
    ebitda_lfy: Optional[float] = None
    ebitda_lfy1: Optional[float] = None
    name: str = ticker

    quote_source = f"{provider_name}:{today.isoformat()}"
    funda_source = quote_source

    if provider is not None:
        try:
            quotes = provider.get_quotes([ticker])
            if quotes:
                q = quotes[0]
                trading_ccy = q.currency
                name = q.name or ticker
                # Best-effort numeric price extraction. Quote.price is a
                # display string ("485p", "$82.15"); providers that expose
                # the raw numeric current price as an extra attr land it
                # via getattr fallback.
                shr_price = _safe_float(getattr(q, "current_price", None))
                if shr_price is None:
                    shr_price = _safe_float(getattr(q, "price_numeric", None))
        except Exception as e:  # noqa: BLE001
            log.warning("comps: provider.get_quotes(%s) failed: %s", ticker, e)
        try:
            f = provider.get_fundamentals(ticker, years=3)
            fs_ccy = f.currency
            name = f.name or name
            # FYE — providers may expose either a month string ("December") or
            # an ISO date. Surface the raw value either way.
            fye_iso = getattr(f, "fiscal_year_end", None)
            if f.years:
                lfy = f.years[0]
                rev_lfy = lfy.revenue
                ebitda_lfy = lfy.ebitda
                # Balance-sheet debt + leases + cash → net debt incl/excl leases.
                total_debt_bs = _safe_float(lfy.total_debt)
                lease_bs = _safe_float(getattr(lfy, "capital_lease_obligations", None))
                cash_bs = _safe_float(lfy.cash_and_equivalents)
                # Prior-year figures = LFY-1. EBITDA LFY-1 is template column O;
                # Revenue LFY-1 is NOT a template column but is carried on the
                # row so the LFY+1 self-historical growth proxy can compute g
                # from the company's own filings (#21-comps Q2).
                if len(f.years) >= 2:
                    ebitda_lfym1 = f.years[1].ebitda
                    rev_lfym1 = f.years[1].revenue
            # P/E LTM — Yahoo's trailingPE lives on FundamentalsRatios.pe.
            if f.ratios is not None:
                pe_ltm = f.ratios.pe
            # Best-effort optional fields some providers attach to the
            # Fundamentals payload (shares outstanding most commonly).
            shares_out = _safe_float(getattr(f, "shares_outstanding", None))
            if shr_price is None:
                shr_price = _safe_float(getattr(f, "current_price", None))
        except Exception as e:  # noqa: BLE001
            log.warning("comps: provider.get_fundamentals(%s) failed: %s", ticker, e)

        # Market cap (+ shares) for the CoCo Mkt Cap input. Fetched via the
        # provider's optional ``market_cap_and_shares`` method — present on the
        # OpenBB provider, absent on minimal stubs (degrades to operator-paste).
        # Direct market cap is deliberately preferred over price×shares: the
        # quote path exposes only a display-string price, and price×shares
        # mis-scales pence-quoted UK stocks (GBp price × shares ≠ GBP mkt cap).
        mc_fn = getattr(provider, "market_cap_and_shares", None)
        if callable(mc_fn):
            try:
                mc, so = mc_fn(ticker)
                market_cap = _safe_float(mc)
                if so is not None and shares_out is None:
                    shares_out = _safe_float(so)
            except Exception as e:  # noqa: BLE001
                log.warning("comps: market_cap_and_shares(%s) failed: %s", ticker, e)

    # Currency-mismatch flags. The unit-mismatch (GBp vs GBP) is the most
    # common — surface as a distinct flag kind so the route can render the
    # operator-chooses-display chip.
    if trading_ccy and fs_ccy and trading_ccy != fs_ccy:
        if trading_ccy.lower() in ("gbp", "gbx", "gbp_p", "p") and fs_ccy.upper() == "GBP":
            ccy_flags.append({
                "ticker": ticker, "kind": "unit_p_vs_pound",
                "trading_currency": trading_ccy, "fs_currency": fs_ccy,
            })
        else:
            ccy_flags.append({
                "ticker": ticker, "kind": "fs_vs_trading",
                "trading_currency": trading_ccy, "fs_currency": fs_ccy,
            })

    # net_debt — from the provider's balance sheet, split INCL vs EXCL IFRS-16
    # leases (the operator's restructured v2 template carries both columns).
    #   net_debt_incl = total_debt − cash           (total_debt is lease-inclusive)
    #   net_debt_excl = total_debt − leases − cash   (financial net debt only)
    # When the provider reports no lease line, excl == incl (assume no leases).
    # Sourced to ``<provider>:<as_of>``. IRON LAW: a populated net-debt value ⇒
    # net_debt_source set; absent both ⇒ None (guard skips the source check).
    # The lease COST (for EBITDAaL) is NOT reliably available from the free
    # provider, so EBITDAaL stays an operator input.
    net_debt: Optional[float] = None        # backward-compat = incl-leases
    net_debt_incl: Optional[float] = None
    net_debt_excl: Optional[float] = None
    net_debt_source = ""
    if total_debt_bs is not None and cash_bs is not None:
        net_debt_incl = total_debt_bs - cash_bs
        net_debt_excl = total_debt_bs - (lease_bs or 0.0) - cash_bs
        net_debt = net_debt_incl
        net_debt_source = funda_source

    # Normalize CoCo financial magnitudes to MILLIONS to match the template's
    # "(m)" columns + align with the CoTrans block (whose EVs are already in
    # £m). The provider returns full currency units (e.g. revenue 466,403,000);
    # the headline-median + multiple maths are ratios → scale-invariant, so
    # this is a display/consistency fix, not a valuation change. Applies to the
    # amount fields only — NOT to per-share price or P/E (which are not "(m)").
    def _to_m(x: Optional[float]) -> Optional[float]:
        return (x / 1_000_000.0) if x is not None else None
    rev_lfy = _to_m(rev_lfy)
    rev_lfym1 = _to_m(rev_lfym1)
    ebitda_lfy = _to_m(ebitda_lfy)
    ebitda_lfym1 = _to_m(ebitda_lfym1)
    market_cap = _to_m(market_cap)
    net_debt = _to_m(net_debt)
    net_debt_incl = _to_m(net_debt_incl)
    net_debt_excl = _to_m(net_debt_excl)

    # Per-row assumptions list. Stage 2 collects assumptions across all rows
    # then surfaces them as a single approval_pending payload. Each assumption
    # carries (subsector_slug, ticker, field) — the identity the assumptions
    # token binds to via `_assumption_identity`.
    assumptions: list[dict[str, Any]] = []

    # LFY+1 — surfaced as a structured assumption-pending entry (NOT silently
    # filled): the operator chooses one of three options (#21-comps Q2,
    # decision 2026-06-02). lfy1_source stays empty until the operator approves.
    lfy1_source = ""
    if rev_lfy is not None or ebitda_lfy is not None:
        g_rev = _self_historical_growth(rev_lfy, rev_lfym1)
        g_ebitda = _self_historical_growth(ebitda_lfy, ebitda_lfym1)
        assumptions.append({
            "subsector_slug": subsector_slug,
            "ticker": ticker,
            "field": "revenue_lfy1_m + ebitda_lfy1_m",
            "candidate_source": "yahoo-consensus" if provider_name.startswith("openbb") else "n/a",
            # The three operator options + the data needed to choose between them.
            "options": ["blank", "self_historical", "operator_input"],
            "lfy_values": {
                "revenue_lfy_m": rev_lfy, "revenue_lfym1_m": rev_lfym1,
                "ebitda_lfy_m": ebitda_lfy, "ebitda_lfym1_m": ebitda_lfym1,
            },
            "self_historical_growth": {"revenue": g_rev, "ebitda": g_ebitda},
            "ask": (
                "No consensus LFY+1 — choose one (Q2): "
                "(a) 'blank' = leave empty; "
                "(b) 'self_historical' = apply the company's own realised "
                "LFY-over-LFY-1 growth; "
                "(c) 'operator_input' = supply growth_rate + a REQUIRED "
                "justification (un-justified rates are refused). Approved "
                f"values are tagged operator-approved:...:{today.isoformat()}."
            ),
        })

    # EBITDAaL (#21-comps-step-4, 2026-06-03) — surfaced as an operator-gated
    # assumption when the peer has IFRS-16 lease exposure (lease_liability > 0)
    # AND we have any EBITDA value to anchor the proxy commentary. EBITDAaL =
    # EBITDA after lease costs; lease COST (depreciation of RoU assets +
    # interest on lease liabilities) isn't reliably on the balance sheet, so
    # the free-provider path can't auto-compute. Operator has THREE paths
    # (orchestrator session drives them via chat — see
    # session-briefs/SESSION-COMPS-ORCHESTRATION.md §EBITDAaL):
    #   * 'blank' (default) — leave None; operator fills in Excel from the
    #     IFRS-16 note. Iron Law passes (no value ⇒ no source needed).
    #   * 'operator_input' — operator supplies 3 values + a REQUIRED
    #     justification. Justification can be the operator's own reasoning OR a
    #     deep-research source URL (option-iii in the brief: orchestrator
    #     session invokes deep-research against the company's IR portal /
    #     annual report, surfaces EBITDAaL or its components, threads back as
    #     operator_input with the source URL as justification).
    # No silent proxy applied — operator vetoed baking a default ratio
    # (2026-06-03 design conversation); the lease_liability + ebitda numbers
    # are surfaced so the operator decides with context, not blind.
    lease_liability_m = _to_m(lease_bs) if lease_bs is not None else None
    if (lease_liability_m and lease_liability_m > 0
            and (ebitda_lfy is not None or ebitda_lfym1 is not None)):
        assumptions.append({
            "subsector_slug": subsector_slug,
            "ticker": ticker,
            # Field identity covers all 3 EBITDAaL periods (LFY-1, LFY, LFY+1)
            # — operator decides for all three at once.
            "field": "ebitdaal_lfym1_m + ebitdaal_lfy_m + ebitdaal_lfy1_m",
            "candidate_source": (
                "balance-sheet-lease-liability"
                if provider_name.startswith("openbb") else "n/a"
            ),
            "options": ["blank", "operator_input"],
            "lease_data": {
                "lease_liability_m": lease_liability_m,
                # The implied annual lease cost depends on (weighted-lease-term
                # + interest_rate). Surfaced as context — not a baked default —
                # so the operator decides the ratio.
            },
            "ebitda_values": {
                "ebitda_lfym1_m": ebitda_lfym1,
                "ebitda_lfy_m": ebitda_lfy,
                "ebitda_lfy1_m": ebitda_lfy1,
            },
            "ask": (
                f"EBITDAaL not pullable from free provider for {ticker}: "
                f"lease COST (depreciation of RoU + interest on lease "
                f"liabilities) isn't on the balance sheet. Lease liability = "
                f"{lease_liability_m:.1f}m; EBITDA LFY = "
                f"{ebitda_lfy if ebitda_lfy is not None else 'n/a'}m. "
                "Three operator paths — drive via the orchestrator session: "
                "(i) supply 3 values + justification ('operator_input'); "
                "(ii) leave blank, fill in Excel from the IFRS-16 note ('blank'); "
                "(iii) invoke deep-research against the company's IR portal / "
                "annual report for EBITDAaL or lease-cost components, then "
                "thread the finding back as 'operator_input' with the source URL "
                "as justification."
            ),
        })

    row = {
        "ticker": ticker,
        "name": name,
        # Trading currency = the price-quote provider's currency (CCY (Px)).
        "currency": trading_ccy,
        "ccy_px": trading_ccy,
        # FS currency = the fundamentals provider's reporting currency.
        "fs_currency": fs_ccy,
        # New v2-template inputs:
        "shr_price": shr_price,
        "shares_out": shares_out,
        "fye": fye_iso,
        "pe_ltm": pe_ltm,
        # Net debt — split incl/excl IFRS-16 leases (v2 lease columns H/I).
        # net_debt_m retained = incl-leases (headline-median basis + back-compat).
        "net_debt_m": net_debt,
        "net_debt_incl_m": net_debt_incl,
        "net_debt_excl_m": net_debt_excl,
        # EBITDAaL — operator input (lease cost not reliably pullable); the
        # acquisition leaves these None so the template's EV/EBITDAaL stays
        # blank until the operator enters EBITDAaL from the IFRS-16 note.
        "ebitdaal_lfym1_m": None,
        "ebitdaal_lfy_m": None,
        "ebitdaal_lfy1_m": None,
        "revenue_lfy_m": rev_lfy,
        # revenue_lfym1_m: LFY-1 revenue — carried for the LFY+1 self-historical
        # growth proxy (Q2); NOT a template column, so it is never stamped.
        "revenue_lfym1_m": rev_lfym1,
        "revenue_lfy1_m": rev_lfy1,
        "ebitda_lfym1_m": ebitda_lfym1,
        "ebitda_lfy_m": ebitda_lfy,
        "ebitda_lfy1_m": ebitda_lfy1,
        # Mkt Cap (direct from provider profile; stamped into template col G,
        # overriding the price×shares formula to dodge the GBp/GBP unit trap).
        "market_cap_m": market_cap,
        # Source: provider + today (Iron Law mechanical guard accepts the
        # `<provider>:<as_of>` form).
        "source": funda_source,
        "net_debt_source": net_debt_source,
        "lfy1_source": lfy1_source,
        # EBITDAaL provenance — empty until the operator approves a non-blank
        # path at Stage 2. Stamped into the Strategic Commentary column at
        # Stage 3 when populated, so the operator can audit how each row's
        # EBITDAaL was sourced (operator-typed / deep-research URL / Excel-
        # filled).
        "ebitdaal_source": "",
    }
    return row, ccy_flags, assumptions


# ─────────────────────────────────────────────────────────────────────────────
# Stage 3 — Populate + deliver (stamp + archive + capture + mirror)
# ─────────────────────────────────────────────────────────────────────────────


def deal_valuation_folder(deal_name: str) -> Path:
    """Per workspace-write-policy: the deal's Valuation/01. COMPS/ folder."""
    return (
        Path("<workspace-root>/1. Projects") / deal_name
        / "3. Financials & analysis" / "2. Valuation" / "01. COMPS"
    )


def comps_output_filename(deal_name: str, today: date, version: int = 1) -> str:
    return f"Project_{deal_name}_COMPS_{today.isoformat()}_v{version}.xlsx"


def _next_version(target_dir: Path, deal_name: str, today: date) -> int:
    """Return the next version number for today's stamp in ``target_dir``."""
    prefix = f"Project_{deal_name}_COMPS_{today.isoformat()}_v"
    existing = [
        int(p.stem.removeprefix(prefix))
        for p in target_dir.glob(f"{prefix}*.xlsx")
        if p.stem.removeprefix(prefix).isdigit()
    ]
    return (max(existing) + 1) if existing else 1


def _archive_prior_same_day(target_dir: Path, deal_name: str, today: date) -> Optional[Path]:
    """Move the most-recent same-day same-skill version (if any) to 00. OLD/.
    Returns the archived path, or None if no prior version existed."""
    archive_dir = target_dir / "00. OLD"
    prefix = f"Project_{deal_name}_COMPS_{today.isoformat()}_v"
    candidates = sorted(target_dir.glob(f"{prefix}*.xlsx"))
    if not candidates:
        return None
    # Move all prior versions to keep the lineage intact.
    archive_dir.mkdir(parents=True, exist_ok=True)
    last_moved: Optional[Path] = None
    for c in candidates:
        dest = archive_dir / c.name
        shutil.move(str(c), str(dest))
        last_moved = dest
    return last_moved


def _emit_sector_mirror_proposal(
    vault_root: Path, parent_sector: str, deal_name: str, template_path: Path,
    today: date, *, run_id: str = "",
    ev_ebitda_median: Optional[float] = None,
    ev_revenue_median: Optional[float] = None,
    peer_count: int = 0, deal_count: int = 0,
    subsectors: Optional[list[str]] = None,
    peer_tickers: Optional[list[str]] = None,
) -> Optional[Path]:
    """Emit an operator-gated ``deliverable-outcome`` proposal that, on operator
    Route, appends a dated **comps-run snapshot** bullet to the ``## Comps runs``
    section of ``Sectors/<sector>/Comps.md`` pointing at the new workbook.

    A comps run is a *valuation snapshot* (peer-set trading multiples as-of a
    date) — NOT a *precedent transaction* (one deal's terms). The two are
    different content types living in different sections of the canonical
    ``sector-claim`` note: precedent transactions under ``## Precedent
    transactions`` (``### comp-<id>`` blocks, fed by #43 deal capture) and
    comps-run snapshots under ``## Comps runs`` (flat dated bullets, fed HERE).
    So the snapshot stays a FLAT bullet — it does NOT use the ``body_md``
    structured-block branch of ``_route_deliverable_outcome`` (#43-sector-
    template-align §2).

    Closes the #61 capability gap (#61-capabilities): the comps skill is
    ``workspace_scope: project`` and the registry validator
    (``_validate_capabilities``) rejects ``vault_write`` outside ``Projects/**``
    on project-scoped skills. The pre-fix code wrote the mirror DIRECTLY to
    ``Sectors/<sector>/Comps.md`` — a runtime write outside the declared
    surface. Routing through ``Routines/deliverable-outcomes/`` means there's
    NO ungated runtime write into ``Sectors/``; the operator's Route action on
    the proposal does the append (via ``_route_deliverable_outcome``), and
    that path is governed by the proposals lifecycle, not by the skill's
    declared capability surface.

    The proposal's frontmatter declares ``target: Sectors/<sector>/Comps.md``
    + ``section: Comps runs`` so the existing ``_route_deliverable_outcome``
    handler appends the bullet under the correct section on Route, creating
    the sector note from the ``sector-comps`` template if missing (the route's
    create-from-template is path-aware — see ``_new_note_from_template``).
    Idempotent: if the proposal already exists in operator-triaged state
    (applied/routed/rejected), it is left alone.

    Returns the proposal-file path, or ``None`` if a same-day same-deal
    proposal has already been operator-triaged.
    """
    import frontmatter as _fm

    from routines.skills._runtime.capture import (
        PROPOSAL_DIR_REL, SKIP_STATUSES, _slug,
    )

    sector_slug = parent_sector.strip().lower().replace(" ", "-")
    target_rel = f"Sectors/{sector_slug}/Comps.md"

    # ── Snapshot bullet (#43-sector-template-align §2): date · subject deal ·
    # peer set · median EV/EBITDA + EV/Rev · subsector · → deliverable link.
    # The deliverable is embedded as a markdown link (filename text → full
    # path href) so the rendered bullet reads cleanly; `workspace_artefact` is
    # therefore left empty (below) to avoid `_format_fact_bullet` repeating the
    # path as a separate `artefact:` tail. The leading date (`date:` frontmatter
    # = the as-of date) is the v1 staleness signal — NOT a note-level `expires:`
    # (a sector note accrues many snapshots; #54a `expires` is note-level and
    # would decay the whole note — wrong; see §3).
    def _fmt_x(v: Optional[float]) -> str:
        return f"{v:g}x" if isinstance(v, (int, float)) else "n/a"

    peer_set = f" ({'/'.join(peer_tickers)})" if peer_tickers else ""
    subsector_label = ", ".join(dict.fromkeys(subsectors)) if subsectors else "—"
    # Angle-bracket the href — deliverable paths contain spaces (`1. Projects`,
    # `3. Financials & analysis`), which a bare markdown link would break.
    deliverable_link = f"[`{template_path.name}`](<{template_path.as_posix()}>)"
    headline = (
        f"{deal_name} comps · {peer_count} CoCo peers{peer_set} "
        f"+ {deal_count} CoTrans · median **EV/EBITDA {_fmt_x(ev_ebitda_median)}**, "
        f"**EV/Rev {_fmt_x(ev_revenue_median)}** · subsector: {subsector_label} "
        f"· → {deliverable_link}"
    )

    # Same naming convention as emit_deliverable_proposal so the proposal
    # file slot is stable across re-runs (per-day, per-deal, per-skill —
    # but we suffix with -sector-mirror so it doesn't collide with the
    # primary Companies/<target>.md capture proposal).
    filename = (
        f"{today.isoformat()}-{_slug(deal_name)}-comps-sector-mirror.md"
    )
    proposal_path = vault_root / PROPOSAL_DIR_REL / filename

    if proposal_path.is_file():
        try:
            existing = _fm.load(proposal_path)
            status = str(existing.metadata.get("status") or "").strip().lower()
            if status in SKIP_STATUSES:
                log.info(
                    "comps: skipping sector-mirror proposal %s — status=%r",
                    proposal_path, status,
                )
                return None
        except Exception as e:  # noqa: BLE001 — unreadable → overwrite
            log.warning(
                "comps: failed to parse existing sector-mirror proposal %s "
                "(%s) — overwriting", proposal_path, e,
            )

    body = (
        f"# {headline}\n\n"
        f"## Captured conclusion\n\n"
        f"{headline}\n\n"
        f"On **Route**, this conclusion is appended as a dated, sourced "
        f"comps-run snapshot bullet to `{target_rel}` under the `Comps runs` "
        f"section — append-only, never overwriting prior snapshots (CLAUDE.md "
        f"§3 rule 9). The Sectors/ write is operator-gated: this skill never "
        f"writes directly into `Sectors/**` (#61-capabilities).\n\n"
        f"*Emitted by the `comps` deliverable→vault capture loop (#76 sector-"
        f"mirror sibling) on {today.isoformat()}. Routes through the standard "
        f"proposals lifecycle (#8 + #58): Review and Route / Reject / Skip / "
        f"Request revision.*\n"
    )
    post = _fm.Post(body)
    post.metadata["type"] = "deliverable-outcome"
    post.metadata["kind"] = "deliverable-outcome"
    post.metadata["status"] = "pending-review"
    post.metadata["date"] = today.isoformat()
    post.metadata["skill"] = "comps"
    post.metadata["target"] = target_rel
    post.metadata["section"] = "Comps runs"
    post.metadata["headline"] = headline
    post.metadata["provenance"] = f"runs:skill.comps.{run_id}"
    # Deliverable is embedded as a markdown link inside `headline`; leave
    # `workspace_artefact` empty so `_route_deliverable_outcome` /
    # `_format_fact_bullet` does NOT repeat the path as a separate `artefact:`
    # tail (the bullet would otherwise carry the deliverable twice).
    post.metadata["workspace_artefact"] = ""
    post.metadata["run_id"] = run_id
    post.metadata["sensitivity"] = "internal"
    post.metadata["tldr"] = headline

    proposal_path.parent.mkdir(parents=True, exist_ok=True)
    serialised = _fm.dumps(post) + "\n"
    tmp = proposal_path.with_suffix(proposal_path.suffix + ".tmp")
    tmp.write_text(serialised, encoding="utf-8")
    tmp.replace(proposal_path)
    log.info("comps: emitted sector-mirror proposal %s", proposal_path)
    return proposal_path


def _coco_rows_have_sources(blocks: list[dict[str, Any]]) -> tuple[bool, int, list[str]]:
    """Iron Law piece 1 — mechanical pre-stamp guard. Returns
    ``(all_sourced, rows_checked, offending_descriptions)``."""
    offending: list[str] = []
    checked = 0
    for block in blocks:
        ss = block.get("subsector_slug") or "?"
        for row in block.get("coco_rows") or []:
            checked += 1
            src = row.get("source")
            if not _looks_sourced(src):
                offending.append(f"CoCo {ss}/{row.get('ticker', '?')}: source={src!r}")
            # The LFY+1 source is allowed to be empty WHEN the row's
            # revenue_lfy1/ebitda_lfy1 are also empty (operator dropped the
            # row at Stage 2). If those are populated, lfy1_source must be
            # sourced.
            if (row.get("revenue_lfy1_m") is not None or row.get("ebitda_lfy1_m") is not None):
                if not _looks_sourced(row.get("lfy1_source")):
                    offending.append(
                        f"CoCo {ss}/{row.get('ticker', '?')}: lfy1_source missing"
                    )
            if (row.get("net_debt_m") is not None
                    and not _looks_sourced(row.get("net_debt_source"))):
                offending.append(
                    f"CoCo {ss}/{row.get('ticker', '?')}: net_debt_source missing"
                )
        for row in block.get("cotrans_rows") or []:
            checked += 1
            if not _looks_sourced(row.get("source")):
                offending.append(f"CoTrans {ss}/{row.get('deal_id', '?')}: source={row.get('source')!r}")
    return (len(offending) == 0, checked, offending)


_LFY1_FIELD = "revenue_lfy1_m + ebitda_lfy1_m"
_EBITDAAL_FIELD = "ebitdaal_lfym1_m + ebitdaal_lfy_m + ebitdaal_lfy1_m"


def _apply_approved_assumptions(
    blocks: list[dict[str, Any]],
    approved: list[dict[str, Any]],
    today: date,
) -> None:
    """Mutate ``blocks`` in place: apply each operator-approved assumption to
    the matching CoCo row. Routes by ``field`` between the LFY+1 family
    (#21-comps Q2, decision 2026-06-02; three choices = blank /
    self_historical / operator_input) and the EBITDAaL family
    (#21-comps-step-4, 2026-06-03; two choices = blank / operator_input).
    Unknown field types are refused (never silently ignore an approval).

    See ``_apply_lfy1_assumption`` and ``_apply_ebitdaal_assumption`` for the
    per-family contracts."""
    by_key = {
        (a.get("subsector_slug"), a.get("ticker"), a.get("field")): a
        for a in approved
    }
    for block in blocks:
        ss = block.get("subsector_slug")
        for row in block.get("coco_rows") or []:
            ticker = row.get("ticker")
            # Route per-field. A row can carry BOTH an LFY+1 and an EBITDAaL
            # assumption — apply each if present (independent paths).
            lfy1_entry = by_key.get((ss, ticker, _LFY1_FIELD))
            if lfy1_entry is not None:
                _apply_lfy1_assumption(row, lfy1_entry, today, ss=ss)
            ebitdaal_entry = by_key.get((ss, ticker, _EBITDAAL_FIELD))
            if ebitdaal_entry is not None:
                _apply_ebitdaal_assumption(row, ebitdaal_entry, today, ss=ss)
            # Anything that doesn't match a known field but IS keyed against
            # this row is a programming error — refuse.
            for key in [
                k for k in by_key
                if k[0] == ss and k[1] == ticker and k[2] not in (_LFY1_FIELD, _EBITDAAL_FIELD)
            ]:
                raise UnsourcedFigureError(
                    f"assumption for {ss}/{ticker}: unknown field "
                    f"{key[2]!r} (expected one of: {_LFY1_FIELD!r}, "
                    f"{_EBITDAAL_FIELD!r})."
                )


def _apply_lfy1_assumption(
    row: dict[str, Any],
    entry: dict[str, Any],
    today: date,
    *,
    ss: Optional[str] = None,
) -> None:
    """Per-row LFY+1 application. Three choices:

      * ``"blank"``           → leave LFY+1 empty; no source set (the Iron Law
                                needs a source only when a value is present).
      * ``"self_historical"`` → LFY+1 = LFY × (1 + g), g = the row's own realised
                                LFY/LFY-1 growth (per metric); source tagged
                                ``operator-approved:self-historical-growth:<today>``.
      * ``"operator_input"``  → LFY+1 = LFY × (1 + growth_rate); REQUIRES a
                                non-empty ``justification`` (refused otherwise);
                                source tags the rate + justification.

    Back-compat: an entry with an explicit ``approved_value`` dict and no
    ``choice`` sets the values directly (the pre-Q2 manual path)."""
    choice = str(entry.get("choice") or "").strip().lower()

    # Back-compat: explicit values, no choice → set directly.
    if not choice and entry.get("approved_value"):
        av = entry.get("approved_value") or {}
        if "revenue_lfy1_m" in av:
            row["revenue_lfy1_m"] = _safe_float(av["revenue_lfy1_m"])
        if "ebitda_lfy1_m" in av:
            row["ebitda_lfy1_m"] = _safe_float(av["ebitda_lfy1_m"])
        row["lfy1_source"] = f"operator-approved:{today.isoformat()}"
        return

    if choice in ("", "blank"):
        # Leave LFY+1 empty; absent value needs no source.
        return

    if choice == "self_historical":
        g_rev = _self_historical_growth(
            row.get("revenue_lfy_m"), row.get("revenue_lfym1_m"))
        g_eb = _self_historical_growth(
            row.get("ebitda_lfy_m"), row.get("ebitda_lfym1_m"))
        applied = False
        if g_rev is not None and row.get("revenue_lfy_m") is not None:
            row["revenue_lfy1_m"] = row["revenue_lfy_m"] * (1 + g_rev)
            applied = True
        if g_eb is not None and row.get("ebitda_lfy_m") is not None:
            row["ebitda_lfy1_m"] = row["ebitda_lfy_m"] * (1 + g_eb)
            applied = True
        if applied:
            row["lfy1_source"] = (
                f"operator-approved:self-historical-growth:{today.isoformat()}")
        return

    if choice == "operator_input":
        justification = str(entry.get("justification") or "").strip()
        if not justification:
            raise UnsourcedFigureError(
                f"LFY+1 operator_input for {ss}/{row.get('ticker', '?')} "
                "requires a non-empty 'justification' (Q2: un-justified "
                "operator growth rates are refused).")
        rate = _safe_float(entry.get("growth_rate"))
        if rate is None:
            raise UnsourcedFigureError(
                f"LFY+1 operator_input for {ss}/{row.get('ticker', '?')} "
                "requires a numeric 'growth_rate'.")
        if row.get("revenue_lfy_m") is not None:
            row["revenue_lfy1_m"] = row["revenue_lfy_m"] * (1 + rate)
        if row.get("ebitda_lfy_m") is not None:
            row["ebitda_lfy1_m"] = row["ebitda_lfy_m"] * (1 + rate)
        row["lfy1_source"] = (
            f"operator-approved:growth={rate};{justification[:80]}:"
            f"{today.isoformat()}")
        return

    # Unknown choice → refuse (never silently ignore an approval).
    raise UnsourcedFigureError(
        f"LFY+1 assumption for {ss}/{row.get('ticker', '?')}: unknown "
        f"choice {choice!r} (expected blank / self_historical / "
        "operator_input).")


def _apply_ebitdaal_assumption(
    row: dict[str, Any],
    entry: dict[str, Any],
    today: date,
    *,
    ss: Optional[str] = None,
) -> None:
    """Per-row EBITDAaL application (#21-comps-step-4, 2026-06-03). Two
    choices:

      * ``"blank"``           → leave EBITDAaL_{lfym1,lfy,lfy1} as None; no
                                source set (the Iron Law needs a source only
                                when a value is present). Operator fills in
                                Excel from the IFRS-16 note. Default path.
      * ``"operator_input"``  → operator supplies an ``approved_value`` dict
                                with any subset of {ebitdaal_lfym1_m,
                                ebitdaal_lfy_m, ebitdaal_lfy1_m} (each
                                optional — partial fills are fine) AND a
                                REQUIRED non-empty ``justification``.
                                Justification can be the operator's own
                                reasoning OR a deep-research source URL
                                (option-iii in SESSION-COMPS-ORCHESTRATION.md
                                §EBITDAaL — orchestrator session invokes
                                deep-research against the company's IR
                                portal, surfaces EBITDAaL or its components,
                                threads back here with source URL as
                                justification). Source is tagged
                                ``operator-approved:ebitdaal;<justification>:<today>``.

    No silent proxy applied — by design (operator vetoed baking a default
    lease-cost ratio; the assumption surfaces lease_liability + ebitda as
    context so the operator decides with it, not blind).
    """
    choice = str(entry.get("choice") or "").strip().lower()

    if choice in ("", "blank"):
        # Leave EBITDAaL empty; absent values need no source.
        return

    if choice == "operator_input":
        justification = str(entry.get("justification") or "").strip()
        if not justification:
            raise UnsourcedFigureError(
                f"EBITDAaL operator_input for {ss}/{row.get('ticker', '?')} "
                "requires a non-empty 'justification' (the operator's own "
                "reasoning OR a deep-research source URL — un-justified "
                "EBITDAaL is refused, same Iron Law as LFY+1)."
            )
        av = entry.get("approved_value") or {}
        # Codex-review finding SEV-3 (2026-06-03): convert first + refuse
        # non-numeric values. Pre-fix, the loop set ``applied=True`` based on
        # the presence-not-None of av[fld], BEFORE checking whether
        # ``_safe_float`` actually returned a number. So ``approved_value=
        # {"ebitdaal_lfy_m": ""}`` (or "not-a-number") would: leave row[fld]
        # = None (silently), still stamp ebitdaal_source, and pass the
        # "applied" guard — un-typed values landing in the deliverable
        # with a fabricated "operator-approved" source tag.
        applied = False
        for fld in ("ebitdaal_lfym1_m", "ebitdaal_lfy_m", "ebitdaal_lfy1_m"):
            if fld in av and av[fld] is not None:
                val = _safe_float(av[fld])
                if val is None:
                    raise UnsourcedFigureError(
                        f"EBITDAaL operator_input for {ss}/{row.get('ticker', '?')}: "
                        f"{fld!r} value {av[fld]!r} is not numeric (Iron Law: "
                        "un-typed values are refused; supply a float / int or "
                        "omit the field for the blank path)."
                    )
                row[fld] = val
                applied = True
        if not applied:
            raise UnsourcedFigureError(
                f"EBITDAaL operator_input for {ss}/{row.get('ticker', '?')} "
                "requires at least one of {ebitdaal_lfym1_m, ebitdaal_lfy_m, "
                "ebitdaal_lfy1_m} in approved_value."
            )
        row["ebitdaal_source"] = (
            f"operator-approved:ebitdaal;{justification[:80]}:"
            f"{today.isoformat()}"
        )
        return

    # Unknown choice → refuse (never silently ignore an approval).
    raise UnsourcedFigureError(
        f"EBITDAaL assumption for {ss}/{row.get('ticker', '?')}: unknown "
        f"choice {choice!r} (expected blank / operator_input)."
    )


def _block_headline_medians(blocks: list[dict[str, Any]]) -> tuple[Optional[float], Optional[float], int]:
    """Return (median_ev_ebitda, median_ev_revenue, peer_count_total).
    Computed from the CoCo rows; tolerates missing data (None medians)."""
    ev_ebitdas: list[float] = []
    ev_revs: list[float] = []
    peer_total = 0
    for block in blocks:
        for row in block.get("coco_rows") or []:
            peer_total += 1
            # If the template's downstream formulas land EV/EBITDA on the row
            # at stamp time, the row dict carries the computed value; otherwise
            # we re-derive from market_cap + net_debt + ebitda when possible.
            #
            # Iron-law correctness: a true EV requires BOTH market_cap AND
            # net_debt to be present. Using `(market_cap or 0) + (net_debt or
            # 0)` would silently land EV=0 → EV/EBITDA = 0.0 on rows where
            # either input is missing, polluting the median with spurious
            # zeros. Skip the row when either input is None.
            mc = row.get("market_cap_m")
            # Post v2-restructure: market_cap_m is template-formula-derived
            # (G = E*F). If the row dict has the upstream inputs but no
            # pre-computed mkt_cap, derive it here so the headline-median path
            # still works server-side without round-tripping through Excel.
            if mc is None:
                sp = row.get("shr_price")
                so = row.get("shares_out")
                if sp is not None and so is not None:
                    mc = sp * so
            nd = row.get("net_debt_m")
            ev: Optional[float] = (mc + nd) if (mc is not None and nd is not None) else None

            ee = row.get("ev_ebitda_lfy_x")
            er = row.get("ev_revenue_lfy_x")
            if ee is None and row.get("ebitda_lfy_m") and ev is not None:
                ee = ev / row["ebitda_lfy_m"] if row["ebitda_lfy_m"] else None
            if er is None and row.get("revenue_lfy_m") and ev is not None:
                er = ev / row["revenue_lfy_m"] if row["revenue_lfy_m"] else None
            if isinstance(ee, (int, float)):
                ev_ebitdas.append(float(ee))
            if isinstance(er, (int, float)):
                ev_revs.append(float(er))
    ee_med = round(statistics.median(ev_ebitdas), 2) if ev_ebitdas else None
    er_med = round(statistics.median(ev_revs), 2) if ev_revs else None
    return ee_med, er_med, peer_total


def run_stage_3(
    inputs: CompsBuildInput,
    *,
    vault_root: Path,
    run_id: str,
    blocks: list[dict[str, Any]],
    today: Optional[date] = None,
    template_src: Optional[Path] = None,
    output_root_override: Optional[Path] = None,
) -> StageResult:
    """Stamp the v2 template, archive prior, capture to vault, refresh mirror.

    ``blocks`` carries the Stage 2 acquired rows (with approved assumptions
    applied). ``template_src`` defaults to ``os-templates/Project_x_Comps_v2.xlsx``
    (operator-controlled); tests pass a tmp synthetic template. ``output_root_override``
    redirects the per-deal Valuation folder under a tmp root (tests use this)."""
    if not inputs.approved_subsectors:
        raise MissingApprovalToken("Stage 3 requires approved_subsectors")
    if not inputs.approved_peers_by_subsector or not inputs.approved_deals_by_subsector:
        raise MissingApprovalToken("Stage 3 requires approved_peers + approved_deals")
    # WS-B + #21-comps-step-3: re-verify the Stage-1 peers/deals tokens at
    # Stage 3 too (defence in depth — catches a caller who landed a valid
    # Stage-2 cache entry then mutates approved_peers / approved_deals at
    # Stage 3 entry). When the operator narrowed at Stage 2 via proposed_*,
    # those proposed_* must also be threaded through Stage 3 so the gate
    # verifies the HMAC over the same universe (codex-review finding P1,
    # 2026-06-03 — previously Stage 3 ignored proposed_* and rejected any
    # narrowing that Stage 2 had accepted).
    ok_peers, err_peers = _verify_subset_approval(
        deal_name=inputs.deal_name,
        stage_label=STAGE_LABEL_PEERS,
        approved=inputs.approved_peers_by_subsector,
        proposed=inputs.proposed_peers_by_subsector,
        token=inputs.peers_approval_token or "",
    )
    if not ok_peers:
        raise MissingApprovalToken(
            f"Stage 3: invalid or missing peers_approval_token — {err_peers}. "
            "Either echo the Stage-1 proposal verbatim as "
            "approved_peers_by_subsector, OR thread proposed_peers_by_subsector "
            "(the full Stage-1 proposal) + approved_peers_by_subsector (any "
            "subset of it) all the way through Stage 3."
        )
    ok_deals, err_deals = _verify_subset_approval(
        deal_name=inputs.deal_name,
        stage_label=STAGE_LABEL_DEALS,
        approved=inputs.approved_deals_by_subsector,
        proposed=inputs.proposed_deals_by_subsector,
        token=inputs.deals_approval_token or "",
    )
    if not ok_deals:
        raise MissingApprovalToken(
            f"Stage 3: invalid or missing deals_approval_token — {err_deals}. "
            "Either echo the Stage-1 proposal verbatim as "
            "approved_deals_by_subsector, OR thread proposed_deals_by_subsector "
            "(the full Stage-1 proposal) + approved_deals_by_subsector (any "
            "subset of it) all the way through Stage 3."
        )
    today = today or datetime.now(timezone.utc).date()

    # Apply any operator-approved assumptions (sets lfy1_source on the rows).
    # The assumptions_approval_token (issued at Stage 2 when unsourced LFY+1
    # candidates were surfaced) must verify against the supplied assumptions
    # payload. If approved_assumptions is empty / None there's nothing to
    # verify — but if it's present the token MUST also be present and valid.
    if inputs.approved_assumptions:
        if not _verify_token(
            inputs.deal_name, STAGE_LABEL_ASSUMPTIONS,
            _assumption_identity(inputs.approved_assumptions),
            inputs.assumptions_approval_token or "",
        ):
            raise MissingApprovalToken(
                "Stage 3: approved_assumptions present but assumptions_approval_"
                "token is missing or does not match the Stage-2-signed payload. "
                "Re-run Stage 2 + present the returned assumptions token."
            )
        _apply_approved_assumptions(blocks, inputs.approved_assumptions, today)

    # Iron Law piece 1 pre-stamp guard — every populated row sourced.
    all_sourced, rows_checked, offending = _coco_rows_have_sources(blocks)
    if not all_sourced:
        raise UnsourcedFigureError(
            "Iron Law: unsourced populated rows detected — refusing to stamp:\n  - "
            + "\n  - ".join(offending)
        )

    # Resolve the workbook target per workspace-write-policy.
    if output_root_override is not None:
        target_dir = output_root_override / inputs.deal_name / "3. Financials & analysis" / "2. Valuation" / "01. COMPS"
    else:
        target_dir = deal_valuation_folder(inputs.deal_name)
    target_dir.mkdir(parents=True, exist_ok=True)

    # Archive any prior same-day version BEFORE the new write.
    prior_archived = _archive_prior_same_day(target_dir, inputs.deal_name, today)

    # Stamp the template (block-flex, one CoCo block per subsector + CoTrans
    # grouped). Real impl orchestrates anthropic-skills:xlsx; the shim copies
    # the template and is mocked in tests.
    version = _next_version(target_dir, inputs.deal_name, today)
    output_path = target_dir / comps_output_filename(inputs.deal_name, today, version)
    template_path = template_src or Path("os-templates/Project_x_Comps_v2.xlsx")
    try:
        _shim_xlsx_stamp(template_path, output_path, blocks,
                         sensitivity=inputs.workspace_sensitivity)
    except Exception as e:  # noqa: BLE001
        raise TemplateStampFailed(f"xlsx stamp failed: {e}") from e

    # Headline medians + counts for the capture proposal, the sector mirror,
    # and the final bubble. Computed BEFORE the mirror emit so the snapshot
    # bullet (#43-sector-template-align §2) can render the medians + peer/deal
    # counts directly into the dated `## Comps runs` bullet.
    ee_med, er_med, peer_total = _block_headline_medians(blocks)
    deal_total = sum(len(b.get("cotrans_rows") or []) for b in blocks)
    subsector_slugs = [
        ss for b in blocks
        if (ss := str(b.get("subsector_slug") or "").strip())
    ]
    peer_tickers = [
        tk for b in blocks for r in (b.get("coco_rows") or [])
        if (tk := str(r.get("ticker") or "").strip())
    ]

    # Sector mirror. Emitted as an operator-gated deliverable-outcome
    # proposal (closes the #61 capability gap — no direct Sectors/** write).
    # The operator's Route action on the proposal performs the append via
    # _route_deliverable_outcome, under the `## Comps runs` section of the
    # canonical `Sectors/<sector>/Comps.md` (a valuation-snapshot, NOT a
    # precedent transaction — #43-sector-template-align).
    mirror_path: Optional[Path] = None
    try:
        mirror_path = _emit_sector_mirror_proposal(
            vault_root, inputs.parent_sector, inputs.deal_name, output_path,
            today, run_id=run_id,
            ev_ebitda_median=ee_med, ev_revenue_median=er_med,
            peer_count=peer_total, deal_count=deal_total,
            subsectors=subsector_slugs, peer_tickers=peer_tickers,
        )
    except Exception as e:  # noqa: BLE001
        log.warning("comps: sector-mirror proposal emission failed: %s", e)

    # The provider tag for the audit / capture line — comma-joined sources used.
    provider_label = _build_provider_label(blocks)

    return StageResult(
        ok=True,
        stage="complete",
        stage_just_completed=3,
        deal_name=inputs.deal_name,
        target=inputs.target,
        run_id=run_id,
        warnings=[],
        approved_subsectors=list(inputs.approved_subsectors),
        blocks=blocks,
        headline_ev_ebitda_median=ee_med,
        headline_ev_revenue_median=er_med,
        peer_count=peer_total,
        deal_count=deal_total,
        as_of=today.isoformat(),
        provider=provider_label,
        template_path=str(output_path),
        prior_archived_path=str(prior_archived) if prior_archived else None,
        mirror_refresh_path=str(mirror_path) if mirror_path else None,
        iron_law_assertion={"all_rows_sourced": True, "rows_checked": rows_checked},
    )


def _build_provider_label(blocks: list[dict[str, Any]]) -> str:
    """Comma-joined provider names parsed from row Source cells. Surfaces the
    full provenance bench so the capture proposal's headline carries the truth."""
    names: list[str] = []
    seen: set[str] = set()
    for block in blocks:
        for row in (block.get("coco_rows") or []):
            src = row.get("source") or ""
            if ":" in src and not src.startswith(_ACCEPTED_SOURCE_PREFIXES):
                name = src.split(":", 1)[0]
                if name and name not in seen:
                    seen.add(name); names.append(name)
            elif src.startswith("http"):
                if "IR-scrape" not in seen:
                    seen.add("IR-scrape"); names.append("IR-scrape")
            elif src.startswith("tracker:"):
                if "tracker" not in seen:
                    seen.add("tracker"); names.append("tracker")
            elif src.startswith("operator-approved:"):
                if "operator-approved" not in seen:
                    seen.add("operator-approved"); names.append("operator-approved")
        for row in (block.get("cotrans_rows") or []):
            src = row.get("source") or ""
            if src.startswith("tracker:") and "tracker" not in seen:
                seen.add("tracker"); names.append("tracker")
            elif src.startswith("http") and "IR-scrape" not in seen:
                seen.add("IR-scrape"); names.append("IR-scrape")
    return " + ".join(names) if names else "n/a"


# ─────────────────────────────────────────────────────────────────────────────
# Top-level dispatch
# ─────────────────────────────────────────────────────────────────────────────


def new_run_id() -> str:
    return uuid.uuid4().hex[:8]


def run(
    inputs: CompsBuildInput,
    *,
    vault_root: Path,
    run_id: Optional[str] = None,
    today: Optional[date] = None,
    template_src: Optional[Path] = None,
    tracker_workbook: Optional[Path] = None,
    output_root_override: Optional[Path] = None,
    stage_2_blocks: Optional[list[dict[str, Any]]] = None,
) -> StageResult:
    """Top-level Stage dispatch. The route picks ``inputs.stage`` from the
    operator's request; each Stage call returns either a Stage-N approval-
    pending payload or the Stage 3 complete payload."""
    rid = run_id or new_run_id()
    if inputs.stage == 0:
        return run_stage_0(inputs, vault_root=vault_root, run_id=rid)
    if inputs.stage == 1:
        return run_stage_1(
            inputs, vault_root=vault_root, run_id=rid,
            tracker_workbook=tracker_workbook, today=today,
        )
    if inputs.stage == 2:
        return run_stage_2(
            inputs, vault_root=vault_root, run_id=rid, today=today,
            tracker_workbook=tracker_workbook,
        )
    if inputs.stage == 3:
        # WS-C: Stage 3 reads Stage-2 blocks from the in-process cache keyed
        # by the stage_2_blocks_approval_token. The composition with WS-B is
        # the point: the HMAC token IS the cache key (signed over
        # (deal_name, STAGE_LABEL_STAGE_2_BLOCKS, h(blocks))), so a tampered
        # cache entry would mismatch on the verify-on-read defence in depth.
        #
        # Priority order:
        #   1. Direct ``stage_2_blocks`` arg (smoke tests / direct Python
        #      callers — bypasses cache; intended for fixture-driven tests).
        #   2. Cache lookup by ``stage_2_blocks_approval_token`` (the
        #      production path on Stage 3 re-fire from the bridge).
        #   3. Re-run Stage 2 (LEGACY fallback for callers who pass
        #      neither — preserved for backward-compat with smoke tests
        #      that pre-date WS-C; the brief's contract is that Stage 3
        #      from the bridge ALWAYS supplies the token).
        blocks = stage_2_blocks
        if blocks is None and inputs.stage_2_blocks_approval_token:
            cached = _get_cached_stage_2_blocks(
                inputs.stage_2_blocks_approval_token,
            )
            if cached is None:
                raise MissingApprovalToken(
                    "Stage 3: stage-2 cache miss for the supplied "
                    "stage_2_blocks_approval_token — the cache entry has been "
                    "evicted (TTL expiry or process restart), OR the token "
                    "was never issued by this process. Re-run Stage 2 to "
                    "re-seed the cache + present the new token."
                )
            # Defence in depth: verify the cached blocks still hash to the
            # presented token (a tampered cache entry would mismatch).
            if not _verify_token(
                inputs.deal_name, STAGE_LABEL_STAGE_2_BLOCKS, cached,
                inputs.stage_2_blocks_approval_token,
            ):
                raise MissingApprovalToken(
                    "Stage 3: cached stage-2 blocks failed HMAC re-verification "
                    "against the presented stage_2_blocks_approval_token "
                    "(cache integrity violation). Refusing to stamp."
                )
            blocks = cached
        if blocks is None:
            # Legacy fallback — re-run Stage 2 in-process. The production
            # bridge always supplies the token; this path supports smoke
            # tests that pre-date the cache.
            r2 = run_stage_2(inputs, vault_root=vault_root, run_id=rid, today=today)
            blocks = r2.blocks or []
        return run_stage_3(
            inputs, vault_root=vault_root, run_id=rid, blocks=blocks,
            today=today, template_src=template_src,
            output_root_override=output_root_override,
        )
    raise CompsSkillError(f"unknown stage {inputs.stage!r}")


__all__ = [
    # IO
    "CompsBuildInput", "ApprovalPayload", "StageResult",
    # Stages
    "run_stage_0", "run_stage_1", "run_stage_2", "run_stage_3", "run",
    # Helpers
    "new_run_id",
    "deal_valuation_folder", "comps_output_filename",
    # Exceptions
    "CompsSkillError", "MissingApprovalToken", "UnsourcedFigureError",
    "TargetBriefMissing", "TemplateStampFailed",
    # WS-B HMAC token helpers (exposed for the bridge UI + tests).
    "STAGE_LABEL_SUBSECTORS", "STAGE_LABEL_PEERS", "STAGE_LABEL_DEALS",
    "STAGE_LABEL_STAGE_2_BLOCKS", "STAGE_LABEL_ASSUMPTIONS",
]
