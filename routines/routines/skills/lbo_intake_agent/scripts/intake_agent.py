"""lbo-intake-agent helpers — deterministic doc extraction + judgment parsing.

Pure functions (no governance, no HTTP): the route owns the ``@anton_skill``
jacket and the governed ``llm()`` calls; this module owns everything testable
without either — reading documents into bounded text, building the two prompt
shapes, and parsing/validating the model's JSON judgment against the boxes
manifest.

NO-LLM-MATHS (CLAUDE.md): the judgment contract is TRANSCRIBE-ONLY — the model
reports values it can point at in a document (value + location + verbatim
quote); anything it cannot source becomes an open question for the operator.
``parse_judgment`` enforces the shape and demotes anything unsourced or
un-coercible; the prompts state the rule; the operator confirms every box in
the modal regardless.
"""

from __future__ import annotations

import json
import logging
import math
import re
from dataclasses import dataclass, field
from datetime import date
from pathlib import Path
from typing import Any, Optional

logger = logging.getLogger(__name__)

# Bounded extraction: a CIM can be 200 pages; the local lane's context window is
# finite (and an override-window Opus call bills by token). Caps are per-doc —
# generous enough for the financial sections that matter; the per-doc digest
# pass condenses further before synthesis.
MAX_PDF_PAGES = 80
MAX_SHEET_ROWS = 200
MAX_SHEETS = 12
MAX_DOC_CHARS = 24_000
# Refuse before the parser libraries even open the file (codex slice-2 SEV-2:
# a multi-GB path must not reach pypdf/openpyxl).
MAX_DOC_BYTES = 50 * 1024 * 1024

_SUPPORTED_SUFFIXES = (".pdf", ".xlsx", ".xlsm", ".txt", ".md", ".csv")


# ─────────────────────────────────────────────────────────────────────────────
# Document reading (deterministic, in-process — pypdf / openpyxl)
# ─────────────────────────────────────────────────────────────────────────────


def read_document(path_str: str) -> dict[str, Any]:
    """Read one document into bounded plain text.

    Returns ``{"name", "kind", "text", "truncated", "error"}`` — ``error`` is
    ``None`` on success; a failed read NEVER raises (the route reports
    unreadable paths in the suspension note instead of burning the run).
    """
    p = Path(path_str)
    out: dict[str, Any] = {
        "name": p.name or path_str, "kind": p.suffix.lower().lstrip("."),
        "text": "", "truncated": False, "error": None,
    }
    if not p.is_file():
        out["error"] = "not found / not a file"
        return out
    suffix = p.suffix.lower()
    if suffix not in _SUPPORTED_SUFFIXES:
        out["error"] = (
            f"unsupported type {suffix!r} "
            f"(supported: {', '.join(_SUPPORTED_SUFFIXES)})"
        )
        return out
    try:
        size = p.stat().st_size
    except OSError as e:
        out["error"] = f"unreadable: {e}"
        return out
    if size > MAX_DOC_BYTES:
        out["error"] = f"file too large ({size} bytes > {MAX_DOC_BYTES} cap)"
        return out
    try:
        if suffix == ".pdf":
            out["text"], out["truncated"] = _read_pdf(p)
        elif suffix in (".xlsx", ".xlsm"):
            out["text"], out["truncated"] = _read_xlsx(p)
        else:
            raw = p.read_text(encoding="utf-8", errors="replace")
            out["text"] = raw[:MAX_DOC_CHARS]
            out["truncated"] = len(raw) > MAX_DOC_CHARS
    except Exception as e:  # noqa: BLE001 — surface as a per-doc warning, not a 500
        out["error"] = f"{type(e).__name__}: {e}"
        return out
    if not out["text"].strip():
        # Typical cause: a scanned/image-only PDF. The multimodal local lane can
        # read those, but wiring page-render → gemma is a follow-up; for now the
        # operator is told exactly which doc came back empty.
        out["error"] = "no extractable text (scanned/image-only PDF?)"
    return out


def _read_pdf(p: Path) -> tuple[str, bool]:
    from pypdf import PdfReader

    reader = PdfReader(str(p))
    truncated = len(reader.pages) > MAX_PDF_PAGES
    chunks: list[str] = []
    total = 0
    for i, page in enumerate(reader.pages[:MAX_PDF_PAGES], start=1):
        t = (page.extract_text() or "").strip()
        if not t:
            continue
        entry = f"[page {i}]\n{t}"
        if total + len(entry) > MAX_DOC_CHARS:
            truncated = True
            break
        chunks.append(entry)
        total += len(entry)
    return "\n\n".join(chunks), truncated


def _read_xlsx(p: Path) -> tuple[str, bool]:
    from openpyxl import load_workbook

    wb = load_workbook(str(p), read_only=True, data_only=True)
    try:
        truncated = len(wb.worksheets) > MAX_SHEETS
        chunks: list[str] = []
        total = 0
        for ws in wb.worksheets[:MAX_SHEETS]:
            rows_out: list[str] = []
            for r_i, row in enumerate(ws.iter_rows(values_only=True), start=1):
                if r_i > MAX_SHEET_ROWS:
                    truncated = True
                    break
                if all(v is None for v in row):
                    continue
                rows_out.append(
                    "\t".join("" if v is None else str(v) for v in row)
                )
            if not rows_out:
                continue
            entry = f"[sheet {ws.title}]\n" + "\n".join(rows_out)
            if total + len(entry) > MAX_DOC_CHARS:
                truncated = True
                break
            chunks.append(entry)
            total += len(entry)
        return "\n\n".join(chunks), truncated
    finally:
        wb.close()


# ─────────────────────────────────────────────────────────────────────────────
# Prompt builders
# ─────────────────────────────────────────────────────────────────────────────


def digest_prompt(doc: dict[str, Any]) -> str:
    """Per-document condensation pass: every LBO-relevant fact + its location."""
    trunc = " — TRUNCATED, later pages/sheets not shown" if doc["truncated"] else ""
    return (
        "Extract deal facts from ONE document for an LBO intake.\n"
        f"Document: {doc['name']} ({doc['kind']}{trunc})\n"
        "---\n"
        f"{doc['text']}\n"
        "---\n"
        "Return a CONCISE digest (plain text, max 80 lines): every financial "
        "fact relevant to a leveraged buyout — EBITDA by year AND its basis "
        "(reported/adjusted/management), revenue, existing net debt, capex, "
        "working-capital movement, D&A, fiscal year end, any proposed leverage "
        "or entry multiple, tax rate, deal dates — each on its own line with "
        "its location in this document (e.g. 'p.12' or 'sheet FY26 row 8').\n"
        "TRANSCRIBE-ONLY: copy figures verbatim with their units. Never "
        "compute, derive, annualise or infer a number that is not printed."
    )


def synthesis_prompt(
    deal_name: str,
    deal_context: str,
    digests: list[dict[str, str]],
    manifest: list[dict[str, Any]],
) -> str:
    """The judgment pass: digests → boxes JSON + optional client_fs + questions."""
    fields_desc = "\n".join(
        f"- {f['key']}: {f['label']}"
        f" [type={f['type']}{', unit=' + f['unit'] if f.get('unit') else ''}"
        f"{', options=' + repr(f['options']) if f.get('options') else ''}"
        f"{', default=' + repr(f['default']) if 'default' in f and f['default'] is not None else ''}]"
        for f in manifest
    )
    digest_text = "\n\n".join(
        f"=== {d['name']} ===\n{d['digest']}" for d in digests
    )
    return (
        f"LBO intake judgment for deal {deal_name!r}.\n"
        f"Operator context: {deal_context or '(none given)'}\n\n"
        "Per-document fact digests (location references are into the original "
        "documents):\n"
        f"{digest_text}\n\n"
        "The deal-assumption boxes (fill ONLY keys from this list):\n"
        f"{fields_desc}\n\n"
        "Units: unit=m means £ millions as a plain number (13.5 for £13.5m); "
        "unit=x a multiple; unit=dec a decimal fraction (0.25 for 25%); dates "
        "ISO YYYY-MM-DD; select values must be one of the listed options.\n\n"
        "Return STRICT JSON, nothing else:\n"
        "{\n"
        '  "boxes": {"<key>": {"value": <number|string>, "source": "<doc name> '
        '<location>", "quote": "<verbatim source text, max 30 words>"}},\n'
        '  "client_fs": null OR {"dates": [<exactly 10 ISO dates>], "rows": '
        '{"6": [<10 revenue values>], "13": [<10 EBITDA values>], "25": '
        '[<10 D&A values>], "32": [<10 working-capital-movement values>], '
        '"38": [<10 capex values>]}},\n'
        '  "open_questions": [{"key": "<box key or topic>", "question": '
        '"<the precise question for the operator>"}],\n'
        '  "notes": "<max 80 words on basis choices / conflicts seen>"\n'
        "}\n\n"
        "Rules:\n"
        "1. TRANSCRIBE-ONLY: a box value must appear in a digest verbatim. "
        "Never compute, blend, annualise or infer. Both source AND a non-empty "
        "verbatim quote are REQUIRED per box — no source or no quote → no box "
        "→ put it in open_questions instead.\n"
        "2. Conflicting candidates (e.g. management vs organic EBITDA builds) "
        "→ do NOT pick one: raise an open_question naming both, with sources.\n"
        "3. client_fs ONLY if a clear 10-period operating model can be "
        "transcribed; values in FULL currency units (17900000.0 for £17.9m — "
        "NOT millions). Otherwise client_fs = null.\n"
        "4. Defaulted convention boxes: only fill when a document explicitly "
        "states a different value; silence means the default stands.\n"
    )


def repair_prompt(bad_output: str, error: str) -> str:
    return (
        "Your previous output could not be parsed as the required JSON "
        f"contract ({error}). Reply with ONLY the corrected strict JSON object "
        "— no prose, no code fences.\n"
        "Previous output:\n"
        f"{bad_output[:6000]}"
    )


# ─────────────────────────────────────────────────────────────────────────────
# Judgment parsing + validation
# ─────────────────────────────────────────────────────────────────────────────


class JudgmentParseError(ValueError):
    """The model's output does not satisfy the JSON judgment contract."""


@dataclass
class Judgment:
    """Validated judgment: only manifest keys, only sourced + coercible values;
    everything else demoted to ``open_questions``."""

    boxes: dict[str, dict[str, Any]] = field(default_factory=dict)
    client_fs: Optional[dict[str, Any]] = None
    open_questions: list[dict[str, str]] = field(default_factory=list)
    notes: str = ""


def parse_judgment(text: str, manifest: list[dict[str, Any]]) -> Judgment:
    """Parse + validate the synthesis output. Raises JudgmentParseError when no
    JSON object can be recovered; ALL soft problems (unknown keys, unsourced or
    un-coercible values) demote to open questions — the agent must degrade to
    "ask the operator", never to a guess."""
    try:
        data = json.loads(_extract_json(text))
    except (ValueError, TypeError) as e:
        raise JudgmentParseError(f"not valid JSON: {e}") from e
    if not isinstance(data, dict):
        raise JudgmentParseError(f"top level is {type(data).__name__}, expected object")

    fields = {f["key"]: f for f in manifest}
    j = Judgment(notes=str(data.get("notes") or "")[:600])

    raw_boxes = data.get("boxes")
    for key, entry in (raw_boxes.items() if isinstance(raw_boxes, dict) else []):
        f = fields.get(key)
        if f is None:
            logger.info("judgment box %r is not a manifest key — dropped", key)
            continue
        source = str(entry.get("source") or "").strip() if isinstance(entry, dict) else ""
        quote = str(entry.get("quote") or "").strip() if isinstance(entry, dict) else ""
        # codex slice-2 SEV-3: a plausible source STRING alone is too easy to
        # hallucinate/inject — the verbatim quote is part of the transcribe-only
        # contract, so its absence demotes just like a missing source.
        if not isinstance(entry, dict) or "value" not in entry or not source or not quote:
            j.open_questions.append({
                "key": key,
                "question": f"The agent proposed {f['label']!r} without a usable "
                            "source + verbatim quote — provide the value and "
                            "where it comes from.",
            })
            continue
        coerced = _coerce(f, entry["value"])
        if coerced is None:
            j.open_questions.append({
                "key": key,
                "question": f"Could not validate {f['label']!r} = "
                            f"{entry['value']!r} as {f['type']} — provide it.",
            })
            continue
        j.boxes[key] = {
            "value": coerced,
            "source": source[:160],
            "quote": quote[:240],
        }

    raw_qs = data.get("open_questions")
    for q in (raw_qs if isinstance(raw_qs, list) else []):
        if isinstance(q, dict) and str(q.get("question") or "").strip():
            j.open_questions.append({
                "key": str(q.get("key") or "general")[:64],
                "question": str(q["question"])[:400],
            })

    raw_cfs = data.get("client_fs")
    if isinstance(raw_cfs, dict):
        j.client_fs = raw_cfs

    return j


def _coerce(f: dict[str, Any], value: Any) -> Any:
    """Coerce a raw judgment value to the manifest field type; None = invalid.
    (None is never a legitimate coerced value — every type below excludes it.)"""
    t = f["type"]
    try:
        if t == "number":
            if isinstance(value, bool) or value is None:
                return None
            v = float(value)
            return v if math.isfinite(v) else None
        if t == "int":
            if isinstance(value, bool) or value is None:
                return None
            return int(value)
        if t == "date":
            date.fromisoformat(str(value))
            return str(value)
        if t == "select":
            return value if value in (f.get("options") or []) else None
        if t == "text":
            s = str(value).strip()
            return s or None
    except (ValueError, TypeError):
        return None
    return None


def _extract_json(text: str) -> str:
    """The JSON object out of a possibly-fenced / prose-wrapped reply."""
    m = re.search(r"```(?:json)?\s*(\{.*?\})\s*```", text, re.S)
    if m:
        return m.group(1)
    start, end = text.find("{"), text.rfind("}")
    if start != -1 and end > start:
        return text[start:end + 1]
    raise ValueError("no JSON object found in output")
