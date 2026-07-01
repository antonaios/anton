"""Excel writer for the precedent transactions tracker workbook.

Canonical path (post 2026-06-01 retarget per WS2 of ``SESSION-COMPS-
IMPLEMENTATION.md``)::

    <workspace-root>/4. Research & data/Precedent transactions tracker/
        Precedent_transactions_tracker.xlsx

Sheet: ``Precedent transactions``. Lean 19-column header (see
``schema.COLUMNS``). Created on first append if missing; subsequent appends
are pure data rows. The live filename is STABLE (no per-append rename) —
date-stamped snapshots live under ``./Archive/`` and are produced by a
separate cadence job, not by every append.

Uses openpyxl (no Excel runtime required). The workbook's threaded-comment
"source" annotations in the operator's hand-maintained section are stripped
by openpyxl on save — that's why the lean schema has a dedicated ``Source``
column: provenance lives in a cell, not a comment.
"""

from __future__ import annotations

import logging
from datetime import date, datetime
from pathlib import Path

from routines.dealtracker.schema import COLUMNS, DealRecord

logger = logging.getLogger(__name__)


# Canonical workbook + sheet — single source of truth for the cron's auto-feed
# and the human-paste route. Operator override via ``append_deal(workbook_path=)``
# (tests use ``tmp_path``; CLI defaults to this; sector-news pipeline points at
# this; the route resolves against vault root + this rel).
CANONICAL_WORKBOOK_PATH = Path(
    r"<workspace-root>\4. Research & data"
    r"\Precedent transactions tracker\Precedent_transactions_tracker.xlsx"
)
CANONICAL_SHEET_NAME = "Precedent transactions"


def append_deal(
    workbook_path: Path,
    deal: DealRecord,
    *,
    sheet_name: str = CANONICAL_SHEET_NAME,
) -> dict[str, str]:
    """Append a deal to the workbook. Creates the workbook with the lean
    19-column header row if missing. Append-in-place — no per-append rename
    of the live file. Returns ``{"status": ..., "row" | "existing_row": ...}``.

    Statuses: ``"appended" | "skipped_duplicate" | "error"``.

    Idempotency: ``(announced_date, target_company)`` (case- and whitespace-
    insensitive on the target). Same key → skipped with the existing row #.
    """
    import openpyxl
    from openpyxl import Workbook

    workbook_path.parent.mkdir(parents=True, exist_ok=True)

    # Open or create. On create we write the lean 19-col header row and re-open
    # so the rest of the function operates on a single load path.
    if workbook_path.exists():
        wb = openpyxl.load_workbook(str(workbook_path))
    else:
        wb = Workbook()
        # openpyxl creates a default "Sheet" — replace with the named sheet.
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]
        ws = wb.create_sheet(sheet_name)
        ws.append(COLUMNS)
        wb.save(str(workbook_path))
        logger.info("created new precedent tracker workbook: %s", workbook_path)
        wb = openpyxl.load_workbook(str(workbook_path))

    # Find sheet (first sheet if requested name absent — operators may rename
    # the live sheet without breaking the cron).
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
    else:
        ws = wb.worksheets[0]

    # Idempotency check. openpyxl reads dates back as datetime; convert to
    # date so the dedupe_key (which uses .isoformat()) matches between the
    # incoming record and existing rows. Column layout: row[0] is "Announced
    # Date", row[1] is "Target" in the lean schema (previously row[2]).
    new_key = deal.dedupe_key()
    if new_key.split("|")[1]:  # only check when target_company is present
        for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if not row or len(row) < 2:
                continue
            cell0 = row[0]
            if isinstance(cell0, datetime):
                existing_date: date | None = cell0.date()
            elif isinstance(cell0, date):
                existing_date = cell0
            else:
                existing_date = None
            existing = DealRecord(
                announced_date=existing_date,
                target_company=str(row[1] or ""),
            )
            if existing.dedupe_key() == new_key:
                logger.info(
                    "duplicate deal skipped: %s (existing row %d)", new_key, row_num,
                )
                return {"status": "skipped_duplicate", "existing_row": str(row_num)}

    # Append the lean 19-col row in place — no rename of the live file.
    ws.append(deal.to_row())
    new_row_num = ws.max_row
    wb.save(str(workbook_path))
    logger.info("appended deal at row %d: %s", new_row_num, deal.target_company)
    return {"status": "appended", "row": str(new_row_num)}
