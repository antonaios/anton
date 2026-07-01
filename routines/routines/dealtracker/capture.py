"""#43 — deal-tracker vault enrichment (the "compounding intelligence" half).

:func:`emit_deal_capture` is the multi-target sibling of #76's
:func:`routines.skills._runtime.capture.emit_deliverable_proposal`. For one
appended deal it emits up to **three** operator-gated ``deliverable-outcome``
proposals — one per target — reusing the #76 substrate:

================  ==============================  ====================  ==============================
Target            File                            ``section``           Append shape
================  ==============================  ====================  ==============================
Acquirer          ``Companies/<Acquirer>.md``     ``Transaction history``  flat role-aware bullet (#76 path)
Target / Vendor   ``Companies/<Target>.md``       ``Transaction history``  flat role-aware bullet (#76 path)
Sector            ``Sectors/<slug>/Comps.md``     ``<YYYY>``               structured ``### comp-<id>`` block (``body_md``)
================  ==============================  ====================  ==============================

The company bullets ride the existing #76 flat-bullet route verbatim (the route
handler renders the dated bullet from ``headline``). The sector block is the one
richer render: the proposal carries an optional ``body_md`` (the fully-rendered
``### comp-<id>`` block) that ``_route_deliverable_outcome`` appends heading-aware
under the year, idempotent on the ``### comp-<id>`` anchor (SESSION-43 decision 4,
Option B).

**Trigger** (SESSION-43 decision 5): the two *non-comps* ingestion sites — the
sector-news auto-feed loop and the manual ``deal-tracker add`` CLI — call this
best-effort after a successful (non-duplicate) ``append_deal``. ``append_deal``
itself stays pure I/O. Comps is the deliverable-level capture owner and never
fires #43 (decision 6) — so a comps-sourced CoTrans write-back is NOT
per-party enriched here, avoiding a double-emit.

Best-effort everywhere: a capture miss logs and returns whatever was written;
it never raises into the ingestion path.
"""

from __future__ import annotations

import hashlib
import logging
from datetime import date, datetime, timezone
from pathlib import Path
from typing import Any, Optional

import frontmatter

from routines.dealtracker.schema import COLUMNS, DealRecord, _target_slug, build_deal_id
from routines.skills._runtime.capture import (
    PROPOSAL_DIR_REL,
    SKIP_STATUSES,
    _slug,
)

log = logging.getLogger(__name__)

# Minimum tracked subsector rows (excluding the deal itself) before the outlier
# triage hint is meaningful — below this a "range" is noise. The hard
# materiality threshold is deferred to #46c; this is the interim heuristic
# (SESSION-43 §5). Swapped at this call site when #46c lands.
OUTLIER_MIN_ROWS = 3

# Auto-extracted deals are medium-confidence by construction (LLM extraction /
# operator paste, not a hand-verified disclosure). The operator can promote on
# Route or via a later edit.
_DEFAULT_CONFIDENCE = "medium"

_CURRENCY_SYMBOLS = {"GBP": "£", "USD": "$", "EUR": "€", "JPY": "¥", "CHF": "CHF "}


# ─────────────────────────────────────────────────────────────────────────────
# Small render helpers
# ─────────────────────────────────────────────────────────────────────────────


def _money(value: Optional[float], currency: str) -> str:
    """Render a millions-denominated figure: ``4300 → "£4.3bn"``,
    ``137.91 → "£138m"``. ``None`` → ``"n/a"``."""
    if value is None:
        return "n/a"
    sym = _CURRENCY_SYMBOLS.get((currency or "").upper().strip(), "")
    if not sym and currency:
        sym = f"{currency.strip().upper()} "
    if abs(value) >= 1000:
        return f"{sym}{value / 1000:.1f}bn"
    return f"{sym}{value:,.0f}m"


def _mult(value: Optional[float]) -> str:
    """``9.0 → "9.0x"``; ``None`` → ``""`` (omit rather than render n/a inline)."""
    return f"{value:.1f}x" if value is not None else ""


def _first_sentence(text: str) -> str:
    """First sentence of a free-text blurb, for the sector block's Structure
    line. Empty string when there's nothing to say."""
    t = (text or "").strip()
    if not t:
        return ""
    for stop in (". ", ".\n"):
        i = t.find(stop)
        if i != -1:
            return t[: i + 1].strip()
    return t.split("\n", 1)[0].strip()


def _company_link(name: str) -> str:
    """``[[Companies/<name>]]`` wikilink, or the bare (possibly empty) name."""
    name = (name or "").strip()
    return f"[[Companies/{name}]]" if name else ""


def _sector_slug_for(deal: DealRecord) -> str:
    """The broad-sector slug for ``Sectors/<slug>/Comps.md``.

    The auto-feed path stamps ``subsector_slug`` with the cron's sector context
    (e.g. ``"hospitality"`` — see ``sectornews/pipeline.py``); the manual-add
    path leaves it blank, so fall back to the first Mergermarket category.
    Returns ``""`` when no sector is known — the caller then skips the sector
    proposal (a deal we can't address to a sector note can't enrich one).
    """
    if deal.subsector_slug.strip():
        return deal.subsector_slug.strip().lower().replace(" ", "-")
    first_cat = (deal.target_sector or "").split(",")[0].strip()
    return _slug(first_cat) if first_cat else ""


def _disambig(deal: DealRecord) -> str:
    """Short stable suffix from the deal's CANONICAL identity — the tracker's own
    ``dedupe_key()`` (announced_date + normalised target).

    Collision-resistant where slugs aren't: two distinct deals whose target
    names slugify to the same string (``"A&B"`` vs ``"A B"``), or the same target
    in the same month on different days, get DISTINCT suffixes — so their
    proposal filenames and ``### comp-<id>`` anchors don't collide and silently
    drop enrichment. A re-emit of the SAME deal yields the SAME suffix, so
    filename-slot + route-side idempotency still hold."""
    return hashlib.sha1(deal.dedupe_key().encode("utf-8")).hexdigest()[:8]


def _safe_note_name(name: str) -> bool:
    """A vault note stem must be a single safe path segment. Rejects names that
    would escape the note's directory (defense in depth — the route's
    ``_safe_vault_target`` is the hard gate, but we avoid emitting un-routable
    proposals from an adversarial LLM-extracted company name)."""
    n = (name or "").strip()
    return bool(n) and ".." not in n and not any(c in n for c in ("/", "\\", ":"))


def _comp_id(deal: DealRecord, fact_date: date) -> str:
    """``comp-<YYYY>-<MM>-<target-slug>-<h>`` — the sector-block anchor + the
    in-bullet ``#<comp-id>`` reference. The ``<h>`` suffix (:func:`_disambig`)
    keeps the anchor unique across slug-colliding / same-month deals while
    staying stable across re-emits of the same deal."""
    slug = _target_slug(deal.target_company) or "deal"
    return f"comp-{fact_date.year:04d}-{fact_date.month:02d}-{slug}-{_disambig(deal)}"


# ─────────────────────────────────────────────────────────────────────────────
# Tracker read-side context (step 4 decoration + step 8 outlier hint)
# ─────────────────────────────────────────────────────────────────────────────


def _tracker_context(
    workbook_path: Optional[Path], deal: DealRecord, *, sheet_name: str
) -> tuple[list[float], int]:
    """One best-effort read of the tracker for the two decorations:

    - the EV/EBITDA values of OTHER rows in this deal's subsector (step 8
      outlier hint), excluding the deal's own just-appended row, and
    - how many OTHER tracked deals already name this acquirer (step 4
      prior-encounter cross-check).

    Returns ``([], 0)`` on any failure (missing/locked workbook, no subsector
    context) — the decorations simply don't appear. Never raises.
    """
    if workbook_path is None or not deal:
        return [], 0
    subsector = _sector_slug_for(deal)
    acquirer = (deal.bidder_company or "").strip().lower()
    if not subsector and not acquirer:
        return [], 0
    try:
        import openpyxl

        wb = openpyxl.load_workbook(str(workbook_path), read_only=True, data_only=True)
        ws = wb[sheet_name] if sheet_name in wb.sheetnames else wb.worksheets[0]
        i_date = COLUMNS.index("Announced Date")
        i_target = COLUMNS.index("Target")
        i_sub = COLUMNS.index("Subsector (slug)")
        i_acq = COLUMNS.index("Acquirer")
        i_ebitda_x = COLUMNS.index("EV/EBITDA")

        own_target = (deal.target_company or "").strip().lower()
        own_date = deal.announced_date

        vals: list[float] = []
        prior_acq = 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or len(row) <= i_ebitda_x:
                continue
            cell_date = row[i_date]
            row_date: date | None = (
                cell_date.date() if isinstance(cell_date, datetime)
                else cell_date if isinstance(cell_date, date) else None
            )
            row_target = str(row[i_target] or "").strip().lower()
            is_own_row = row_target == own_target and row_date == own_date

            row_sub = str(row[i_sub] or "").strip().lower().replace(" ", "-")
            ebitda_x = row[i_ebitda_x]
            if (
                subsector
                and row_sub == subsector
                and isinstance(ebitda_x, (int, float))
                and not is_own_row
            ):
                vals.append(float(ebitda_x))

            if acquirer and str(row[i_acq] or "").strip().lower() == acquirer and not is_own_row:
                prior_acq += 1
        wb.close()
        return vals, prior_acq
    except Exception as e:  # noqa: BLE001 — openpyxl/IO can raise broadly; best-effort
        log.warning("deal-capture: tracker context read failed (%s) — decorations skipped", e)
        return [], 0


def _outlier_hint(deal: DealRecord, subsector_vals: list[float]) -> str:
    """``"⚠ Outlier: 9.0x EV/EBITDA vs subsector 3.5–5.6x (n=6)"`` when the
    deal's EV/EBITDA falls outside the tracked subsector range, else ``""``.

    Requires the deal's own multiple and ≥``OUTLIER_MIN_ROWS`` comparison rows
    (the hint stays silent on a thin tracker — no false signal)."""
    me = deal.reported_ebitda_multiple_y1
    if me is None or len(subsector_vals) < OUTLIER_MIN_ROWS:
        return ""
    lo, hi = min(subsector_vals), max(subsector_vals)
    if me < lo or me > hi:
        return (
            f"⚠ Outlier: {me:.1f}x EV/EBITDA vs subsector "
            f"{lo:.1f}–{hi:.1f}x (n={len(subsector_vals)})"
        )
    return ""


# ─────────────────────────────────────────────────────────────────────────────
# Renderers
# ─────────────────────────────────────────────────────────────────────────────


def _multiples_phrase(deal: DealRecord) -> str:
    """``"£4.3bn EV, 9.0x EV/EBITDA"`` — the shared money+multiple phrase used in
    both the company bullets and the sector block's Implied line."""
    parts = [f"{_money(deal.enterprise_value_m, deal.currency)} EV"]
    ev_ebitda = _mult(deal.reported_ebitda_multiple_y1)
    if ev_ebitda:
        parts.append(f"{ev_ebitda} EV/EBITDA")
    ev_rev = _mult(deal.reported_revenue_multiple_y1)
    if ev_rev:
        parts.append(f"{ev_rev} EV/Revenue")
    return ", ".join(parts)


def _company_headline(
    deal: DealRecord, *, role: str, sector_slug: str, comp_id: str,
    outlier: str, prior_acq: int,
) -> str:
    """Role-aware event text for a ``Companies/<party>.md`` bullet. The route
    handler wraps this with the dated/confirmed/provenance scaffold."""
    phrase = _multiples_phrase(deal)
    if role == "acquirer":
        other = _company_link(deal.target_company) or deal.target_company or "(unnamed target)"
        lead = f"Acquired {other}"
    else:  # target / vendor
        other = _company_link(deal.bidder_company) or deal.bidder_company or "(unnamed acquirer)"
        lead = f"Acquired by {other}"
    text = f"{lead} — {phrase}"
    if sector_slug and comp_id:
        text += f" → [[Sectors/{sector_slug}/Comps#{comp_id}]]"
    # Step 4 decoration — only on the acquirer voice (the prior-encounter party).
    if role == "acquirer" and prior_acq > 0:
        text += f" (acquirer in {prior_acq} prior tracked deal{'s' if prior_acq != 1 else ''})"
    if outlier:
        text += f" — {outlier}"
    return text


def _sector_block(
    deal: DealRecord, *, comp_id: str, fact_date: date, sector_slug: str, outlier: str,
) -> str:
    """The canonical ``### comp-<id>`` block (mirrors ``Sectors/telecoms/Comps.md``)
    for append under the ``## <YYYY>`` year heading via ``body_md``."""
    acquirer = _company_link(deal.bidder_company) or (deal.bidder_company or "n/a")
    structure = _first_sentence(deal.strategic_commentary) or _first_sentence(deal.deal_description) or "—"
    significance = outlier or "—"
    sources = (deal.source or deal.source_url or "Public announcement").strip()
    implied = _multiples_phrase(deal)
    lines = [
        f"### {comp_id}",
        f"- **Date:** {fact_date.isoformat()}",
        f"- **Target:** {deal.target_company or 'n/a'}",
        f"- **Acquirer:** {acquirer}",
        f"- **Consideration:** {_money(deal.enterprise_value_m, deal.currency)} EV",
        f"- **Implied:** {implied}",
        f"- **Subsector:** `{sector_slug or 'n/a'}`",
        f"- **Structure:** {structure}",
        f"- **Significance:** {significance}",
        f"- **Sources:** {sources}",
        f"- **Confidence:** {_DEFAULT_CONFIDENCE}",
    ]
    return "\n".join(lines)


# ─────────────────────────────────────────────────────────────────────────────
# Proposal write (skip-status idempotency + atomic) — mirrors
# emit_deliverable_proposal's write block, which is reused unchanged per
# SESSION-43 decision (single-target emitter stays as-is).
# ─────────────────────────────────────────────────────────────────────────────


def _write_proposal(path: Path, *, metadata: dict[str, Any], body: str) -> Optional[Path]:
    """Atomically write a ``deliverable-outcome`` proposal, skipping if an
    operator-triaged version already occupies the slot. Returns the path
    written, or ``None`` if skipped."""
    if path.is_file():
        try:
            existing = frontmatter.load(path)
            status = str(existing.metadata.get("status") or "").strip().lower()
            if status in SKIP_STATUSES:
                log.info("deal-capture: skipping %s — operator-triaged status=%r", path, status)
                return None
        except Exception as e:  # noqa: BLE001 — unreadable existing file → overwrite
            log.warning("deal-capture: failed to parse existing %s (%s) — overwriting", path, e)

    post = frontmatter.Post(body)
    post.metadata.update(metadata)
    path.parent.mkdir(parents=True, exist_ok=True)
    serialised = frontmatter.dumps(post) + "\n"
    tmp = path.with_suffix(path.suffix + ".tmp")
    tmp.write_text(serialised, encoding="utf-8")
    tmp.replace(path)
    log.info("deal-capture: wrote %s", path)
    return path


def _proposal_body(headline: str, *, target: str, section: str, now: datetime) -> str:
    ts = now.strftime("%Y-%m-%d %H:%M %Z").strip()
    return (
        f"# {headline}\n\n"
        f"## Captured conclusion\n\n"
        f"{headline}\n\n"
        f"On **Route**, this is appended (append-only — CLAUDE.md §3 rule 9) to "
        f"`{target}` under the `{section}` section.\n\n"
        f"*Emitted by the `dealtracker` vault-enrichment capture (#43) on {ts}. "
        f"Routes through the standard proposals lifecycle (#8 + #58): "
        f"Review and Route / Reject / Skip / Request revision.*\n"
    )


def _base_metadata(
    *, target: str, section: str, headline: str, deal: DealRecord,
    fact_date: date, provenance: str, run_id: str, artefact: str, sensitivity: str,
) -> dict[str, Any]:
    return {
        "type": "deliverable-outcome",
        "kind": "deliverable-outcome",
        "status": "pending-review",
        "date": fact_date.isoformat(),
        "skill": "dealtracker",
        "target": target,
        "section": section,
        "headline": headline,
        "fields": {
            "deal_id": deal.deal_id or build_deal_id(deal.announced_date, deal.target_company),
            "enterprise_value_m": deal.enterprise_value_m,
            "ev_ebitda_x": deal.reported_ebitda_multiple_y1,
            "ev_revenue_x": deal.reported_revenue_multiple_y1,
            "acquirer_type": deal.acquirer_type,
            "subsector_slug": deal.subsector_slug,
        },
        "provenance": provenance,
        "workspace_artefact": artefact,
        "run_id": run_id,
        "sensitivity": sensitivity,
        "tldr": headline,
    }


# ─────────────────────────────────────────────────────────────────────────────
# Public entrypoint
# ─────────────────────────────────────────────────────────────────────────────


def emit_deal_capture(
    deal: DealRecord,
    *,
    vault_root: Path,
    run_id: str = "",
    workbook_path: Optional[Path] = None,
    sheet_name: str = "Precedent transactions",
    sensitivity: str = "internal",
    now: Optional[datetime] = None,
) -> list[Path]:
    """Emit up to three operator-gated ``deliverable-outcome`` proposals for one
    appended deal (acquirer, target, sector). Returns the proposal paths
    actually written (skips: a deal with no ``announced_date`` — no stable
    identity; an unnamed party; a missing sector context; or an operator-triaged
    existing slot).

    Caller contract: best-effort. Pass a workbook to enable the step-4
    prior-encounter and step-8 outlier decorations; omit it and they're skipped.
    """
    now = now or datetime.now(timezone.utc)
    # An undated deal has NO stable capture identity: the comp-id, the proposal
    # filename, the `## <YYYY>` sector section, and the dated bullet all need the
    # deal's date — falling back to `now` would (a) drift across re-emits (the
    # same undated deal captured on different days yields different anchors) and
    # (b) fabricate a date the source never stated. Surface the gap, don't invent
    # (Iron Law): the row still lands in the tracker; vault enrichment waits until
    # the date is known. Capture is thus a pure function of (announced_date, target).
    if deal.announced_date is None:
        log.info(
            "deal-capture: skipping %r — no announced_date (no stable capture identity)",
            deal.target_company,
        )
        return []
    fact_date = deal.announced_date
    date_str = fact_date.isoformat()
    deal_slug = _slug(deal.target_company or "deal")
    disambig = _disambig(deal)  # collision-resistant slot/anchor suffix
    sector_slug = _sector_slug_for(deal)
    comp_id = _comp_id(deal, fact_date)
    provenance = f"runs:dealtracker.{run_id}" if run_id else "runs:dealtracker"
    artefact = str(workbook_path) if workbook_path else ""

    subsector_vals, prior_acq = _tracker_context(workbook_path, deal, sheet_name=sheet_name)
    outlier = _outlier_hint(deal, subsector_vals)

    proposal_dir = vault_root / PROPOSAL_DIR_REL
    written: list[Path] = []

    # ── steps 4 + 6: the two party pages (flat-bullet path, handler unchanged) ──
    party_specs = [
        ("acquirer", deal.bidder_company, deal.target_company),  # acquirer page; "other" = target
        ("target", deal.target_company, deal.bidder_company),    # target page;   "other" = acquirer
    ]
    for role, page_name, _other in party_specs:
        if not _safe_note_name(page_name):
            continue  # unnamed or un-addressable party — skip (emit N−1)
        headline = _company_headline(
            deal, role=role, sector_slug=sector_slug, comp_id=comp_id,
            outlier=outlier, prior_acq=prior_acq,
        )
        target_rel = f"Companies/{page_name.strip()}.md"
        meta = _base_metadata(
            target=target_rel, section="Transaction history", headline=headline,
            deal=deal, fact_date=fact_date, provenance=provenance, run_id=run_id,
            artefact=artefact, sensitivity=sensitivity,
        )
        path = proposal_dir / f"{date_str}-{deal_slug}-{disambig}-dealtracker-{role}.md"
        body = _proposal_body(headline, target=target_rel, section="Transaction history", now=now)
        if (p := _write_proposal(path, metadata=meta, body=body)) is not None:
            written.append(p)

    # ── step 7: the sector comp block (body_md path, additive handler branch) ──
    if sector_slug and _safe_note_name(sector_slug) and deal.target_company.strip():
        block = _sector_block(
            deal, comp_id=comp_id, fact_date=fact_date, sector_slug=sector_slug, outlier=outlier,
        )
        year = f"{fact_date.year:04d}"
        target_rel = f"Sectors/{sector_slug}/Comps.md"
        s_headline = (
            f"{deal.target_company} / {deal.bidder_company or 'n/a'} — "
            f"{sector_slug} precedent ({_multiples_phrase(deal)})"
        )
        meta = _base_metadata(
            target=target_rel, section=year, headline=s_headline,
            deal=deal, fact_date=fact_date, provenance=provenance, run_id=run_id,
            artefact=artefact, sensitivity=sensitivity,
        )
        meta["body_md"] = block  # consumed by _route_deliverable_outcome's body_md branch
        meta["comp_id"] = comp_id
        path = proposal_dir / f"{date_str}-{deal_slug}-{disambig}-dealtracker-sector.md"
        body = _proposal_body(s_headline, target=target_rel, section=year, now=now)
        if (p := _write_proposal(path, metadata=meta, body=body)) is not None:
            written.append(p)

    return written


__all__ = ["emit_deal_capture", "OUTLIER_MIN_ROWS"]
