"""CLI for deal-tracker routine."""

from __future__ import annotations

import logging
import sys
from pathlib import Path

import click

from routines.dealtracker.extract import extract_deal
from routines.dealtracker.snapshot import snapshot_tracker
from routines.dealtracker.workbook import (
    CANONICAL_SHEET_NAME, CANONICAL_WORKBOOK_PATH, append_deal,
)
from routines.shared import audit
from routines.shared.ollama_client import OllamaClient, OllamaError
from routines.shared.vault_writer import VaultPaths


DEFAULT_VAULT = Path("/mnt/x/OS AI Vault")
DEFAULT_AUDIT_DIR = Path(__file__).resolve().parent.parent.parent / "runs"
# Post 2026-06-01 retarget (WS2): the canonical workbook is the operator's
# precedent transactions tracker (absolute path OUTSIDE the vault). The legacy
# ``Projects/_Trackers/M&A Deals.xlsx`` is SUPERSEDED.
DEFAULT_WORKBOOK_PATH = CANONICAL_WORKBOOK_PATH


@click.group()
@click.option("--debug", is_flag=True)
def main(debug: bool) -> None:
    """Deal tracker — append M&A deal records to an Excel workbook."""
    logging.basicConfig(
        level=logging.DEBUG if debug else logging.INFO,
        format="%(asctime)s %(levelname)-7s %(name)s: %(message)s",
        datefmt="%H:%M:%S",
    )


@main.command("add")
@click.option("--url", default="", help="Source URL (recorded for provenance)")
@click.option("--text", default="", help="Pasted news article / press release text")
@click.option("--text-file", type=click.Path(exists=True, dir_okay=False, path_type=Path),
              help="Read source text from a file (alternative to --text)")
@click.option("--vault", type=click.Path(exists=True, file_okay=False, path_type=Path), default=DEFAULT_VAULT)
@click.option("--workbook",
              help="Path to tracker workbook (default: the canonical "
                   "Precedent_transactions_tracker.xlsx on the Corporate "
                   "Finance research drive)")
@click.option("--ollama-url", default="http://localhost:11434")
@click.option("--dry-run", is_flag=True, help="Show extracted record but don't write")
def add_cmd(
    url: str, text: str, text_file: Path | None,
    vault: Path, workbook: str | None,
    ollama_url: str, dry_run: bool,
) -> None:
    """Extract one deal from a URL or text and append to the tracker workbook.

    Examples:
        deal-tracker add --url https://... --text "<paste article body>"
        deal-tracker add --text-file ~/Downloads/press-release.txt --url https://...
    """
    if text_file:
        text = text_file.read_text(encoding="utf-8", errors="replace")
    if not text.strip():
        click.echo("Need --text or --text-file with article content.", err=True)
        sys.exit(2)

    paths = VaultPaths(vault)
    workbook_path = Path(workbook) if workbook else DEFAULT_WORKBOOK_PATH
    DEFAULT_AUDIT_DIR.mkdir(parents=True, exist_ok=True)

    client = OllamaClient(base_url=ollama_url)
    run_id = audit.new_run_id()

    try:
        deal = extract_deal(text=text, source_url=url, client=client)
    except OllamaError as e:
        click.echo(f"Extraction failed: {e}", err=True)
        sys.exit(1)
    deal.extracted_by_run_id = run_id

    # Show what we got
    click.echo(f"\nExtracted deal:")
    click.echo(f"  target:    {deal.target_company or '(none)'}")
    click.echo(f"  bidder:    {deal.bidder_company or '(none)'}")
    click.echo(f"  seller:    {deal.seller_company or '(none)'}")
    click.echo(f"  announced: {deal.announced_date or '(unknown)'}")
    click.echo(f"  EV:        {deal.enterprise_value_m} {deal.currency or ''}m"
               if deal.enterprise_value_m else "  EV:        (not stated)")
    click.echo(f"  multiples: rev={deal.reported_revenue_multiple_y1} "
               f"ebit={deal.reported_ebit_multiple_y1} "
               f"ebitda={deal.reported_ebitda_multiple_y1}")
    click.echo()

    if not deal.target_company:
        click.echo("WARNING: no target company extracted — likely not an M&A announcement.")

    if dry_run:
        click.echo("(dry-run — not appending)")
        return

    result = append_deal(workbook_path, deal)
    if result["status"] == "appended":
        click.echo(f"OK appended to row {result['row']} in {workbook_path}")
        # #43 — best-effort vault enrichment: emit operator-gated capture
        # proposals (acquirer / target / sector). A miss never fails the append.
        try:
            from routines.dealtracker.capture import emit_deal_capture

            proposals = emit_deal_capture(
                deal, vault_root=vault, run_id=run_id, workbook_path=workbook_path,
            )
            if proposals:
                click.echo(f"  + {len(proposals)} vault-capture proposal(s) pending review")
        except Exception as e:  # noqa: BLE001 — capture is best-effort
            logging.getLogger(__name__).warning("deal-capture emit failed: %s", e)
    elif result["status"] == "skipped_duplicate":
        click.echo(f"SKIPPED — duplicate of existing row {result['existing_row']}")

    audit.write_structured(
        actor={"type": "system", "id": "routine:dealtracker"},
        entity_type="vault_note",
        entity_id=str(workbook_path),
        action="ingest",
        routine="dealtracker", run_id=run_id, status="ok",
        audit_dir=DEFAULT_AUDIT_DIR,
        episodic_source=url,
        semantic_target=str(workbook_path),
        inputs={"url": url, "text_chars": len(text)},
        outputs={
            "workbook": str(workbook_path),
            "result": result["status"],
            "target_company": deal.target_company,
            "announced_date": deal.announced_date.isoformat() if deal.announced_date else None,
        },
    )


@main.command("init")
@click.option("--vault", type=click.Path(exists=True, file_okay=False, path_type=Path), default=DEFAULT_VAULT)
@click.option("--workbook",
              help="Path to tracker workbook (default: the canonical "
                   "Precedent_transactions_tracker.xlsx)")
def init_cmd(vault: Path, workbook: str | None) -> None:
    """Create an empty precedent tracker workbook with the lean 19-column
    header row + the ``Precedent transactions`` sheet, if it doesn't already
    exist."""
    paths = VaultPaths(vault)  # noqa: F841 — kept for symmetry with other commands
    workbook_path = Path(workbook) if workbook else DEFAULT_WORKBOOK_PATH
    if workbook_path.exists():
        click.echo(f"Workbook already exists: {workbook_path}")
        return

    from openpyxl import Workbook
    from routines.dealtracker.schema import COLUMNS

    workbook_path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]
    ws = wb.create_sheet(CANONICAL_SHEET_NAME)
    ws.append(COLUMNS)
    wb.save(str(workbook_path))
    click.echo(f"Created: {workbook_path}")
    click.echo(f"Sheet: {CANONICAL_SHEET_NAME}, {len(COLUMNS)} columns, 0 data rows")


@main.command("snapshot")
@click.option("--workbook",
              help="Path to the live tracker workbook (default: canonical "
                   "Precedent_transactions_tracker.xlsx on the corporate-"
                   "finance research drive)")
@click.option("--archive-dir",
              help="Where dated snapshots land (default: <live>/../Archive)")
def snapshot_cmd(workbook: str | None, archive_dir: str | None) -> None:
    """Snapshot the precedent transactions tracker to ./Archive/.

    Idempotent — safe to re-fire any number of times per day. Same-day
    re-fire is a no-op; cross-day re-fire with an unchanged live workbook
    is also a no-op (don't accumulate identical snapshots). Decision-3 of
    COMPS-REDESIGN-2026-06-01 (WS2).

    Examples:
        deal-tracker snapshot
        deal-tracker snapshot --workbook ./my-test-tracker.xlsx \\
            --archive-dir ./test-archive/
    """
    live_path = Path(workbook) if workbook else None     # None → canonical
    arch_dir = Path(archive_dir) if archive_dir else None
    DEFAULT_AUDIT_DIR.mkdir(parents=True, exist_ok=True)
    run_id = audit.new_run_id()

    result = snapshot_tracker(live_path=live_path, archive_dir=arch_dir)
    click.echo(f"status: {result.status}")
    click.echo(f"reason: {result.reason}")
    if result.snapshot_path:
        click.echo(f"snapshot: {result.snapshot_path}")

    audit.write_structured(
        actor={"type": "system", "id": "routine:dealtracker.snapshot"},
        entity_type="vault_note",
        entity_id=str(result.snapshot_path or live_path or CANONICAL_WORKBOOK_PATH),
        action="snapshot",
        routine="dealtracker.snapshot", run_id=run_id, status="ok",
        audit_dir=DEFAULT_AUDIT_DIR,
        inputs={
            "live_path": str(live_path) if live_path else str(CANONICAL_WORKBOOK_PATH),
            "archive_dir": str(arch_dir) if arch_dir else None,
        },
        outputs={
            "status": result.status,
            "snapshot_path": str(result.snapshot_path) if result.snapshot_path else None,
            "reason": result.reason,
        },
    )


if __name__ == "__main__":
    main()
